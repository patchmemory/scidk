"""
Blueprint for UI routes.

Handles:
- Homepage (index)
- Chat interface
- Map visualization
- Datasets (file list and detail views)
- Workbook viewer (Excel files)
- Plugins page
- Interpreters page
- RO-Crate viewer
- Settings page
- Scan form submission
"""
from flask import Blueprint, jsonify, request, render_template, current_app, redirect, url_for
from pathlib import Path
import os
import time
import hashlib

# Create blueprint
bp = Blueprint('ui', __name__)


# Helper to access app extensions
def _get_ext():
    """Get SciDK extensions from current Flask app."""
    return current_app.extensions['scidk']


# Routes
@bp.get('/')
def index():
    """Homepage with dataset and scan summaries."""
    ext = _get_ext()
    datasets = ext['graph'].list_datasets()
    # Build lightweight summaries for the landing page
    by_ext = {}
    interp_types = set()
    for d in datasets:
        by_ext[d.get('extension') or ''] = by_ext.get(d.get('extension') or '', 0) + 1
        for k in (d.get('interpretations') or {}).keys():
            interp_types.add(k)
    schema_summary = ext['graph'].schema_summary()
    telemetry = ext.get('telemetry', {})
    directories = list(ext.get('directories', {}).values())
    directories.sort(key=lambda d: d.get('last_scanned') or 0, reverse=True)
    scans = list(ext.get('scans', {}).values())
    scans.sort(key=lambda s: s.get('ended') or s.get('started') or 0, reverse=True)
    # Add SQLite-backed scan_count for landing summary
    scan_count = None
    try:
        from ...core import path_index_sqlite as pix
        conn = pix.connect()
        try:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(1) FROM scans")
            row = cur.fetchone()
            if row:
                scan_count = int(row[0])
        finally:
            try:
                conn.close()
            except Exception:
                pass
    except Exception:
        scan_count = None
    return render_template('index.html', datasets=datasets, by_ext=by_ext, schema_summary=schema_summary, telemetry=telemetry, directories=directories, scans=scans, scan_count=scan_count)


@bp.get('/chat')
def chat():
    """Chat interface page."""
    return render_template('chat.html')


@bp.get('/map')
def map_page():
    """Map visualization page."""
    ext = _get_ext()
    schema_summary = ext['graph'].schema_summary()
    return render_template('map.html', schema_summary=schema_summary)


@bp.get('/datasets')
def datasets():
    """List all datasets (files), optionally filtered by scan."""
    ext = _get_ext()
    all_datasets = ext['graph'].list_datasets()
    scan_id = (request.args.get('scan_id') or '').strip()
    selected_scan = None
    if scan_id:
        selected_scan = ext.get('scans', {}).get(scan_id)
    if selected_scan:
        checks = set(selected_scan.get('checksums') or [])
        datasets = [d for d in all_datasets if d.get('checksum') in checks]
    else:
        datasets = all_datasets
    directories = list(ext.get('directories', {}).values())
    directories.sort(key=lambda d: d.get('last_scanned') or 0, reverse=True)
    recent_scans = list(ext.get('scans', {}).values())
    recent_scans.sort(key=lambda s: s.get('ended') or s.get('started') or 0, reverse=True)
    # Show only the most recent N scans for dropdown
    N = 20
    recent_scans = recent_scans[:N]
    # files viewer mode: allow query param override, else env, else classic
    files_viewer = (request.args.get('files_viewer') or os.environ.get('SCIDK_FILES_VIEWER') or 'classic').strip()
    return render_template('datasets.html', datasets=datasets, directories=directories, recent_scans=recent_scans, selected_scan=selected_scan, files_viewer=files_viewer)


@bp.get('/datasets/<dataset_id>')
def dataset_detail(dataset_id):
    """Show detail page for a single dataset."""
    ext = _get_ext()
    item = ext['graph'].get_dataset(dataset_id)
    if not item:
        return render_template('dataset_detail.html', dataset=None), 404
    return render_template('dataset_detail.html', dataset=item)


@bp.get('/workbook/<dataset_id>')
def workbook_view(dataset_id):
    """Simple XLSX viewer: list sheets and preview first rows."""
    ext = _get_ext()
    item = ext['graph'].get_dataset(dataset_id)
    if not item:
        return render_template('workbook.html', dataset=None, error="Dataset not found"), 404
    file_path = Path(item['path'])
    if (item.get('extension') or '').lower() not in ['.xlsx', '.xlsm']:
        return render_template('workbook.html', dataset=item, error="Not an Excel workbook (.xlsx/.xlsm)"), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        sheetnames = wb.sheetnames
        previews = []
        max_rows = 20
        max_cols = 20
        for name in sheetnames:
            ws = wb[name]
            rows = []
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True), start=1):
                rows.append(list(row))
            previews.append({'name': name, 'rows': rows})
        wb.close()
        return render_template('workbook.html', dataset=item, sheetnames=sheetnames, previews=previews, error=None)
    except Exception as e:
        return render_template('workbook.html', dataset=item, sheetnames=[], previews=[], error=str(e)), 500


@bp.get('/plugins')
def plugins():
    """Plugins page - show counts from registry."""
    ext = _get_ext()
    reg = ext['registry']
    ext_count = len(reg.by_extension)
    interp_count = len(reg.by_id)
    return render_template('plugins.html', ext_count=ext_count, interp_count=interp_count)


@bp.get('/interpreters')
def interpreters():
    """List registry mappings and selection rules."""
    ext = _get_ext()
    reg = ext['registry']
    mappings = {ext: [getattr(i, 'id', 'unknown') for i in interps] for ext, interps in reg.by_extension.items()}
    rules = list(reg.rules.rules)
    return render_template('extensions.html', mappings=mappings, rules=rules)


@bp.get('/extensions')
def extensions_legacy():
    """Backward-compatible route - redirects to interpreters."""
    return redirect(url_for('ui.interpreters'))


@bp.get('/rocrate_view')
def rocrate_view():
    """Lightweight wrapper page to preview RO-Crate JSON-LD and prepare for embedding Crate-O."""
    prov_id = (request.args.get('provider_id') or 'local_fs').strip() or 'local_fs'
    root_id = (request.args.get('root_id') or '/').strip() or '/'
    sel_path = (request.args.get('path') or root_id).strip()
    try:
        from urllib.parse import urlencode
        qs = urlencode({'provider_id': prov_id, 'root_id': root_id, 'path': sel_path})
    except Exception:
        qs = f"provider_id={prov_id}&root_id={root_id}&path={sel_path}"
    metadata_url = '/api/rocrate?' + qs
    embed_mode = (os.environ.get('SCIDK_ROCRATE_EMBED') or 'json').strip()
    return render_template('rocrate_view.html', metadata_url=metadata_url, embed_mode=embed_mode, prov_id=prov_id, root_id=root_id, path=sel_path)


@bp.get('/settings')
def settings():
    """Basic settings from environment and current in-memory sizes."""
    ext = _get_ext()
    datasets = ext['graph'].list_datasets()
    reg = ext['registry']
    info = {
        'host': os.environ.get('SCIDK_HOST', '127.0.0.1'),
        'port': os.environ.get('SCIDK_PORT', '5000'),
        'debug': os.environ.get('SCIDK_DEBUG', '1'),
        'feature_file_index': os.environ.get('SCIDK_FEATURE_FILE_INDEX', ''),
        'hash_policy': os.environ.get('SCIDK_HASH_POLICY', 'auto'),
        'dataset_count': len(datasets),
        'interpreter_count': len(reg.by_id),
        'channel': os.environ.get('SCIDK_CHANNEL', 'stable'),
        'files_viewer': os.environ.get('SCIDK_FILES_VIEWER', ''),
        'providers': os.environ.get('SCIDK_PROVIDERS', 'local_fs,mounted_fs'),
    }
    # Provide interpreter mappings and rules, and plugin summary counts for the Settings page sections
    mappings = {ext: [getattr(i, 'id', 'unknown') for i in interps] for ext, interps in reg.by_extension.items()}
    rules = list(reg.rules.rules)
    ext_count = len(reg.by_extension)
    interp_count = len(reg.by_id)
    # Rclone mounts UI is now always enabled
    return render_template('settings.html', info=info, mappings=mappings, rules=rules, ext_count=ext_count, interp_count=interp_count, rclone_mounts_feature=True)


@bp.post('/scan')
def ui_scan():
    """Handle scan form submission from UI."""
    ext = _get_ext()
    fs = current_app.extensions['scidk']['fs']
    path = request.form.get('path') or os.getcwd()
    recursive = request.form.get('recursive') == 'on'
    # Pre-scan snapshot
    before = set(ds.get('checksum') for ds in ext['graph'].list_datasets())
    started = time.time()
    count = fs.scan_directory(Path(path), recursive=recursive)
    ended = time.time()
    duration = ended - started
    after = set(ds.get('checksum') for ds in ext['graph'].list_datasets())
    new_checksums = sorted(list(after - before))
    by_ext = {}
    ext_map = {ds.get('checksum'): ds.get('extension') or '' for ds in ext['graph'].list_datasets()}
    for ch in new_checksums:
        ext_val = ext_map.get(ch, '')
        by_ext[ext_val] = by_ext.get(ext_val, 0) + 1
    sid_src = f"{path}|{started}"
    scan_id = hashlib.sha1(sid_src.encode()).hexdigest()[:12]
    scan = {
        'id': scan_id,
        'path': str(path),
        'recursive': bool(recursive),
        'started': started,
        'ended': ended,
        'duration_sec': duration,
        'file_count': int(count),
        'checksums': new_checksums,
        'by_ext': by_ext,
        'source': getattr(fs, 'last_scan_source', 'python'),
        'errors': [],
    }
    scans = ext.setdefault('scans', {})
    scans[scan_id] = scan
    # Persist scan summary to SQLite (best-effort)
    try:
        from ...core import path_index_sqlite as pix
        from ...core import migrations as _migs
        import json as _json
        conn = pix.connect()
        try:
            _migs.migrate(conn)
            cur = conn.cursor()
            cur.execute(
                "INSERT OR REPLACE INTO scans(id, root, started, completed, status, extra_json) VALUES(?,?,?,?,?,?)",
                (
                    scan_id,
                    str(path),
                    float(started or 0.0),
                    float(ended or 0.0),
                    'completed',
                    _json.dumps({
                        'recursive': bool(recursive),
                        'duration_sec': duration,
                        'file_count': int(count),
                        'by_ext': by_ext,
                        'source': getattr(fs, 'last_scan_source', 'python'),
                        'checksums': new_checksums,
                        'committed': bool(scan.get('committed', False)),
                        'committed_at': scan.get('committed_at'),
                    })
                )
            )
            conn.commit()
        finally:
            try:
                conn.close()
            except Exception:
                pass
    except Exception:
        pass
    telem = ext.setdefault('telemetry', {})
    telem['last_scan'] = {
        'path': str(path),
        'recursive': bool(recursive),
        'scanned': int(count),
        'started': started,
        'ended': ended,
        'duration_sec': duration,
    }
    # Track scanned directories here as well
    dirs = ext.setdefault('directories', {})
    drec = dirs.setdefault(str(path), {'path': str(path), 'recursive': bool(recursive), 'scanned': 0, 'last_scanned': 0, 'scan_ids': []})
    drec.update({'recursive': bool(recursive), 'scanned': int(count), 'last_scanned': ended})
    drec.setdefault('scan_ids', []).append(scan_id)
    return redirect(url_for('ui.datasets', scan_id=scan_id))
