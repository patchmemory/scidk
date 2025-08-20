from pathlib import Path
from typing import Dict, List

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - environment without openpyxl
    load_workbook = None  # type: ignore


class XlsxInterpreter:
    id = "xlsx"
    name = "Excel Workbook Interpreter"
    version = "0.1.0"

    def __init__(self, max_bytes: int = 20 * 1024 * 1024):
        self.max_bytes = max_bytes

    def interpret(self, file_path: Path) -> Dict:
        if load_workbook is None:
            return {
                'status': 'error',
                'data': {
                    'error_type': 'XLSX_DEPENDENCY_MISSING',
                    'details': 'openpyxl is not installed'
                }
            }
        try:
            size = file_path.stat().st_size
            if size > self.max_bytes:
                return {
                    'status': 'error',
                    'data': {
                        'error_type': 'FILE_TOO_LARGE',
                        'max_bytes': self.max_bytes,
                        'size_bytes': size,
                        'details': f"Workbook exceeds limit of {self.max_bytes} bytes"
                    }
                }
            wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
            try:
                sheetnames: List[str] = wb.sheetnames
                meta = []
                for name in sheetnames:
                    try:
                        ws = wb[name]
                        rows = int(getattr(ws, 'max_row', 0) or 0)
                        cols = int(getattr(ws, 'max_column', 0) or 0)
                        meta.append({'name': name, 'rows': rows, 'cols': cols})
                    except Exception:
                        meta.append({'name': name, 'rows': 0, 'cols': 0})
                has_macros = file_path.suffix.lower() == '.xlsm'
                return {
                    'status': 'success',
                    'data': {
                        'type': 'xlsx',
                        'total_sheets': len(sheetnames),
                        'sheets': meta,
                        'has_macros': has_macros,
                    }
                }
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
        except Exception as e:
            return {
                'status': 'error',
                'data': {
                    'error_type': 'XLSX_INTERPRET_ERROR',
                    'details': str(e),
                }
            }
