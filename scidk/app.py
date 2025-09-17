from flask import Flask, Blueprint, jsonify, request, render_template, redirect, url_for
from pathlib import Path
import os
from typing import Optional
import time
import json

from .core.graph import InMemoryGraph
from .core.filesystem import FilesystemManager
from .core.registry import InterpreterRegistry
from .interpreters.python_code import PythonCodeInterpreter
from .interpreters.csv_interpreter import CsvInterpreter
from .interpreters.json_interpreter import JsonInterpreter
from .interpreters.yaml_interpreter import YamlInterpreter
from .interpreters.ipynb_interpreter import IpynbInterpreter
from .interpreters.txt_interpreter import TxtInterpreter
from .interpreters.xlsx_interpreter import XlsxInterpreter
from .core.pattern_matcher import Rule
from .core.providers import ProviderRegistry as FsProviderRegistry, LocalFSProvider, MountedFSProvider, RcloneProvider


def _apply_channel_defaults():
    """Apply channel-based defaults for feature flags when unset.
    Channels: stable (default), dev, beta.
    Explicit env values always win; we only set defaults if unset.
    Also soft-disable rclone provider by removing it from SCIDK_PROVIDERS if rclone binary is missing,
    unless SCIDK_FORCE_RCLONE is truthy. Only perform soft-disable when SCIDK_PROVIDERS was not explicitly set by user.
    """
    import shutil
    ch = (os.environ.get('SCIDK_CHANNEL') or 'stable').strip().lower()
    had_prov_env = 'SCIDK_PROVIDERS' in os.environ
    def setdefault_env(name: str, value: str):
        if os.environ.get(name) is None:
            os.environ[name] = value
    if ch in ('dev', 'beta'):
        # Providers default: include rclone
        if os.environ.get('SCIDK_PROVIDERS') is None:
            os.environ['SCIDK_PROVIDERS'] = 'local_fs,mounted_fs,rclone'
        # Mounts UI
        setdefault_env('SCIDK_RCLONE_MOUNTS', '1')
        # Files viewer mode
        setdefault_env('SCIDK_FILES_VIEWER', 'rocrate')
        # File index work in progress
        setdefault_env('SCIDK_FEATURE_FILE_INDEX', '1')
    # Soft rclone detection: remove if missing and not forced, but only when we set providers implicitly
    if not had_prov_env:
        prov_env = os.environ.get('SCIDK_PROVIDERS')
        if prov_env:
            prov_list = [p.strip() for p in prov_env.split(',') if p.strip()]
            if 'rclone' in prov_list and not shutil.which('rclone'):
                force = (os.environ.get('SCIDK_FORCE_RCLONE') or '').strip().lower() in ('1','true','yes','y','on')
                if not force:
                    prov_list = [p for p in prov_list if p != 'rclone']
                    os.environ['SCIDK_PROVIDERS'] = ','.join(prov_list)
    # Record effective channel for UI/debug
    os.environ.setdefault('SCIDK_CHANNEL', ch or 'stable')
    # Default: commit to graph should read from index unless explicitly disabled
    if os.environ.get('SCIDK_COMMIT_FROM_INDEX') is None:
        os.environ['SCIDK_COMMIT_FROM_INDEX'] = '1'


# --- Batched Neo4j commit helper ---
from typing import Iterable, Callable, Dict, Any, List, Optional, Tuple
import time as _time
import random as _random


def _chunked_list(seq: List[Dict[str, Any]], size: int) -> Iterable[List[Dict[str, Any]]]:
    size = max(1, int(size or 1))
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def _neo4j_progress_default(event: str, payload: Dict[str, Any]) -> None:
    try:
        print(f"[neo4j:{event}] {payload}")
    except Exception:
        pass


def commit_to_neo4j_batched(
    rows: List[Dict[str, Any]],
    folder_rows: List[Dict[str, Any]],
    scan: Dict[str, Any],
    neo4j_params: Tuple[str, str, str, str, Optional[str]],
    file_batch_size: int = 5000,
    folder_batch_size: int = 5000,
    max_retries: int = 2,
    on_progress: Callable[[str, Dict[str, Any]], None] = _neo4j_progress_default,
):
    """
    Batched Neo4j commit with per-batch retries and progress reporting.
    Expects rows to include keys: path, filename, extension, size_bytes, created, modified, mime_type, folder
    Expects folder_rows to include keys: path, name, parent
    """
    # Unpack params (support both 4 and 5 tuple forms)
    try:
        uri, user, pwd, database, auth_mode = neo4j_params
    except Exception:
        uri, user, pwd, database = neo4j_params  # type: ignore
        auth_mode = "basic"

    result: Dict[str, Any] = {
        "attempted": False,
        "written_files": 0,
        "written_folders": 0,
        "batches_total": 0,
        "batches_ok": 0,
        "batches_failed": 0,
        "errors": [],
        "error": None,
        "db_scan_exists": None,
        "db_files": None,
        "db_folders": None,
        "db_verified": False,
    }

    if not uri:
        return result

    can_basic = bool(user and pwd)
    if not (auth_mode == "none" or can_basic):
        return result

    # Backoff state
    try:
        from flask import current_app as _cap
        st = _cap.extensions["scidk"].setdefault("neo4j_state", {})
    except Exception:
        st = {"next_connect_after": 0}
    now = _time.time()
    next_after = float(st.get("next_connect_after") or 0)
    if next_after and now < next_after:
        wait_s = int(next_after - now)
        msg = f"neo4j connect backoff active; retry after {wait_s}s"
        result["errors"].append(msg)
        result["error"] = msg
        return result

    result["attempted"] = True

    # Derive host id for composite identity
    node_host = scan.get("host_id")

    from neo4j import GraphDatabase  # type: ignore  # noqa
    driver = None
    try:
        driver = GraphDatabase.driver(uri, auth=None if auth_mode == "none" else (user, pwd))
        with driver.session(database=database) as sess:
            # Emit non-secret connection info for diagnostics
            try:
                on_progress("neo4j_params", {
                    "uri": (uri.split("@")[-1] if uri else None),
                    "auth_mode": auth_mode,
                    "database": database or None,
                })
            except Exception:
                pass
            # Constraints (safe on 5.x; ignore errors otherwise)
            for cql in (
                "CREATE CONSTRAINT file_identity IF NOT EXISTS FOR (f:File) REQUIRE (f.path, f.host) IS UNIQUE",
                "CREATE CONSTRAINT folder_identity IF NOT EXISTS FOR (d:Folder) REQUIRE (d.path, d.host) IS UNIQUE",
            ):
                try:
                    sess.run(cql).consume()
                except Exception:
                    pass

            # Upsert Scan once
            scan_upsert = (
                "MERGE (s:Scan {id: $scan_id}) "
                "SET s.path = $scan_path, s.started = $scan_started, s.ended = $scan_ended, "
                "    s.provider_id = $scan_provider, s.host_type = $scan_host_type, s.host_id = $scan_host_id, "
                "    s.root_id = $scan_root_id, s.root_label = $scan_root_label, s.scan_source = $scan_source "
                "RETURN s.id AS scan_id"
            )
            sess.run(
                scan_upsert,
                scan_id=scan.get("id"),
                scan_path=scan.get("path"),
                scan_started=scan.get("started"),
                scan_ended=scan.get("ended"),
                scan_provider=scan.get("provider_id"),
                scan_host_type=scan.get("host_type"),
                scan_host_id=scan.get("host_id"),
                scan_root_id=scan.get("root_id"),
                scan_root_label=scan.get("root_label"),
                scan_source=scan.get("scan_source") or scan.get("source"),
            ).consume()

            # Cypher templates
            folders_upsert_cql = (
                "WITH $folders AS folders, $node_host AS node_host, $scan AS scan_id, $scan_provider AS scan_provider, $scan_host_type AS scan_host_type, $scan_host_id AS scan_host_id "
                "UNWIND folders AS folder "
                "MERGE (fo:Folder {path: folder.path, host: node_host}) "
                "  SET fo.name = folder.name, fo.provider_id = scan_provider, fo.host_type = scan_host_type, fo.host_id = scan_host_id "
                "WITH fo, scan_id "
                "OPTIONAL MATCH (s:Scan {id: scan_id}) "
                "MERGE (fo)-[:SCANNED_IN]->(s)"
            )

            folders_link_cql = (
                "WITH $folders AS folders, $node_host AS node_host "
                "UNWIND folders AS folder "
                "WITH folder, node_host "
                "WHERE folder.parent IS NOT NULL AND folder.parent <> '' AND folder.parent <> folder.path "
                "MERGE (parent:Folder {path: folder.parent, host: node_host}) "
                "MERGE (child:Folder {path: folder.path, host: node_host}) "
                "MERGE (parent)-[:CONTAINS]->(child)"
            )

            # Optional safety net to compute folder from file path when missing (feature-flagged)
            _files_fallback = (os.environ.get('SCIDK_FILES_COMPUTE_FOLDER_FALLBACK') or '').strip().lower() in ('1','true','yes','y','on')
            if _files_fallback:
                files_cql = (
                    "WITH $rows AS rows, $node_host AS node_host, $scan AS scan_id, $scan_provider AS scan_provider, $scan_host_type AS scan_host_type, $scan_host_id AS scan_host_id "
                    "UNWIND rows AS r "
                    "MERGE (f:File {path: r.path, host: node_host}) "
                    "  SET f.filename = r.filename, f.extension = r.extension, f.size_bytes = r.size_bytes, "
                    "      f.created = r.created, f.modified = r.modified, f.mime_type = r.mime_type, "
                    "      f.provider_id = scan_provider, f.host_type = scan_host_type, f.host_id = scan_host_id "
                    "WITH r, f, scan_id, node_host, CASE WHEN r.folder IS NOT NULL AND r.folder <> '' THEN r.folder ELSE substring(r.path, 0, size(r.path) - size(last(split(r.path, '/'))) - 1) END AS folder_path "
                    "OPTIONAL MATCH (s:Scan {id: scan_id}) "
                    "MERGE (f)-[:SCANNED_IN]->(s) "
                    "WITH folder_path, f, node_host WHERE folder_path IS NOT NULL AND folder_path <> '' "
                    "MERGE (fo:Folder {path: folder_path, host: node_host}) "
                    "MERGE (fo)-[:CONTAINS]->(f)"
                )
            else:
                files_cql = (
                    "WITH $rows AS rows, $node_host AS node_host, $scan AS scan_id, $scan_provider AS scan_provider, $scan_host_type AS scan_host_type, $scan_host_id AS scan_host_id "
                    "UNWIND rows AS r "
                    "MERGE (f:File {path: r.path, host: node_host}) "
                    "  SET f.filename = r.filename, f.extension = r.extension, f.size_bytes = r.size_bytes, "
                    "      f.created = r.created, f.modified = r.modified, f.mime_type = r.mime_type, "
                    "      f.provider_id = scan_provider, f.host_type = scan_host_type, f.host_id = scan_host_id "
                    "WITH r, f, scan_id, node_host "
                    "OPTIONAL MATCH (s:Scan {id: scan_id}) "
                    "MERGE (f)-[:SCANNED_IN]->(s) "
                    "WITH r, f, node_host "
                    "WHERE r.folder IS NOT NULL AND r.folder <> '' "
                    "MERGE (fo:Folder {path: r.folder, host: node_host}) "
                    "MERGE (fo)-[:CONTAINS]->(f)"
                )

            def _run_with_retry(cql: str, params: Dict[str, Any], kind: str, index: int):
                attempt = 0
                while True:
                    try:
                        sess.run(cql, **params).consume()
                        return True, None
                    except Exception as ex:
                        attempt += 1
                        if attempt > max_retries:
                            return False, str(ex)
                        sleep_s = min(1.5 * attempt + _random.random(), 5.0)
                        on_progress("retry", {"batch_kind": kind, "batch_index": index, "attempt": attempt, "sleep_s": round(sleep_s, 2), "error": str(ex)[:200]})
                        _time.sleep(sleep_s)

            folder_batches = list(_chunked_list(folder_rows, folder_batch_size))
            file_batches = list(_chunked_list(rows, file_batch_size))
            result["batches_total"] = len(folder_batches) + len(file_batches)

            # Debug first 10 examples of folder/file mappings
            try:
                sample_files = [{"path": r.get("path"), "folder": r.get("folder")} for r in (rows[:10] if rows else [])]
                sample_folders = [{"path": r.get("path"), "parent": r.get("parent")} for r in (folder_rows[:10] if folder_rows else [])]
                on_progress("debug_rows", {"sample_files": sample_files, "sample_folders": sample_folders})
            except Exception:
                pass

            on_progress("start", {
                "scan_id": scan.get("id"),
                "folders": len(folder_rows),
                "files": len(rows),
                "folder_batches": len(folder_batches),
                "file_batches": len(file_batches),
                "batch_sizes": {"folders": folder_batch_size, "files": file_batch_size},
            })

            common = {
                "scan": scan.get("id"),
                "node_host": node_host,
                "scan_provider": scan.get("provider_id"),
                "scan_host_type": scan.get("host_type"),
                "scan_host_id": scan.get("host_id"),
            }

            # Folders first: upsert nodes and attach to scan
            for i, batch in enumerate(folder_batches, start=1):
                t0 = _time.time()
                ok, err = _run_with_retry(folders_upsert_cql, {"folders": batch, **common}, kind="folders_upsert", index=i)
                dt = _time.time() - t0
                if not ok:
                    result["batches_failed"] += 1
                    result["errors"].append(f"folders upsert batch {i}: {err}")
                    on_progress("batch_error", {"batch_kind": "folders_upsert", "batch_index": i, "items": len(batch), "error": err})
                    continue
                # Link parent-child hierarchy in a second step
                t1 = _time.time()
                ok2, err2 = _run_with_retry(folders_link_cql, {"folders": batch, **common}, kind="folders_link", index=i)
                dt2 = _time.time() - t1
                if ok2:
                    result["batches_ok"] += 2  # count both successful sub-steps
                    result["written_folders"] += len(batch)
                    on_progress("batch_done", {"batch_kind": "folders_upsert", "batch_index": i, "items": len(batch), "sec": round(dt, 3)})
                    on_progress("batch_done", {"batch_kind": "folders_link", "batch_index": i, "items": len(batch), "sec": round(dt2, 3)})
                else:
                    result["batches_ok"] += 1  # upsert ok
                    result["batches_failed"] += 1
                    result["errors"].append(f"folders link batch {i}: {err2}")
                    on_progress("batch_done", {"batch_kind": "folders_upsert", "batch_index": i, "items": len(batch), "sec": round(dt, 3)})
                    on_progress("batch_error", {"batch_kind": "folders_link", "batch_index": i, "items": len(batch), "error": err2})

            # Files
            for i, batch in enumerate(file_batches, start=1):
                t0 = _time.time()
                ok, err = _run_with_retry(files_cql, {"rows": batch, **common}, kind="files", index=i)
                dt = _time.time() - t0
                if ok:
                    result["batches_ok"] += 1
                    result["written_files"] += len(batch)
                    on_progress("batch_done", {"batch_kind": "files", "batch_index": i, "items": len(batch), "sec": round(dt, 3)})
                else:
                    result["batches_failed"] += 1
                    result["errors"].append(f"files batch {i}: {err}")
                    on_progress("batch_error", {"batch_kind": "files", "batch_index": i, "items": len(batch), "error": err})

            # Verify
            verify_q = (
                "OPTIONAL MATCH (s:Scan {id: $scan_id}) "
                "WITH s "
                "OPTIONAL MATCH (s)<-[:SCANNED_IN]-(f:File) "
                "WITH s, count(DISTINCT f) AS files_cnt "
                "OPTIONAL MATCH (s)<-[:SCANNED_IN]-(fo:Folder) "
                "RETURN coalesce(s IS NOT NULL, false) AS scan_exists, files_cnt AS files_cnt, count(DISTINCT fo) AS folders_cnt"
            )
            vrec = sess.run(verify_q, scan_id=scan.get("id")).single()
            if vrec:
                scan_exists = bool(vrec.get("scan_exists"))
                files_cnt = int(vrec.get("files_cnt") or 0)
                folders_cnt = int(vrec.get("folders_cnt") or 0)
                result["db_scan_exists"] = scan_exists
                result["db_files"] = files_cnt
                result["db_folders"] = folders_cnt
                result["db_verified"] = bool(scan_exists and (files_cnt > 0 or folders_cnt > 0))

            on_progress("done", {
                "scan_id": scan.get("id"),
                "written_files": result["written_files"],
                "written_folders": result["written_folders"],
                "batches_ok": result["batches_ok"],
                "batches_failed": result["batches_failed"],
                "db_verified": result["db_verified"],
                "db_files": result["db_files"],
                "db_folders": result["db_folders"],
            })

    except Exception as e:
        msg = str(e)
        result["error"] = msg
        result["errors"].append(msg)
        # auth backoff
        try:
            emsg = msg.lower()
            if ("unauthorized" in emsg) or ("authentication" in emsg):
                prev = float(st.get("next_connect_after") or 0)
                base = 20.0
                delay = base
                now2 = _time.time()
                if prev and now2 < prev:
                    rem = prev - now2
                    delay = min(max(base * 2, rem * 2), 120.0)
                st["next_connect_after"] = now2 + delay
                st["last_error"] = msg
        except Exception:
            pass
    finally:
        try:
            if driver is not None:
                driver.close()
        except Exception:
            pass

    return result


def create_app():
    # Apply channel-based defaults before reading env-driven config
    _apply_channel_defaults()
    app = Flask(__name__, template_folder="ui/templates", static_folder="ui/static")
    # Feature: selective dry-run UI flag (dev default)
    try:
        ch = (os.environ.get('SCIDK_CHANNEL') or 'stable').strip().lower()
        flag_env = (os.environ.get('SCIDK_FEATURE_SELECTIVE_DRYRUN') or '').strip().lower()
        flag = flag_env in ('1','true','yes','y','on')
        if flag_env == '' and ch == 'dev':
            flag = True
        app.config['feature.selectiveDryRun'] = bool(flag)
    except Exception:
        app.config['feature.selectiveDryRun'] = False

    # Auto-migrate SQLite schema on boot (best effort)
    try:
        from .core import migrations as _migs
        _migs.migrate()
    except Exception as _e:
        # Defer reporting to /api/health if needed via app.extensions
        pass

    # Core singletons (select backend)
    backend = (os.environ.get('SCIDK_GRAPH_BACKEND') or 'memory').strip().lower()
    if backend == 'neo4j':
        try:
            uri, user, pwd, database, auth_mode = _get_neo4j_params()
            from .core.neo4j_graph import Neo4jGraph
            auth = None if auth_mode == 'none' else (user, pwd)
            graph = Neo4jGraph(uri=uri, auth=auth, database=database)
        except Exception:
            # Fallback to in-memory if neo4j params invalid
            from .core.graph import InMemoryGraph as _IMG
            graph = _IMG()
    else:
        graph = InMemoryGraph()
    registry = InterpreterRegistry()
    # Load persisted interpreter toggle settings (optional)
    try:
        from .core.settings import InterpreterSettings
        settings = InterpreterSettings(os.environ.get('SCIDK_SETTINGS_DB', 'scidk_settings.db'))
        enabled = settings.load_enabled_interpreters()
        if enabled:
            registry.enabled_interpreters = set(enabled)
    except Exception:
        settings = None

    # Register interpreters
    py_interp = PythonCodeInterpreter()
    csv_interp = CsvInterpreter()
    json_interp = JsonInterpreter()
    yaml_interp = YamlInterpreter()
    ipynb_interp = IpynbInterpreter()
    txt_interp = TxtInterpreter()
    xlsx_interp = XlsxInterpreter()
    registry.register_extension(".py", py_interp)
    registry.register_extension(".csv", csv_interp)
    registry.register_extension(".json", json_interp)
    registry.register_extension(".yml", yaml_interp)
    registry.register_extension(".yaml", yaml_interp)
    registry.register_extension(".ipynb", ipynb_interp)
    registry.register_extension(".txt", txt_interp)
    registry.register_extension(".xlsx", xlsx_interp)
    registry.register_extension(".xlsm", xlsx_interp)
    # Register simple rules to prefer interpreters for extensions
    registry.register_rule(Rule(id="rule.py.default", interpreter_id=py_interp.id, pattern="*.py", priority=10, conditions={"ext": ".py"}))
    registry.register_rule(Rule(id="rule.csv.default", interpreter_id=csv_interp.id, pattern="*.csv", priority=10, conditions={"ext": ".csv"}))
    registry.register_rule(Rule(id="rule.json.default", interpreter_id=json_interp.id, pattern="*.json", priority=10, conditions={"ext": ".json"}))
    registry.register_rule(Rule(id="rule.yml.default", interpreter_id=yaml_interp.id, pattern="*.yml", priority=10, conditions={"ext": ".yml"}))
    registry.register_rule(Rule(id="rule.yaml.default", interpreter_id=yaml_interp.id, pattern="*.yaml", priority=10, conditions={"ext": ".yaml"}))
    registry.register_rule(Rule(id="rule.ipynb.default", interpreter_id=ipynb_interp.id, pattern="*.ipynb", priority=10, conditions={"ext": ".ipynb"}))
    registry.register_rule(Rule(id="rule.txt.default", interpreter_id=txt_interp.id, pattern="*.txt", priority=10, conditions={"ext": ".txt"}))
    registry.register_rule(Rule(id="rule.xlsx.default", interpreter_id=xlsx_interp.id, pattern="*.xlsx", priority=10, conditions={"ext": ".xlsx"}))
    registry.register_rule(Rule(id="rule.xlsm.default", interpreter_id=xlsx_interp.id, pattern="*.xlsm", priority=10, conditions={"ext": ".xlsm"}))

    # Compute effective interpreter enablement (CLI envs > global settings > defaults)
    testing_env = bool(os.environ.get('PYTEST_CURRENT_TEST')) or bool(os.environ.get('SCIDK_DISABLE_SETTINGS'))
    try:
        from .core.settings import InterpreterSettings
        settings = None if testing_env else InterpreterSettings(db_path=str(Path(os.getcwd()) / 'scidk_settings.db'))
    except Exception:
        settings = None
    # Defaults from interpreter attributes (fallback True)
    all_ids = list(registry.by_id.keys())
    default_enabled_ids = set([iid for iid in all_ids if bool(getattr(registry.by_id[iid], 'default_enabled', True))])
    # CLI overrides via env
    # CLI overrides via env (case-insensitive); ignore unknown ids to avoid surprises
    en_raw = [s.strip() for s in (os.environ.get('SCIDK_ENABLE_INTERPRETERS') or '').split(',') if s.strip()]
    dis_raw = [s.strip() for s in (os.environ.get('SCIDK_DISABLE_INTERPRETERS') or '').split(',') if s.strip()]
    # Normalize to lowercase (registry ids are lowercase)
    en_list = [s.lower() for s in en_raw]
    dis_list = [s.lower() for s in dis_raw]
    source = 'default'
    if en_list or dis_list:
        known_ids = set(all_ids)
        unknown_en = [x for x in en_list if x not in known_ids]
        unknown_dis = [x for x in dis_list if x not in known_ids]
        # Start from defaults; remove DISABLE; add ENABLE; ENABLE wins on conflicts
        enabled_set = set(default_enabled_ids)
        for d in dis_list:
            if d in known_ids:
                enabled_set.discard(d)
        for e in en_list:
            if e in known_ids:
                enabled_set.add(e)
        source = 'cli'
        # Do NOT persist CLI-derived sets to settings to avoid masking user intentions
        try:
            _ist = app.extensions.setdefault('scidk', {}).setdefault('interpreters', {})
            _ist['unknown_env'] = {'enable': unknown_en, 'disable': unknown_dis}
        except Exception:
            pass
    else:
        # Load global saved set if any
        loaded = set()
        try:
            if settings:
                loaded = set(settings.load_enabled_interpreters())
        except Exception:
            loaded = set()
        if loaded:
            enabled_set = set(loaded)
            source = 'global'
        else:
            enabled_set = set(default_enabled_ids)
            source = 'default'
    # Store effective on app
    _interp_state = {'effective_enabled': enabled_set, 'source': source}
    # Apply effective enabled set to registry for selection logic
    try:
        registry.enabled_interpreters = set(enabled_set)
    except Exception:
        pass

    fs = FilesystemManager(graph=graph, registry=registry)

    # Initialize filesystem providers (Phase 0)
    prov_enabled = [p.strip() for p in (os.environ.get('SCIDK_PROVIDERS', 'local_fs,mounted_fs').split(',')) if p.strip()]
    # If rclone mounts feature is enabled, ensure rclone provider is also enabled for listremotes validation
    _ff_rc = (os.environ.get('SCIDK_RCLONE_MOUNTS') or os.environ.get('SCIDK_FEATURE_RCLONE_MOUNTS') or '').strip().lower() in ('1','true','yes','y','on')
    if _ff_rc and 'rclone' not in prov_enabled:
        prov_enabled.append('rclone')
    fs_providers = FsProviderRegistry(enabled=prov_enabled)
    p_local = LocalFSProvider(); p_local.initialize(app, {})
    p_mounted = MountedFSProvider(); p_mounted.initialize(app, {})
    p_rclone = RcloneProvider(); p_rclone.initialize(app, {})
    fs_providers.register(p_local)
    fs_providers.register(p_mounted)
    fs_providers.register(p_rclone)

    # Store refs on app for easy access
    app.extensions = getattr(app, 'extensions', {})
    app.extensions['scidk'] = {
        'graph': graph,
        'registry': registry,
        'fs': fs,
        'providers': fs_providers,
        'interpreters': _interp_state,
        # in-session registries
        'scans': {},  # scan_id -> scan session dict
        'directories': {},  # path -> aggregate info incl. scan_ids
        'telemetry': {},
        'tasks': {},  # task_id -> task dict (background jobs like scans)
        'scan_fs': {},  # per-scan filesystem index cache for snapshot navigation
        'neo4j_config': {
            'uri': None,
            'user': None,
            'password': None,
            'database': None,
        },
        'neo4j_state': {
            'connected': False,
            'last_error': None,
        },
        # rclone mounts runtime registry (feature-flagged API will use this)
        'rclone_mounts': {},  # id/name -> { id, remote, subpath, path, read_only, started_at, pid, log_file }
        'settings': settings,
    }

    # Hydrate telemetry.last_scan from SQLite settings on startup (best-effort)
    try:
        from .core import path_index_sqlite as pix
        from .core import migrations as _migs
        import json as _json
        conn = pix.connect()
        try:
            _migs.migrate(conn)
            cur = conn.cursor()
            row = cur.execute("SELECT value FROM settings WHERE key = ?", ("telemetry.last_scan",)).fetchone()
            if row and row[0]:
                try:
                    last_scan = _json.loads(row[0])
                    app.extensions.setdefault('scidk', {}).setdefault('telemetry', {})['last_scan'] = last_scan
                except Exception:
                    pass
        finally:
            try:
                conn.close()
            except Exception:
                pass
    except Exception:
        pass

    # Hydrate rclone interpretation settings (suggest mount threshold and batch size)
    try:
        def _env_int(name: str, dflt: int) -> int:
            try:
                v = os.environ.get(name)
                return int(v) if v is not None and v != '' else dflt
            except Exception:
                return dflt
        suggest_dflt = _env_int('SCIDK_RCLONE_INTERPRET_SUGGEST_MOUNT', 400)
        max_batch_dflt = _env_int('SCIDK_RCLONE_INTERPRET_MAX_FILES', 1000)
        max_batch_dflt = min(max(100, max_batch_dflt), 2000)
        from .core import path_index_sqlite as pix
        from .core import migrations as _migs
        conn = pix.connect()
        try:
            _migs.migrate(conn)
            cur = conn.cursor()
            def _get_setting_int(key: str, dflt: int) -> int:
                row = cur.execute("SELECT value FROM settings WHERE key= ?", (key,)).fetchone()
                if row and row[0] not in (None, ''):
                    try:
                        return int(row[0])
                    except Exception:
                        return dflt
                return dflt
            suggest_mount_threshold = _get_setting_int('rclone.interpret.suggest_mount_threshold', suggest_dflt)
            max_files_per_batch = _get_setting_int('rclone.interpret.max_files_per_batch', max_batch_dflt)
            max_files_per_batch = min(max(100, int(max_files_per_batch)), 2000)
            app.config['rclone.interpret.suggest_mount_threshold'] = int(suggest_mount_threshold)
            app.config['rclone.interpret.max_files_per_batch'] = int(max_files_per_batch)
        finally:
            try:
                conn.close()
            except Exception:
                pass
    except Exception:
        # Defaults if hydration fails
        app.config.setdefault('rclone.interpret.suggest_mount_threshold', 400)
        app.config.setdefault('rclone.interpret.max_files_per_batch', 1000)

    # Feature flag for rclone mount manager (define before first use)
    def _feature_rclone_mounts() -> bool:
        val = (os.environ.get('SCIDK_RCLONE_MOUNTS') or os.environ.get('SCIDK_FEATURE_RCLONE_MOUNTS') or '').strip().lower()
        return val in ('1', 'true', 'yes', 'y', 'on')

    # Rehydrate rclone mounts metadata from SQLite on startup (no process attached)
    if _feature_rclone_mounts():
        try:
            from .core import path_index_sqlite as pix
            from .core import migrations as _migs
            import json as _json
            conn = pix.connect()
            try:
                _migs.migrate(conn)
                cur = conn.cursor()
                cur.execute("SELECT id, provider, root, created, status, extra_json FROM provider_mounts WHERE provider='rclone'")
                rows = cur.fetchall() or []
                rm = app.extensions['scidk'].setdefault('rclone_mounts', {})
                for (mid, provider, remote, created, status_persisted, extra) in rows:
                    try:
                        extra_obj = _json.loads(extra) if extra else {}
                    except Exception:
                        extra_obj = {}
                    rm[mid] = {
                        'id': mid,
                        'name': mid,
                        'remote': remote,
                        'subpath': extra_obj.get('subpath'),
                        'path': extra_obj.get('path'),
                        'read_only': extra_obj.get('read_only'),
                        'started_at': created,
                        'process': None,
                        'pid': None,
                        'log_file': extra_obj.get('log_file'),
                    }
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass

    # API routes
    api = Blueprint('api', __name__, url_prefix='/api')

    # Import SQLite layer for selections/annotations lazily to avoid circular deps
    from .core import annotations_sqlite as ann_db


    # Helper to read Neo4j configuration, preferring in-app settings over environment
    # Returns tuple: (uri, user, password, database, auth_mode)
    # auth_mode: 'basic' (username+password) or 'none' (no authentication)
    def _get_neo4j_params():
        cfg = app.extensions['scidk'].get('neo4j_config', {})
        uri = cfg.get('uri') or os.environ.get('NEO4J_URI') or os.environ.get('BOLT_URI')
        user = cfg.get('user') or os.environ.get('NEO4J_USER') or os.environ.get('NEO4J_USERNAME')
        pwd = cfg.get('password') or os.environ.get('NEO4J_PASSWORD')
        database = cfg.get('database') or os.environ.get('SCIDK_NEO4J_DATABASE') or None
        # Parse NEO4J_AUTH env var if provided (formats: "user/pass" or "none")
        neo4j_auth = (os.environ.get('NEO4J_AUTH') or '').strip()
        if neo4j_auth:
            if neo4j_auth.lower() == 'none':
                user = user or None
                pwd = pwd or None
                auth_mode = 'none'
            else:
                try:
                    # Expecting user/password
                    parts = neo4j_auth.split('/')
                    if len(parts) >= 2 and not (user and pwd):
                        user = user or parts[0]
                        pwd = pwd or '/'.join(parts[1:])
                except Exception:
                    pass
        # If user/password still missing, try to parse from URI (bolt://user:pass@host:port)
        auth_mode = 'basic'
        try:
            if uri and (not user or not pwd):
                from urllib.parse import urlparse, unquote
                parsed = urlparse(uri)
                if parsed.username and parsed.password:
                    user = user or unquote(parsed.username)
                    pwd = pwd or unquote(parsed.password)
        except Exception:
            pass
        # Determine auth mode: none only when explicitly set via NEO4J_AUTH=none
        if (os.environ.get('NEO4J_AUTH') or '').strip().lower() == 'none':
            auth_mode = 'none'
        else:
            auth_mode = 'basic'
        return uri, user, pwd, database, auth_mode

    # Build rows for commit: files (rows) and standalone folders (folder_rows)
    def build_commit_rows(scan, ds_map):
        """Legacy builder from in-memory datasets."""
        try:
            from .services.commit_service import CommitService
            return CommitService().build_rows_legacy_from_datasets(scan, ds_map)
        except Exception:
            # Fallback to empty on unexpected import/runtime error
            return [], []

    # Execute Neo4j commit using simplified, idempotent Cypher
    def commit_to_neo4j(rows, folder_rows, scan, neo4j_params):
        # Support 4-tuple (backward compat) and 5-tuple with auth_mode
        try:
            uri, user, pwd, database, auth_mode = neo4j_params
        except Exception:
            uri, user, pwd, database = neo4j_params
            auth_mode = 'basic'
        result = {'attempted': False, 'written_files': 0, 'written_folders': 0, 'error': None}
        if not uri:
            return result
        # Decide if we can attempt a connection
        can_basic = bool(user and pwd)
        can_connect = (auth_mode == 'none') or can_basic
        if not can_connect:
            return result
        # Backoff on recent auth failures to avoid rate limiting
        st = app.extensions['scidk'].setdefault('neo4j_state', {})
        import time as _t
        now = _t.time()
        next_after = float(st.get('next_connect_after') or 0)
        if next_after and now < next_after:
            result['error'] = f"neo4j connect backoff active; retry after {int(next_after-now)}s"
            return result
        result['attempted'] = True
        try:
            from .services.neo4j_client import Neo4jClient
            client = Neo4jClient(uri, user, pwd, database, auth_mode).connect()
            try:
                client.ensure_constraints()
                wres = client.write_scan(rows, folder_rows, scan)
                result['written_files'] = wres.get('written_files', 0)
                result['written_folders'] = wres.get('written_folders', 0)
                vres = client.verify(scan.get('id'))
                result.update(vres)
            finally:
                client.close()
        except Exception as e:
            msg = str(e)
            result['error'] = msg
            # On auth-related errors, set a backoff to avoid rate limiting
            try:
                emsg = msg.lower()
                if ('unauthorized' in emsg) or ('authentication' in emsg):
                    # Exponential-ish backoff min 20s
                    prev = float(st.get('next_connect_after') or 0)
                    base = 20.0
                    delay = base
                    if prev and now < prev:
                        # increase delay up to 120s
                        rem = prev - now
                        delay = min(max(base*2, rem*2), 120.0)
                    st['next_connect_after'] = now + delay
                    st['last_error'] = msg
            except Exception:
                pass
        return result

    # Build or fetch per-scan filesystem index for snapshot navigation
    def _get_or_build_scan_index(scan_id: str):
        cache = app.extensions['scidk'].setdefault('scan_fs', {})
        if scan_id in cache:
            return cache[scan_id]
        scans = app.extensions['scidk'].get('scans', {})
        s = scans.get(scan_id)
        if not s:
            return None
        checksums = s.get('checksums') or []
        ds_map = app.extensions['scidk']['graph'].datasets  # checksum -> dataset

        from .core.path_utils import parse_remote_path, parent_remote_path
        from pathlib import Path as _P

        folder_info = {}
        children_files = {}

        def ensure_complete_parent_chain(path_str: str):
            """Ensure all parent folders exist in folder_info for any given path"""
            if not path_str or path_str in folder_info:
                return

            info = parse_remote_path(path_str)
            if info.get('is_remote'):
                parent = parent_remote_path(path_str)
                name = (info.get('parts')[-1] if info.get('parts') else info.get('remote_name') or path_str)
            else:
                try:
                    p = _P(path_str)
                    parent = str(p.parent)
                    name = p.name or path_str
                except Exception:
                    parent = ''
                    name = path_str

            folder_info[path_str] = {
                'path': path_str,
                'name': name,
                'parent': parent,
            }

            if parent and parent != path_str:
                ensure_complete_parent_chain(parent)

        # Seed scan base path (stable roots even on empty scans)
        try:
            base_path = s.get('path') or ''
            if base_path:
                ensure_complete_parent_chain(base_path)
        except Exception:
            pass

        # Process files and ensure their parent chains exist
        for ch in checksums:
            d = ds_map.get(ch)
            if not d:
                continue
            file_path = d.get('path')
            if not file_path:
                continue

            info = parse_remote_path(file_path)
            if info.get('is_remote'):
                parent = parent_remote_path(file_path)
                filename = (info.get('parts')[-1] if info.get('parts') else info.get('remote_name') or file_path)
            else:
                try:
                    p = _P(file_path)
                    parent = str(p.parent)
                    filename = p.name or file_path
                except Exception:
                    parent = ''
                    filename = file_path

            file_entry = {
                'id': d.get('id'),
                'path': file_path,
                'filename': d.get('filename') or filename,
                'extension': d.get('extension'),
                'size_bytes': int(d.get('size_bytes') or 0),
                'modified': float(d.get('modified') or 0),
                'mime_type': d.get('mime_type'),
                'checksum': d.get('checksum'),
            }
            children_files.setdefault(parent, []).append(file_entry)

            if parent:
                ensure_complete_parent_chain(parent)

        # Process explicitly recorded folders
        for f in (s.get('folders') or []):
            path = f.get('path')
            if path:
                ensure_complete_parent_chain(path)

        # Build children_folders map
        children_folders = {}
        for fpath, info in folder_info.items():
            par = info.get('parent')
            if par and par in folder_info:
                children_folders.setdefault(par, []).append(fpath)

        # Find actual roots
        roots = sorted([fp for fp, info in folder_info.items()
                        if not info.get('parent') or info.get('parent') not in folder_info])

        # Prefer scan base as visible root and drop its ancestors
        try:
            base_path = s.get('path') or ''
            if base_path and base_path in folder_info:
                if base_path not in roots:
                    roots.append(base_path)
                def _is_ancestor(candidate: str, child: str) -> bool:
                    if not candidate or candidate == child:
                        return False
                    cinf = parse_remote_path(candidate)
                    chinf = parse_remote_path(child)
                    if chinf.get('is_remote') and cinf.get('is_remote'):
                        return child.startswith(candidate.rstrip('/') + '/')
                    try:
                        return str(_P(child)).startswith(str(_P(candidate)) + '/')
                    except Exception:
                        return False
                roots = [r for r in roots if not _is_ancestor(r, base_path) or r == base_path]
                roots = sorted(list(dict.fromkeys(roots)))
        except Exception:
            pass

        # Sort children deterministically
        for k in list(children_folders.keys()):
            children_folders[k].sort(key=lambda p: folder_info.get(p, {}).get('name', '').lower())
        for k in list(children_files.keys()):
            children_files[k].sort(key=lambda f: (f.get('filename') or '').lower())

        idx = {
            'folder_info': folder_info,
            'children_folders': children_folders,
            'children_files': children_files,
            'roots': roots,
        }
        cache[scan_id] = idx
        return idx

    # Feature flags for file indexing
    _ff_index = (os.environ.get('SCIDK_FEATURE_FILE_INDEX') or '').strip().lower() in ('1','true','yes','y','on')

    @api.post('/scan/dry-run')
    def api_scan_dry_run():
        from fnmatch import fnmatch
        data = request.get_json(force=True, silent=True) or {}
        path = data.get('path') or os.getcwd()
        include = data.get('include') or []
        exclude = data.get('exclude') or []
        max_depth = data.get('max_depth')
        use_ignore = bool(data.get('use_ignore', True))
        base = Path(path)
        if not base.exists() or not base.is_dir():
            return jsonify({'status':'error','error':'invalid path'}), 400
        # Load .scidkignore patterns (gitignore-like globs, one per line) at root
        ignore_patterns = []
        if use_ignore:
            ign = base / '.scidkignore'
            try:
                if ign.exists():
                    for line in ign.read_text(encoding='utf-8').splitlines():
                        s = line.strip()
                        if not s or s.startswith('#'):
                            continue
                        ignore_patterns.append(s)
            except Exception:
                pass
        files = []
        total_bytes = 0
        base_parts = len(base.resolve().parts)
        try:
            for p in base.rglob('*'):
                try:
                    if p.is_file():
                        rel = p.resolve().relative_to(base.resolve()).as_posix()
                        # skip control file itself
                        if rel == '.scidkignore':
                            continue
                        # depth filter
                        if isinstance(max_depth, int):
                            depth = len(p.resolve().parts) - base_parts
                            if depth > max_depth:
                                continue
                        # ignore patterns
                        ignored = any(fnmatch(rel, pat) for pat in ignore_patterns)
                        if ignored:
                            continue
                        # include/exclude
                        if include:
                            if not any(fnmatch(rel, pat) for pat in include):
                                continue
                        if exclude and any(fnmatch(rel, pat) for pat in exclude):
                            continue
                        files.append(rel)
                        try:
                            total_bytes += int(p.stat().st_size)
                        except Exception:
                            pass
                except Exception:
                    continue
        except Exception:
            files = []
            total_bytes = 0
        files.sort()
        return jsonify({
            'status': 'ok',
            'root': str(base.resolve()),
            'total_files': len(files),
            'total_bytes': int(total_bytes),
            'files': files
        })

    @api.post('/scan')
    def api_scan():
        data = request.get_json(force=True, silent=True) or {}
        try:
            from .services.metrics import record_event_time
            record_event_time(app, 'scan_started_times')
        except Exception:
            pass
        provider_id = (data.get('provider_id') or 'local_fs').strip() or 'local_fs'
        root_id = (data.get('root_id') or '/').strip() or '/'
        path = data.get('path') or (root_id if provider_id != 'local_fs' else os.getcwd())
        recursive = bool(data.get('recursive', True))
        fast_list = bool(data.get('fast_list', False))
        # Prefer fast_list by default for recursive rclone scans if client omitted it
        _client_specified_fast_list = ('fast_list' in data)
        # Delegate to ScansService (refactor): preserve payload and behavior
        try:
            from .services.scans_service import ScansService
            svc = ScansService(app)
            result = svc.run_scan({
                'provider_id': provider_id,
                'root_id': root_id,
                'path': path,
                'recursive': recursive,
                'fast_list': fast_list,
            })
            if isinstance(result, dict) and result.get('status') == 'ok':
                return jsonify(result), 200
            # Error path with optional http_status
            if isinstance(result, dict) and result.get('status') == 'error':
                code = int(result.get('http_status', 400))
                payload = {'status': 'error', 'error': result.get('error')}
                return jsonify(payload), code
        except Exception:
            # On service failure, fallback to legacy in-place implementation below
            pass
        try:
            import time, hashlib, json
            from .core import path_index_sqlite as pix
            # Pre-scan snapshot of checksums
            before = set(ds.get('checksum') for ds in app.extensions['scidk']['graph'].list_datasets())
            started = time.time()
            # Precompute scan id early for SQLite tagging
            sid_src = f"{path}|{started}"
            scan_id = hashlib.sha1(sid_src.encode()).hexdigest()[:12]
            count = 0
            ingested = 0
            folders = []
            files_skipped = 0
            files_hashed = 0
            if provider_id in ('local_fs', 'mounted_fs'):
                # Local/Mounted: enumerate filesystem and ingest into SQLite index
                base = Path(path)
                # Build list of files and folders
                items_files = []
                items_dirs = set()
                # Preserve source detection semantics (ncdu > gdu > python)
                try:
                    probe_ncdu = fs._list_files_with_ncdu(base, recursive=recursive)  # type: ignore
                    if probe_ncdu:
                        fs.last_scan_source = 'ncdu'
                    else:
                        probe_gdu = fs._list_files_with_gdu(base, recursive=recursive)  # type: ignore
                        if probe_gdu:
                            fs.last_scan_source = 'gdu'
                        else:
                            fs.last_scan_source = 'python'
                except Exception:
                    fs.last_scan_source = 'python'
                try:
                    if recursive:
                        for p in base.rglob('*'):
                            try:
                                if p.is_dir():
                                    items_dirs.add(p)
                                else:
                                    items_files.append(p)
                                    # ensure parent chain exists in dirs set
                                    parent = p.parent
                                    while parent and parent != parent.parent and str(parent).startswith(str(base)):
                                        items_dirs.add(parent)
                                        if parent == base:
                                            break
                                        parent = parent.parent
                            except Exception:
                                continue
                        # include base itself as a folder
                        items_dirs.add(base)
                    else:
                        for p in base.iterdir():
                            try:
                                if p.is_dir():
                                    items_dirs.add(p)
                                else:
                                    items_files.append(p)
                            except Exception:
                                continue
                        items_dirs.add(base)
                except Exception:
                    items_files = []
                    items_dirs = set()
                # Map to rows
                from .core import path_index_sqlite as pix
                rows = []
                files_skipped = 0
                files_hashed = 0
                hash_policy = (os.environ.get('SCIDK_HASH_POLICY') or 'auto').strip().lower()
                def _row_from_local(pth: Path, typ: str) -> tuple:
                    nonlocal files_skipped, files_hashed
                    full = str(pth.resolve())
                    parent = str(pth.parent.resolve()) if pth != pth.parent else ''
                    name = pth.name or full
                    depth = 0 if pth == base else max(0, len(str(pth.resolve()).rstrip('/').split('/')) - len(str(base.resolve()).rstrip('/').split('/')))
                    size = 0
                    mtime = None
                    ext = ''
                    mime = None
                    etag = None
                    ahash = None
                    if typ == 'file':
                        try:
                            st = pth.stat()
                            size = int(st.st_size)
                            mtime = float(st.st_mtime)
                        except Exception:
                            size = 0
                            mtime = None
                        ext = pth.suffix.lower()
                        # Skip logic: reuse previous hash if unchanged (size + mtime)
                        try:
                            prev = pix.get_latest_file_meta(full)
                        except Exception:
                            prev = None
                        if prev is not None and prev[0] == size and prev[1] == mtime and (prev[2] or '') != '':
                            ahash = prev[2]
                            files_skipped += 1
                        else:
                            # Compute content hash with policy
                            try:
                                ahash = pix.compute_content_hash(full, hash_policy)
                            except Exception:
                                ahash = None
                            files_hashed += 1
                    remote = f"local:{os.uname().nodename}" if provider_id == 'local_fs' else f"mounted:{root_id}"
                    return (full, parent, name, depth, typ, size, mtime, ext, mime, etag, ahash, remote, scan_id, None)
                # Insert folder rows first for structure consistency
                for d in sorted(items_dirs, key=lambda x: str(x)):
                    rows.append(_row_from_local(d, 'folder'))
                # Then files
                for fpath in items_files:
                    rows.append(_row_from_local(fpath, 'file'))
                ingested = pix.batch_insert_files(rows)
                # Also create in-memory datasets (keep legacy behavior)
                count = 0
                for fpath in items_files:
                    try:
                        ds = fs.create_dataset_node(fpath)
                        app.extensions['scidk']['graph'].upsert_dataset(ds)
                        interps = registry.select_for_dataset(ds)
                        for interp in interps:
                            try:
                                result = interp.interpret(fpath)
                                payload = {
                                    'status': result.get('status', 'success'),
                                    'data': result.get('data', result),
                                    'interpreter_version': getattr(interp, 'version', '0.0.1'),
                                }
                                app.extensions['scidk']['graph'].add_interpretation(ds['checksum'], interp.id, payload)
                                # Persist interpretation metadata into SQLite files table for this path
                                try:
                                    from .core import path_index_sqlite as pix
                                    conn_i = pix.connect(); pix.init_db(conn_i)
                                    try:
                                        cur_i = conn_i.cursor()
                                        import json as _json
                                        # Determine the canonical key used in the index for this file path
                                        key_path = None
                                        try:
                                            # For rclone/remote scans, the index stores canonical remote paths like "remote:rel/path"
                                            # Prefer dataset-provided original path if present
                                            key_path = ds.get('path') or None
                                        except Exception:
                                            key_path = None
                                        if not key_path:
                                            # Fallback to absolute local path for local filesystem scans
                                            key_path = str(fpath.resolve())
                                        cur_i.execute(
                                            "UPDATE files SET interpreted_as = ?, interpretation_json = ? WHERE path = ? AND type = 'file' AND scan_id = ?",
                                            (interp.id, _json.dumps(payload.get('data')), key_path, scan_id)
                                        )
                                        conn_i.commit()
                                    finally:
                                        conn_i.close()
                                except Exception:
                                    pass
                            except Exception as e:
                                err_payload = {
                                    'status': 'error',
                                    'data': {'error': str(e)},
                                    'interpreter_version': getattr(interp, 'version', '0.0.1'),
                                }
                                app.extensions['scidk']['graph'].add_interpretation(ds['checksum'], interp.id, err_payload)
                                try:
                                    from .core import path_index_sqlite as pix
                                    conn_i = pix.connect(); pix.init_db(conn_i)
                                    try:
                                        cur_i = conn_i.cursor()
                                        import json as _json
                                        # Determine canonical key as above
                                        key_path = None
                                        try:
                                            key_path = ds.get('path') or None
                                        except Exception:
                                            key_path = None
                                        if not key_path:
                                            key_path = str(fpath.resolve())
                                        cur_i.execute(
                                            "UPDATE files SET interpreted_as = ?, interpretation_json = ? WHERE path = ? AND type = 'file' AND scan_id = ?",
                                            (interp.id, _json.dumps(err_payload.get('data')), key_path, scan_id)
                                        )
                                        conn_i.commit()
                                    finally:
                                        conn_i.close()
                                except Exception:
                                    pass
                        count += 1
                    except Exception:
                        continue
                # Collect folders metadata for scan record
                folders = []
                for d in items_dirs:
                    try:
                        parent = str(d.parent.resolve()) if d != d.parent else ''
                        folders.append({'path': str(d.resolve()), 'name': d.name, 'parent': parent, 'parent_name': Path(parent).name if parent else ''})
                    except Exception:
                        continue
            elif provider_id == 'rclone':
                # Use rclone lsjson to enumerate remote files; ingest into SQLite and create lightweight datasets.
                provs = app.extensions['scidk'].get('providers')
                prov = provs.get('rclone') if provs else None
                if not prov:
                    raise RuntimeError('rclone provider not available')
                # Normalize relative Rclone paths to full remote targets using root_id
                try:
                    from .core.path_utils import parse_remote_path, join_remote_path
                    info = parse_remote_path(path or '')
                    is_remote = bool(info.get('is_remote'))
                except Exception:
                    is_remote = False
                if not is_remote:
                    # path is relative or empty; compose with root_id
                    from .core.path_utils import join_remote_path as _join
                    path = _join(root_id, (path or '').lstrip('/'))
                # If recursive rclone and client did not specify fast_list, enable it for robustness
                if provider_id == 'rclone' and recursive and not _client_specified_fast_list:
                    fast_list = True

                # ALWAYS RECORD THE SCAN BASE FOLDER for rclone scans
                # Ensures the target path appears as a folder node, preventing flattened view
                try:
                    from .core.path_utils import parse_remote_path, parent_remote_path
                    from .core import path_index_sqlite as pix

                    info_t = parse_remote_path(path)
                    base_name = (info_t.get('parts')[-1] if info_t.get('parts') else info_t.get('remote_name') or path)
                    base_parent = parent_remote_path(path)

                    # Create synthetic base folder SQLite row and initialize rows with it
                    base_item = {"Name": base_name, "Path": "", "IsDir": True, "Size": 0}
                    rows = [pix.map_rclone_item_to_row(base_item, path, scan_id)]

                    # Compute parent display name for folders list
                    try:
                        info_par = parse_remote_path(base_parent) if base_parent else {}
                        if info_par.get('is_remote'):
                            parts = info_par.get('parts') or []
                            parent_name = (info_par.get('remote_name') or '') if not parts else parts[-1]
                        else:
                            from pathlib import Path as _P
                            parent_name = _P(base_parent).name if base_parent else ''
                    except Exception:
                        parent_name = ''

                    folders.append({
                        'path': path,
                        'name': base_name,
                        'parent': base_parent,
                        'parent_name': parent_name,
                    })
                except Exception:
                    # Non-fatal; initialize to empty rows if base insertion fails
                    rows = []

                try:
                    # In testing mode, allow metadata-only scans for rclone to avoid external binary dependency
                    if app.config.get('TESTING') and not recursive:
                        items = []
                    else:
                        items = prov.list_files(path, recursive=recursive, fast_list=fast_list)  # type: ignore[attr-defined]
                except Exception as ee:
                    return jsonify({"status": "error", "error": str(ee)}), 400
                # Map to SQLite rows (files and folders); only files will have size > 0 typically.
                # rows is initialized above with the base folder row; continue accumulating
                # Collect folders set for both non-recursive and recursive scans
                seen_folders = set()
                def _add_folder(full_path: str, name: str, parent: str):
                    if full_path in seen_folders:
                        return
                    seen_folders.add(full_path)
                    try:
                        from .core.path_utils import parse_remote_path
                        info_par = parse_remote_path(parent)
                        if info_par.get('is_remote'):
                            parts = info_par.get('parts') or []
                            parent_name = (info_par.get('remote_name') or '') if not parts else parts[-1]
                        else:
                            parent_name = Path(parent).name if parent else ''
                    except Exception:
                        parent_name = ''
                    folders.append({'path': full_path, 'name': name, 'parent': parent, 'parent_name': parent_name})
                for it in (items or []):
                    try:
                        # Track folders for both modes
                        if it.get('IsDir'):
                            rel = it.get('Path') or it.get('Name') or ''
                            if rel:
                                from .core.path_utils import join_remote_path, parent_remote_path
                                full = join_remote_path(path, rel)
                                parent = parent_remote_path(full)
                                leaf = rel.rsplit('/',1)[-1] if isinstance(rel, str) and '/' in rel else rel
                                # record folder regardless of recursive flag so empty dirs are preserved
                                _add_folder(full, leaf, parent)
                            # Still insert folder rows into SQLite for depth/structure awareness
                            rows.append(pix.map_rclone_item_to_row(it, path, scan_id))
                            continue
                        # File entry
                        rows.append(pix.map_rclone_item_to_row(it, path, scan_id))
                                # Synthesize intermediate folders from file rel paths (even if not recursive)
                        rel = it.get('Path') or it.get('Name') or ''
                        if rel:
                            from .core.path_utils import join_remote_path, parent_remote_path
                            parts = [p for p in (rel.split('/') if isinstance(rel, str) else []) if p]
                            cur_rel = ''
                            for i in range(len(parts)-1):  # exclude the file itself
                                cur_rel = parts[i] if i == 0 else (cur_rel + '/' + parts[i])
                                full = join_remote_path(path, cur_rel)
                                parent = parent_remote_path(full)
                                _add_folder(full, parts[i], parent)
                                # ensure a folder row exists in SQLite, even if rclone didn't emit it
                                try:
                                    folder_item = {"Name": parts[i], "Path": cur_rel, "IsDir": True, "Size": 0}
                                    rows.append(pix.map_rclone_item_to_row(folder_item, path, scan_id))
                                except Exception:
                                    pass
                        # Create datasets only when backend is not neo4j (to reduce RAM)
                        backend = (os.environ.get('SCIDK_GRAPH_BACKEND') or 'memory').strip().lower()
                        if backend != 'neo4j':
                            size = int(it.get('Size') or 0)
                            from .core.path_utils import join_remote_path
                            full = join_remote_path(path, rel)
                            ds = fs.create_dataset_remote(full, size_bytes=size, modified_ts=0.0, mime=None)
                            app.extensions['scidk']['graph'].upsert_dataset(ds)
                            count += 1
                    except Exception:
                        continue
                # Batch insert into SQLite (10k/txn) always (remove feature flag gating for rclone)
                try:
                    # Deduplicate rows by (path,type) before insert
                    try:
                        seen = set(); uniq = []
                        for r in rows:
                            key = (r[0], r[4])  # path, type
                            if key in seen:
                                continue
                            seen.add(key); uniq.append(r)
                        rows = uniq
                    except Exception:
                        pass
                    ingested = pix.batch_insert_files(rows, batch_size=10000)
                    # Minimal change detection to populate file_history
                    try:
                        _chg = pix.apply_basic_change_history(scan_id, path)
                        app.extensions['scidk'].setdefault('telemetry', {})['last_change_counts'] = _chg
                    except Exception as __e:
                        app.extensions['scidk'].setdefault('telemetry', {})['last_change_error'] = str(__e)
                except Exception as _e:
                    # Surface as non-fatal for now; continue app flow but record error
                    app.extensions['scidk'].setdefault('telemetry', {})['last_sqlite_error'] = str(_e)
            else:
                return jsonify({"status": "error", "error": f"provider {provider_id} not supported for scan"}), 400
            ended = time.time()
            duration = ended - started
            after = set(ds.get('checksum') for ds in app.extensions['scidk']['graph'].list_datasets())
            new_checksums = sorted(list(after - before))
            # Build by_ext
            by_ext = {}
            backend = (os.environ.get('SCIDK_GRAPH_BACKEND') or 'memory').strip().lower()
            if backend == 'neo4j':
                # derive from SQLite for the current scan
                try:
                    from .core import path_index_sqlite as pix
                    conn = pix.connect(); pix.init_db(conn)
                    cur = conn.cursor()
                    cur.execute("SELECT file_extension FROM files WHERE scan_id = ? AND type='file'", (scan_id,))
                    for (ext,) in cur.fetchall():
                        ext = ext or ''
                        by_ext[ext] = by_ext.get(ext, 0) + 1
                    conn.close()
                except Exception:
                    by_ext = {}
            else:
                ext_map = {}
                for ds in app.extensions['scidk']['graph'].list_datasets():
                    ext_map[ds.get('checksum')] = ds.get('extension') or ''
                for ch in new_checksums:
                    ext = ext_map.get(ch, '')
                    by_ext[ext] = by_ext.get(ext, 0) + 1
            # For non-recursive local scans, include immediate subfolders for later commit/merge
            if provider_id in ('local_fs', 'mounted_fs'):
                try:
                    if not recursive:
                        base = Path(path)
                        for child in base.iterdir():
                            if child.is_dir():
                                parent = str(child.parent)
                                folders.append({
                                    'path': str(child.resolve()),
                                    'name': child.name,
                                    'parent': parent,
                                    'parent_name': Path(parent).name if parent else '',
                                })
                except Exception:
                    pass
            # Provider metadata for scan/session records
            provs = app.extensions['scidk'].get('providers')
            prov = provs.get(provider_id) if provs else None
            root_label = None
            try:
                if prov:
                    root_label = Path(root_id).name or str(root_id)
            except Exception:
                root_label = None
            # Host/provider tagging
            host_type = provider_id
            host_id = None
            try:
                if provider_id == 'rclone':
                    host_id = f"rclone:{(root_id or '').rstrip(':')}"
                elif provider_id == 'local_fs':
                    import socket as _sock
                    host_id = f"local:{_sock.gethostname()}"
                elif provider_id == 'mounted_fs':
                    host_id = f"mounted:{root_id}"
            except Exception:
                host_id = f"{provider_id}:{root_id}" if root_id else provider_id
            scan = {
                'id': scan_id,
                'path': str(path),
                'recursive': bool(recursive),
                'started': started,
                'ended': ended,
                'duration_sec': duration,
                'file_count': int(count),
                'folder_count': len(folders),
                'checksums': new_checksums,
                'folders': folders,
                'by_ext': by_ext,
                'source': getattr(fs, 'last_scan_source', 'python') if provider_id in ('local_fs','mounted_fs') else f"provider:{provider_id}",
                'errors': [],
                'committed': False,
                'committed_at': None,
                'provider_id': provider_id,
                'host_type': host_type,
                'host_id': host_id,
                'root_id': root_id,
                'root_label': root_label,
                'scan_source': f"provider:{provider_id}",
                'ingested_rows': int(ingested),
                'config_json': {
                    'interpreters': {
                        'effective_enabled': sorted(list(app.extensions['scidk'].get('interpreters', {}).get('effective_enabled', []))),
                        'source': app.extensions['scidk'].get('interpreters', {}).get('source', 'default'),
                    }
                },
            }
            scans = app.extensions['scidk'].setdefault('scans', {})
            scans[scan_id] = scan
            # Persist scan summary to SQLite (best-effort)
            try:
                from .core import path_index_sqlite as pix
                from .core import migrations as _migs
                conn = pix.connect()
                import json as _json
                try:
                    _migs.migrate(conn)
                    cur = conn.cursor()
                    cur.execute(
                        "INSERT OR REPLACE INTO scans(id, root, started, completed, status, extra_json) VALUES(?,?,?,?,?,?)",
                        (
                            scan_id,
                            str(path),
                            float(started or 0.0),
                            float(ended or 0.0),
                            'completed',
                            _json.dumps({
                                'recursive': bool(recursive),
                                'duration_sec': duration,
                                'file_count': int(count),
                                'by_ext': by_ext,
                                'source': scan.get('source'),
                                'checksums': new_checksums,
                                'committed': False,
                                'committed_at': None,
                                'provider_id': provider_id,
                                'root_id': root_id,
                                'host_type': host_type,
                                'host_id': host_id,
                                'root_label': root_label,
                            })
                        )
                    )
                    conn.commit()
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            except Exception:
                pass
            # Clear cached fs index for this scan so next request rebuilds with fresh data
            try:
                app.extensions['scidk'].setdefault('scan_fs', {}).pop(scan_id, None)
            except Exception:
                pass
            # Save telemetry on app
            telem = app.extensions['scidk'].setdefault('telemetry', {})
            telem['last_scan'] = {
                'path': str(path),
                'recursive': bool(recursive),
                'scanned': int(count),
                'started': started,
                'ended': ended,
                'duration_sec': duration,
                'source': getattr(fs, 'last_scan_source', 'python') if provider_id in ('local_fs','mounted_fs') else f"provider:{provider_id}",
                'provider_id': provider_id,
                'root_id': root_id,
                'files_skipped': int(files_skipped),
                'files_hashed': int(files_hashed),
            }
            # Persist telemetry.last_scan to SQLite (best-effort)
            try:
                from .core import path_index_sqlite as pix
                from .core import migrations as _migs
                import json as _json
                conn = pix.connect()
                try:
                    _migs.migrate(conn)
                    cur = conn.cursor()
                    cur.execute(
                        "INSERT OR REPLACE INTO settings(key, value) VALUES(?, ?)",
                        ("telemetry.last_scan", _json.dumps(telem.get('last_scan') or {}))
                    )
                    conn.commit()
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            except Exception:
                pass
            # Track scanned directories (in-session registry)
            dirs = app.extensions['scidk'].setdefault('directories', {})
            drec = dirs.setdefault(str(path), {
                'path': str(path),
                'recursive': bool(recursive),
                'scanned': 0,
                'last_scanned': 0,
                'scan_ids': [],
                'source': getattr(fs, 'last_scan_source', 'python') if provider_id in ('local_fs','mounted_fs') else f"provider:{provider_id}",
                'provider_id': provider_id,
                'root_id': root_id,
                'root_label': root_label,
            })
            drec.update({
                'recursive': bool(recursive),
                'scanned': int(count),
                'last_scanned': ended,
                'source': getattr(fs, 'last_scan_source', 'python') if provider_id in ('local_fs','mounted_fs') else f"provider:{provider_id}",
                'provider_id': provider_id,
                'root_id': root_id,
                'root_label': root_label,
            })
            drec.setdefault('scan_ids', []).append(scan_id)
            return jsonify({"status": "ok", "scan_id": scan_id, "scanned": count, "folder_count": len(folders), "ingested_rows": int(ingested), "duration_sec": duration, "path": str(path), "recursive": bool(recursive), "provider_id": provider_id}), 200
        except Exception as e:
            return jsonify({"status": "error", "error": str(e)}), 400

    @api.post('/tasks')
    def api_tasks_create():
        """Create a background task. Supports type=scan and type=commit."""
        data = request.get_json(force=True, silent=True) or {}
        ttype = (data.get('type') or 'scan').strip().lower()
        import time, hashlib, threading
        started = time.time()

        # Enforce max concurrent tasks (running)
        try:
            max_tasks = int(os.environ.get('SCIDK_MAX_BG_TASKS', '2'))
        except Exception:
            max_tasks = 2
        running = sum(1 for t in app.extensions['scidk'].get('tasks', {}).values() if t.get('status') == 'running')
        if running >= max_tasks:
            return jsonify({'error': 'too many tasks running', 'code': 'max_tasks', 'max': max_tasks}), 429

        if ttype == 'scan':
            provider_id = (data.get('provider_id') or 'local_fs').strip() or 'local_fs'
            root_id = (data.get('root_id') or ('/' if provider_id != 'rclone' else 'remote:')).strip()
            path = data.get('path') or (root_id if provider_id != 'local_fs' else os.getcwd())
            recursive = bool(data.get('recursive', True))
            # Normalize rclone path to full remote target if needed
            if provider_id == 'rclone':
                try:
                    from .core.path_utils import parse_remote_path, join_remote_path
                    info = parse_remote_path(path or '')
                    if not bool(info.get('is_remote')):
                        path = join_remote_path(root_id, (path or '').lstrip('/'))
                except Exception:
                    pass
            tid_src = f"scan|{provider_id}|{path}|{started}"
            task_id = hashlib.sha1(tid_src.encode()).hexdigest()[:12]
            task = {
                'id': task_id,
                'type': 'scan',
                'status': 'running',
                'path': str(path),
                'recursive': bool(recursive),
                'started': started,
                'ended': None,
                'total': 0,
                'processed': 0,
                'progress': 0.0,
                'scan_id': None,
                'error': None,
                'cancel_requested': False,
            }
            app.extensions['scidk'].setdefault('tasks', {})[task_id] = task

            def _worker():
                try:
                    import hashlib as _h
                    from .core import path_index_sqlite as pix
                    scans = app.extensions['scidk'].setdefault('scans', {})
                    # Pre snapshot for in-memory dataset delta
                    before = set(ds.get('checksum') for ds in app.extensions['scidk']['graph'].list_datasets())
                    started_ts = time.time()
                    scan_id = _h.sha1(f"{path}|{started_ts}".encode()).hexdigest()[:12]
                    file_count = 0
                    folder_count = 0
                    ingested = 0
                    folders_meta = []

                    if provider_id in ('local_fs', 'mounted_fs'):
                        base = Path(path)
                        # Estimate total: Python traversal
                        files_list = [p for p in fs._iter_files_python(base, recursive=recursive)]
                        task['total'] = len(files_list)
                        # Build rows like api_scan
                        items_files = []
                        items_dirs = set()
                        if recursive:
                            for p in base.rglob('*'):
                                if task.get('cancel_requested'):
                                    task['status'] = 'canceled'; task['ended'] = time.time(); return
                                try:
                                    if p.is_dir():
                                        items_dirs.add(p)
                                    else:
                                        items_files.append(p)
                                        parent = p.parent
                                        while parent and parent != parent.parent and str(parent).startswith(str(base)):
                                            items_dirs.add(parent)
                                            if parent == base:
                                                break
                                            parent = parent.parent
                                except Exception:
                                    continue
                            items_dirs.add(base)
                        else:
                            try:
                                for p in base.iterdir():
                                    if p.is_dir(): items_dirs.add(p)
                                    else: items_files.append(p)
                            except Exception:
                                pass
                            items_dirs.add(base)
                        # Map to rows
                        def _row_from_local(pth: Path, typ: str) -> tuple:
                            full = str(pth.resolve())
                            parent = str(pth.parent.resolve()) if pth != pth.parent else ''
                            name = pth.name or full
                            depth = 0 if pth == base else max(0, len(str(pth.resolve()).rstrip('/').split('/')) - len(str(base.resolve()).rstrip('/').split('/')))
                            size = 0; mtime = None; ext = ''; mime = None
                            if typ == 'file':
                                try:
                                    st = pth.stat(); size = int(st.st_size); mtime = float(st.st_mtime)
                                except Exception:
                                    size = 0; mtime = None
                                ext = pth.suffix.lower()
                            remote = f"local:{os.uname().nodename}" if provider_id == 'local_fs' else f"mounted:{root_id}"
                            return (full, parent, name, depth, typ, size, mtime, ext, mime, None, None, remote, scan_id, None)
                        rows = []
                        for d in sorted(items_dirs, key=lambda x: str(x)):
                            rows.append(_row_from_local(d, 'folder'))
                        for fpath in items_files:
                            rows.append(_row_from_local(fpath, 'file'))
                        # dedupe
                        try:
                            seen = set(); uniq = []
                            for r in rows:
                                key = (r[0], r[4])
                                if key in seen: continue
                                seen.add(key); uniq.append(r)
                            rows = uniq
                        except Exception:
                            pass
                        ingested = pix.batch_insert_files(rows)
                        # In-memory datasets and progress
                        processed = 0
                        for fpath in items_files:
                            if task.get('cancel_requested'):
                                task['status'] = 'canceled'; task['ended'] = time.time(); return
                            try:
                                ds = fs.create_dataset_node(fpath)
                                app.extensions['scidk']['graph'].upsert_dataset(ds)
                            except Exception:
                                pass
                            processed += 1; task['processed'] = processed
                            if task['total']:
                                task['progress'] = processed / task['total']
                        file_count = len(items_files)
                        # Folders meta
                        for d in items_dirs:
                            try:
                                parent = str(d.parent.resolve()) if d != d.parent else ''
                                folders_meta.append({'path': str(d.resolve()), 'name': d.name, 'parent': parent, 'parent_name': Path(parent).name if parent else ''})
                            except Exception:
                                continue
                        folder_count = len(items_dirs)

                    elif provider_id == 'rclone':
                        provs = app.extensions['scidk'].get('providers')
                        prov = provs.get('rclone') if provs else None
                        if not prov:
                            raise RuntimeError('rclone provider not available')
                        # Prefer fast_list for recursive unless specified
                        fast_list = True if recursive else False
                        try:
                            items = prov.list_files(path, recursive=recursive, fast_list=fast_list)  # type: ignore[attr-defined]
                        except Exception as ee:
                            raise RuntimeError(str(ee))
                        rows = []
                        seen_rows = set()
                        seen_folders = set()
                        def _add_folder(full_path: str, name: str, parent: str):
                            nonlocal folders_meta
                            if full_path in seen_folders: return
                            seen_folders.add(full_path)
                            try:
                                from .core.path_utils import parse_remote_path
                                info_par = parse_remote_path(parent)
                                if info_par.get('is_remote'):
                                    parts = info_par.get('parts') or []
                                    parent_name = (info_par.get('remote_name') or '') if not parts else parts[-1]
                                else:
                                    parent_name = Path(parent).name if parent else ''
                            except Exception:
                                parent_name = ''
                            folders_meta.append({'path': full_path, 'name': name, 'parent': parent, 'parent_name': parent_name})
                        from .core.path_utils import join_remote_path, parent_remote_path
                        for it in (items or []):
                            name = it.get('Name') or it.get('Path') or ''
                            if it.get('IsDir'):
                                if name:
                                    full = join_remote_path(path, name)
                                    parent = parent_remote_path(full)
                                    _add_folder(full, name, parent)
                                # rclone folder row
                                rrow = pix.map_rclone_item_to_row(it, path, scan_id)
                                key = (rrow[0], rrow[4])
                                if key not in seen_rows:
                                    seen_rows.add(key)
                                    rows.append(rrow)
                                continue
                            # rclone file row
                            rrow = pix.map_rclone_item_to_row(it, path, scan_id)
                            key = (rrow[0], rrow[4])
                            if key not in seen_rows:
                                seen_rows.add(key)
                                rows.append(rrow)
                            if recursive and name:
                                parts = [p for p in (name.split('/') if isinstance(name, str) else []) if p]
                                cur = ''
                                for i in range(len(parts)-1):
                                    cur = parts[i] if i == 0 else (cur + '/' + parts[i])
                                    full = join_remote_path(path, cur)
                                    parent = parent_remote_path(full)
                                    _add_folder(full, parts[i], parent)
                            # In-memory dataset for file
                            try:
                                size = int(it.get('Size') or 0)
                                full = join_remote_path(path, name)
                                ds = fs.create_dataset_remote(full, size_bytes=size, modified_ts=0.0, mime=None)
                                app.extensions['scidk']['graph'].upsert_dataset(ds)
                            except Exception:
                                pass
                            file_count += 1
                            task['processed'] = file_count
                        folder_count = len(seen_folders)
                        ingested = pix.batch_insert_files(rows)
                    else:
                        raise RuntimeError(f"provider {provider_id} not supported for background scan")

                    # Build scan record
                    ended = time.time()
                    after = set(ds.get('checksum') for ds in app.extensions['scidk']['graph'].list_datasets())
                    new_checksums = sorted(list(after - before))
                    by_ext = {}
                    ext_map = {ds.get('checksum'): ds.get('extension') or '' for ds in app.extensions['scidk']['graph'].list_datasets()}
                    for ch in new_checksums:
                        ext = ext_map.get(ch, ''); by_ext[ext] = by_ext.get(ext, 0) + 1
                    # Host/provider tagging
                    host_type = provider_id
                    host_id = None
                    try:
                        if provider_id == 'rclone':
                            host_id = f"rclone:{(root_id or '').rstrip(':')}"
                        elif provider_id == 'local_fs':
                            import socket as _sock
                            host_id = f"local:{_sock.gethostname()}"
                        elif provider_id == 'mounted_fs':
                            host_id = f"mounted:{root_id}"
                    except Exception:
                        host_id = f"{provider_id}:{root_id}" if root_id else provider_id
                    scan = {
                        'id': scan_id,
                        'path': str(path),
                        'recursive': bool(recursive),
                        'started': started_ts,
                        'ended': ended,
                        'duration_sec': ended - started_ts,
                        'file_count': int(file_count),
                        'folder_count': int(folder_count),
                        'checksums': new_checksums,
                        'folders': folders_meta,
                        'by_ext': by_ext,
                        'source': getattr(fs, 'last_scan_source', 'python') if provider_id in ('local_fs','mounted_fs') else f"provider:{provider_id}",
                        'errors': [],
                        'committed': False,
                        'committed_at': None,
                        'provider_id': provider_id,
                        'host_type': host_type,
                        'host_id': host_id,
                        'root_id': root_id,
                        'root_label': Path(root_id).name if root_id else None,
                        'scan_source': f"provider:{provider_id}",
                        'ingested_rows': int(ingested),
                        'config_json': {
                            'interpreters': {
                                'effective_enabled': sorted(list(app.extensions['scidk'].get('interpreters', {}).get('effective_enabled', []))),
                                'source': app.extensions['scidk'].get('interpreters', {}).get('source', 'default'),
                            }
                        },
                    }
                    scans[scan_id] = scan
                    # Persist scan summary to SQLite (best-effort)
                    try:
                        from .core import path_index_sqlite as pix
                        from .core import migrations as _migs
                        import json as _json
                        conn = pix.connect()
                        try:
                            _migs.migrate(conn)
                            cur = conn.cursor()
                            cur.execute(
                                "INSERT OR REPLACE INTO scans(id, root, started, completed, status, extra_json) VALUES(?,?,?,?,?,?)",
                                (
                                    scan_id,
                                    str(path),
                                    float(started_ts or 0.0),
                                    float(ended or 0.0),
                                    'completed',
                                    _json.dumps({
                                        'recursive': bool(recursive),
                                        'duration_sec': ended - started_ts,
                                        'file_count': int(file_count),
                                        'by_ext': by_ext,
                                        'source': scan.get('source'),
                                        'checksums': new_checksums,
                                        'committed': False,
                                        'committed_at': None,
                                        'provider_id': provider_id,
                                        'root_id': root_id,
                                        'host_type': host_type,
                                        'host_id': host_id,
                                        'root_label': scan.get('root_label'),
                                    })
                                )
                            )
                            conn.commit()
                        finally:
                            try:
                                conn.close()
                            except Exception:
                                pass
                    except Exception:
                        pass
                    # Telemetry and directories
                    app.extensions['scidk'].setdefault('telemetry', {})['last_scan'] = {
                        'path': str(path), 'recursive': bool(recursive), 'scanned': int(file_count),
                        'started': started_ts, 'ended': ended, 'duration_sec': ended - started_ts,
                        'source': scan['source'], 'provider_id': provider_id, 'root_id': root_id,
                    }
                    dirs = app.extensions['scidk'].setdefault('directories', {})
                    drec = dirs.setdefault(str(path), {'path': str(path), 'recursive': bool(recursive), 'scanned': 0, 'last_scanned': 0, 'scan_ids': [], 'source': scan['source'], 'provider_id': provider_id, 'root_id': root_id, 'root_label': scan.get('root_label')})
                    drec.update({'recursive': bool(recursive), 'scanned': int(file_count), 'last_scanned': ended, 'source': scan['source'], 'provider_id': provider_id, 'root_id': root_id, 'root_label': scan.get('root_label')})
                    drec.setdefault('scan_ids', []).append(scan_id)

                    # Complete task
                    task['ended'] = ended
                    task['status'] = 'completed'
                    task['scan_id'] = scan_id
                    task['progress'] = 1.0
                except Exception as e:
                    import time as _t
                    task['ended'] = _t.time()
                    task['status'] = 'error'
                    task['error'] = str(e)
            threading.Thread(target=_worker, daemon=True).start()
            return jsonify({'task_id': task_id, 'status': 'running'}), 202

        elif ttype == 'commit':
            scan_id = (data.get('scan_id') or '').strip()
            scans = app.extensions['scidk'].setdefault('scans', {})
            s = scans.get(scan_id)
            if not s:
                return jsonify({'error': 'scan not found'}), 404
            checksums = s.get('checksums') or []
            total = len(checksums)
            tid_src = f"commit|{scan_id}|{started}"
            task_id = hashlib.sha1(tid_src.encode()).hexdigest()[:12]
            task = {
                'id': task_id,
                'type': 'commit',
                'status': 'running',
                'scan_id': scan_id,
                'path': s.get('path'),
                'started': started,
                'ended': None,
                # include one extra step for the Neo4j write phase so progress doesn't hit 100% before completion
                'total': total + 1,
                'processed': 0,
                'progress': 0.0,
                'neo4j_attempted': False,
                'neo4j_written': 0,
                'neo4j_error': None,
                'error': None,
                'cancel_requested': False,
            }
            app.extensions['scidk'].setdefault('tasks', {})[task_id] = task

            def _worker_commit():
                try:
                    if task.get('cancel_requested'):
                        task['status'] = 'canceled'
                        task['ended'] = time.time()
                        return
                    g = app.extensions['scidk']['graph']
                    # In-memory commit first (idempotent)
                    g.commit_scan(s)
                    s['committed'] = True
                    s['committed_at'] = time.time()
                    # Persist commit status to SQLite (best-effort)
                    try:
                        from .core import path_index_sqlite as pix
                        import json as _json
                        conn = pix.connect()
                        try:
                            cur = conn.cursor()
                            # fetch existing extra_json to merge
                            cur.execute("SELECT extra_json FROM scans WHERE id = ?", (s.get('id'),))
                            row = cur.fetchone()
                            extra_obj = {}
                            try:
                                if row and row[0]:
                                    extra_obj = _json.loads(row[0])
                            except Exception:
                                extra_obj = {}
                            extra_obj['committed'] = True
                            extra_obj['committed_at'] = s.get('committed_at')
                            cur.execute(
                                "UPDATE scans SET status = ?, extra_json = ? WHERE id = ?",
                                ('committed', _json.dumps(extra_obj), s.get('id'))
                            )
                            conn.commit()
                        finally:
                            try:
                                conn.close()
                            except Exception:
                                pass
                    except Exception:
                        pass
                    # Build rows once using shared builder when index mode is enabled
                    use_index = (os.environ.get('SCIDK_COMMIT_FROM_INDEX') or '').strip().lower() in ('1','true','yes','y','on')
                    if use_index:
                        from .core.commit_rows_from_index import build_rows_for_scan_from_index
                        rows, folder_rows = build_rows_for_scan_from_index(scan_id, s, include_hierarchy=True)
                    else:
                        ds_map = getattr(g, 'datasets', {})
                        rows, folder_rows = build_commit_rows(s, ds_map)
                    # Update progress for the file-processing phase
                    task['processed'] = total
                    if total:
                        task['progress'] = total / (task.get('total') or (total + 1))
                    # Allow cancel before Neo4j step
                    if task.get('cancel_requested'):
                        task['status'] = 'canceled'
                        task['ended'] = time.time()
                        return
                    # Neo4j write if configured via helper
                    uri, user, pwd, database, auth_mode = _get_neo4j_params()
                    def _on_prog(e, p):
                        try:
                            app.logger.info(f"neo4j {e}: {p}")
                        except Exception:
                            pass
                    if app.config.get('TESTING'):
                        result = commit_to_neo4j(rows, folder_rows, s, (uri, user, pwd, database, auth_mode))
                    else:
                        result = commit_to_neo4j_batched(
                            rows=rows,
                            folder_rows=folder_rows,
                            scan=s,
                            neo4j_params=(uri, user, pwd, database, auth_mode),
                            file_batch_size=int(os.environ.get('SCIDK_NEO4J_FILE_BATCH') or 5000),
                            folder_batch_size=int(os.environ.get('SCIDK_NEO4J_FOLDER_BATCH') or 5000),
                            max_retries=2,
                            on_progress=_on_prog
                        )
                    if result['attempted']:
                        task['neo4j_attempted'] = True
                    if result['error']:
                        task['neo4j_error'] = result['error']
                    task['neo4j_written'] = int(result.get('written_files', 0)) + int(result.get('written_folders', 0))
                    # Include DB verification results if available
                    if 'db_verified' in result:
                        task['neo4j_db_verified'] = bool(result.get('db_verified'))
                        task['neo4j_db_files'] = int(result.get('db_files') or 0)
                        task['neo4j_db_folders'] = int(result.get('db_folders') or 0)
                        if task['neo4j_attempted'] and not task['neo4j_db_verified'] and not task.get('neo4j_error'):
                            task['neo4j_error'] = 'Post-commit verification found 0 SCANNED_IN edges for this scan. Check Neo4j credentials/database or permissions.'
                    # Done
                    # mark final step (Neo4j write) as processed so progress reaches 100% only at the end
                    task['processed'] = task.get('total') or task.get('processed')
                    task['ended'] = time.time()
                    task['status'] = 'completed'
                    task['progress'] = 1.0
                except Exception as e:
                    task['ended'] = time.time()
                    task['status'] = 'error'
                    task['error'] = str(e)
            threading.Thread(target=_worker_commit, daemon=True).start()
            return jsonify({'task_id': task_id, 'status': 'running'}), 202

        else:
            return jsonify({"error": "unsupported task type"}), 400

    @api.get('/tasks')
    def api_tasks_list():
        tasks = list(app.extensions['scidk'].get('tasks', {}).values())
        # sort newest first
        tasks.sort(key=lambda t: t.get('started') or 0, reverse=True)
        return jsonify(tasks), 200

    @api.get('/tasks/<task_id>')
    def api_tasks_detail(task_id):
        task = app.extensions['scidk'].get('tasks', {}).get(task_id)
        if not task:
            return jsonify({"error": "not found"}), 404
        return jsonify(task), 200

    @api.post('/tasks/<task_id>/cancel')
    def api_tasks_cancel(task_id):
        tasks = app.extensions['scidk'].setdefault('tasks', {})
        task = tasks.get(task_id)
        if not task:
            return jsonify({'error': 'not found'}), 404
        # only running tasks can be canceled
        if task.get('status') != 'running':
            return jsonify({'status': task.get('status'), 'message': 'task not running'}), 400
        task['cancel_requested'] = True
        return jsonify({'status': 'canceling'}), 202

    @api.get('/datasets')
    def api_datasets():
        items = graph.list_datasets()
        return jsonify(items)

    @api.get('/datasets/<dataset_id>')
    def api_dataset(dataset_id):
        item = graph.get_dataset(dataset_id)
        if not item:
            return jsonify({"error": "not found"}), 404
        return jsonify(item)

    @api.post('/interpret')
    def api_interpret():
        data = request.get_json(force=True, silent=True) or {}
        dataset_id = data.get('dataset_id')
        interpreter_id = data.get('interpreter_id')
        if not dataset_id:
            return jsonify({"status": "error", "error": "dataset_id required"}), 400
        ds = graph.get_dataset(dataset_id)
        if not ds:
            return jsonify({"status": "error", "error": "dataset not found"}), 404
        file_path = Path(ds['path'])
        if interpreter_id:
            interp = registry.get_by_id(interpreter_id)
            if not interp:
                return jsonify({"status": "error", "error": "interpreter not found"}), 404
            interps = [interp]
        else:
            interps = registry.select_for_dataset(ds)
            if not interps:
                return jsonify({"status": "error", "error": "no interpreters available"}), 400
        results = []
        for interp in interps:
            try:
                _t0 = time.time()
                result = interp.interpret(file_path)
                _t1 = time.time()
                graph.add_interpretation(ds['checksum'], interp.id, {
                    'status': result.get('status', 'success'),
                    'data': result.get('data', result),
                    'interpreter_version': getattr(interp, 'version', '0.0.1'),
                })
                # Record success
                try:
                    registry.record_usage(interp.id, success=True, execution_time_ms=int((_t1 - _t0)*1000))
                except Exception:
                    pass
                results.append({'interpreter_id': interp.id, 'status': 'ok'})
            except Exception as e:
                try:
                    registry.record_usage(interp.id, success=False, execution_time_ms=0)
                except Exception:
                    pass
                graph.add_interpretation(ds['checksum'], interp.id, {
                    'status': 'error',
                    'data': {'error': str(e)},
                    'interpreter_version': getattr(interp, 'version', '0.0.1'),
                })
                results.append({'interpreter_id': interp.id, 'status': 'error', 'error': str(e)})
        return jsonify({"status": "ok", "results": results}), 200

    @api.post('/chat')
    def api_chat():
        data = request.get_json(force=True, silent=True) or {}
        message = (data.get('message') or '').strip()
        if not message:
            return jsonify({"status": "error", "error": "message required"}), 400
        store = app.extensions['scidk'].setdefault('chat', {"history": []})
        # Simple echo bot with count
        reply = f"Echo: {message}"
        entry_user = {"role": "user", "content": message}
        entry_assistant = {"role": "assistant", "content": reply}
        store['history'].append(entry_user)
        store['history'].append(entry_assistant)
        return jsonify({"status": "ok", "reply": reply, "history": store['history']}), 200

    @api.get('/search')
    def api_search():
        q = (request.args.get('q') or '').strip()
        if not q:
            return jsonify([]), 200
        q_lower = q.lower()
        results = []
        for ds in graph.list_datasets():
            matched_on = []
            # Match filename
            if q_lower in (ds.get('filename') or '').lower() or q_lower in (ds.get('path') or '').lower():
                matched_on.append('filename')
            # Match interpreter ids present
            interps = (ds.get('interpretations') or {})
            for interp_id in interps.keys():
                if q_lower in interp_id.lower():
                    if 'interpreter_id' not in matched_on:
                        matched_on.append('interpreter_id')
            if matched_on:
                results.append({
                    'id': ds.get('id'),
                    'path': ds.get('path'),
                    'filename': ds.get('filename'),
                    'extension': ds.get('extension'),
                    'matched_on': matched_on,
                })
        # Simple ordering: filename matches first, then interpreter_id
        def score(r):
            return (0 if 'filename' in r['matched_on'] else 1, r['filename'] or '')
        results.sort(key=score)
        return jsonify(results), 200

    @api.get('/interpreters')
    def api_interpreters():
        # Unified listing: registry metadata + toggle/usage/metrics + effective view override
        reg = app.extensions['scidk']['registry']
        # Build mapping ext -> interpreter ids
        ext_map = {}
        for ext, interps in reg.by_extension.items():
            ext_map[ext] = [getattr(i, 'id', 'unknown') for i in interps]
        items = []
        for iid, interp in reg.by_id.items():
            globs = sorted([ext for ext, ids in ext_map.items() if iid in ids])
            it = {
                'id': iid,
                'name': getattr(interp, 'name', iid),
                'version': getattr(interp, 'version', '0.0.1'),
                'globs': globs,
                'default_enabled': bool(getattr(interp, 'default_enabled', getattr(reg, 'default_enabled', True))),
                'cost': getattr(interp, 'cost', None),
                'extensions': globs,
                'enabled': True,
                'runtime': getattr(interp, 'runtime', 'python'),
                'last_used': getattr(reg, 'get_last_used', lambda _x: None)(iid),
                'success_rate': getattr(reg, 'get_success_rate', lambda _x: 0.0)(iid),
            }
            try:
                it['enabled'] = reg._is_enabled(iid)
            except Exception:
                pass
            items.append(it)
        # Optional effective view from app extensions (e.g., CLI/env overridden)
        view = (request.args.get('view') or '').strip().lower()
        if view == 'effective':
            interp_state = app.extensions['scidk'].get('interpreters', {})
            eff = set(interp_state.get('effective_enabled') or [])
            src = interp_state.get('source') or 'default'
            for it in items:
                it['enabled'] = (it['id'] in eff)
                it['source'] = src
        return jsonify(items), 200

    @api.get('/interpreters/effective_debug')
    def api_interpreters_effective_debug():
        istate = app.extensions.get('scidk', {}).get('interpreters', {})
        eff = sorted(list(istate.get('effective_enabled') or []))
        src = istate.get('source') or 'default'
        unknown_env = istate.get('unknown_env') or {}
        reg = app.extensions['scidk']['registry']
        all_ids = sorted(list(reg.by_id.keys()))
        default_enabled = []
        for iid in all_ids:
            try:
                if bool(getattr(reg.by_id[iid], 'default_enabled', True)):
                    default_enabled.append(iid)
            except Exception:
                pass
        loaded = []
        try:
            settings = app.extensions['scidk'].get('settings')
            if settings is not None and src != 'cli':
                loaded = sorted(list(settings.load_enabled_interpreters() or []))
        except Exception:
            loaded = []
        en_raw = [s.strip() for s in (os.environ.get('SCIDK_ENABLE_INTERPRETERS') or '').split(',') if s.strip()]
        dis_raw = [s.strip() for s in (os.environ.get('SCIDK_DISABLE_INTERPRETERS') or '').split(',') if s.strip()]
        en_norm = [s.lower() for s in en_raw]
        dis_norm = [s.lower() for s in dis_raw]
        return jsonify({
            'source': src,
            'effective_enabled': eff,
            'default_enabled': sorted(default_enabled),
            'loaded_settings': loaded,
            'env': {
                'enable_raw': en_raw,
                'disable_raw': dis_raw,
                'enable_norm': en_norm,
                'disable_norm': dis_norm,
                'unknown': unknown_env,
            }
        }), 200

    @api.post('/interpreters/<interpreter_id>/toggle')
    def api_interpreters_toggle(interpreter_id):
        reg = app.extensions['scidk']['registry']
        data = request.get_json(force=True, silent=True) or {}
        enabled = bool(data.get('enabled', True))
        if enabled:
            reg.enable_interpreter(interpreter_id)
        else:
            reg.disable_interpreter(interpreter_id)
        # Persist if settings available
        try:
            settings = app.extensions['scidk'].get('settings')
            if settings is not None:
                settings.save_enabled_interpreters(reg.enabled_interpreters)
        except Exception:
            pass
        # Refresh effective interpreter view so /api/interpreters?view=effective reflects the change immediately
        try:
            istate = app.extensions['scidk'].setdefault('interpreters', {})
            eff = set(istate.get('effective_enabled') or [])
            # If snapshot missing/empty, rebuild from current registry state
            if not eff:
                eff = set([iid for iid in reg.by_id.keys() if reg._is_enabled(iid)])
            if enabled:
                eff.add(interpreter_id)
            else:
                eff.discard(interpreter_id)
            istate['effective_enabled'] = eff
        except Exception:
            pass
        return jsonify({'status': 'updated', 'enabled': enabled}), 200

    @api.get('/providers')
    def api_providers():
        provs = app.extensions['scidk']['providers']
        out = []
        for d in provs.list():
            out.append({
                'id': d.id,
                'display_name': d.display_name,
                'capabilities': d.capabilities,
                'auth': d.auth,
            })
        return jsonify(out), 200

    @api.get('/provider_roots')
    def api_provider_roots():
        prov_id = (request.args.get('provider_id') or 'local_fs').strip() or 'local_fs'
        try:
            provs = app.extensions['scidk']['providers']
            prov = provs.get(prov_id)
            if not prov:
                return jsonify({'error': 'provider not available'}), 400
            roots = prov.list_roots()
            return jsonify([{'id': r.id, 'name': r.name, 'path': r.path} for r in roots]), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @api.get('/browse')
    def api_browse():
        prov_id = (request.args.get('provider_id') or 'local_fs').strip() or 'local_fs'
        root_id = (request.args.get('root_id') or '/').strip() or '/'
        path_q = (request.args.get('path') or '').strip()
        _t0 = _time.time()
        try:
            provs = app.extensions['scidk']['providers']
            prov = provs.get(prov_id)
            if not prov:
                return jsonify({'error': 'provider not available', 'code': 'provider_not_available'}), 400
            # If path empty, default to root_id
            # Parse rclone browse options
            opts = {}
            if prov_id == 'rclone':
                rec_s = (request.args.get('recursive') or '').strip().lower()
                fast_s = (request.args.get('fast_list') or '').strip().lower()
                depth_s = (request.args.get('max_depth') or '').strip()
                opts['recursive'] = (rec_s in ('1','true','yes','on'))
                opts['fast_list'] = (fast_s in ('1','true','yes','on'))
                try:
                    opts['max_depth'] = int(depth_s) if depth_s else 1
                except Exception:
                    opts['max_depth'] = 1
            listing = prov.list(root_id=root_id, path=path_q or root_id, **opts)
            # Bubble provider-level errors clearly
            if isinstance(listing, dict) and listing.get('error'):
                return jsonify({'error': listing.get('error'), 'code': 'browse_failed'}), 400
            # Augment with provider badge and convenience fields
            for e in listing.get('entries', []):
                e['provider_id'] = prov_id
            try:
                from .services.metrics import record_latency
                record_latency(app, 'browse', _time.time() - _t0)
            except Exception:
                pass
            return jsonify(listing), 200
        except Exception as e:
            try:
                from .services.metrics import record_latency
                record_latency(app, 'browse', _time.time() - _t0)
            except Exception:
                pass
            return jsonify({'error': str(e), 'code': 'browse_exception'}), 500

    @api.get('/directories')
    def api_directories():
        # Prefer SQLite-backed aggregation by root; fallback to in-memory registry
        try:
            from .core import path_index_sqlite as pix
            from .core import migrations as _migs
            import json as _json
            conn = pix.connect()
            try:
                _migs.migrate(conn)
                cur = conn.cursor()
                cur.execute("SELECT id, root, completed, extra_json FROM scans WHERE root IS NOT NULL AND root <> '' ORDER BY coalesce(completed, 0) DESC LIMIT 2000")
                rows = cur.fetchall()
                agg = {}
                for (sid, root, completed, extra) in rows:
                    if not root:
                        continue
                    rec = agg.get(root) or {'path': root, 'scanned': 0, 'last_scanned': 0, 'scan_ids': [], 'recursive': None}
                    rec['scan_ids'].append(sid)
                    try:
                        if completed and float(completed) > float(rec.get('last_scanned') or 0):
                            rec['last_scanned'] = float(completed)
                    except Exception:
                        pass
                    try:
                        ex = _json.loads(extra) if extra else {}
                        # Best-effort fields
                        if ex:
                            if rec.get('recursive') is None:
                                rec['recursive'] = bool(ex.get('recursive', False))
                            if 'file_count' in ex:
                                rec['scanned'] = int(ex.get('file_count') or rec.get('scanned') or 0)
                            if 'source' in ex and not rec.get('source'):
                                rec['source'] = ex.get('source')
                            if 'provider_id' in ex and not rec.get('provider_id'):
                                rec['provider_id'] = ex.get('provider_id')
                            if 'root_id' in ex and not rec.get('root_id'):
                                rec['root_id'] = ex.get('root_id')
                            if 'root_label' in ex and not rec.get('root_label'):
                                rec['root_label'] = ex.get('root_label')
                    except Exception:
                        pass
                    agg[root] = rec
                values = list(agg.values())
                values.sort(key=lambda d: d.get('last_scanned') or 0, reverse=True)
                # Fill defaults
                for v in values:
                    if v.get('recursive') is None:
                        v['recursive'] = False
                return jsonify(values), 200
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass
        # Fallback
        dirs = app.extensions['scidk'].get('directories', {})
        values = list(dirs.values())
        values.sort(key=lambda d: d.get('last_scanned') or 0, reverse=True)
        return jsonify(values), 200

    # -----------------------------
    # Rclone Mount Manager (flagged)
    # -----------------------------
    if _feature_rclone_mounts():
        import time, subprocess, shutil, json as _json

        def _mounts_dir() -> Path:
            d = Path(app.root_path).parent / 'data' / 'mounts'
            d.mkdir(parents=True, exist_ok=True)
            return d

        def _sanitize_name(name: str) -> str:
            safe = ''.join([c for c in (name or '') if c.isalnum() or c in ('-', '_')]).strip()
            return safe[:64] if safe else ''

        def _listremotes() -> list:
            try:
                provs = app.extensions['scidk']['providers']
                rp = provs.get('rclone') if provs else None
                roots = rp.list_roots() if rp else []
                return [r.id for r in roots]
            except Exception:
                return []

        def _rclone_exe() -> Optional[str]:
            return shutil.which('rclone')

        @api.get('/rclone/mounts')
        def api_rclone_mounts_list():
            mounts_mem = app.extensions['scidk'].setdefault('rclone_mounts', {})
            rows = []
            try:
                from .core import path_index_sqlite as pix
                from .core import migrations as _migs
                import json as _json
                conn = pix.connect()
                try:
                    _migs.migrate(conn)
                    cur = conn.cursor()
                    cur.execute("SELECT id, provider, root, created, status, extra_json FROM provider_mounts WHERE provider='rclone'")
                    rows = cur.fetchall() or []
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            except Exception:
                rows = []
            out = []
            for (mid, provider, remote, created, status_persisted, extra) in rows:
                try:
                    extra_obj = json.loads(extra) if extra else {}
                except Exception:
                    extra_obj = {}
                mem = mounts_mem.get(mid) or {}
                proc = mem.get('process')
                alive = (proc is not None) and (proc.poll() is None)
                status = 'running' if alive else ('exited' if proc is not None else (status_persisted or 'unknown'))
                exit_code = None if alive else (proc.returncode if proc is not None else None)
                out.append({
                    'id': mid,
                    'name': mid,
                    'remote': remote,
                    'subpath': extra_obj.get('subpath'),
                    'path': extra_obj.get('path'),
                    'read_only': extra_obj.get('read_only'),
                    'started_at': created,
                    'status': status,
                    'exit_code': exit_code,
                    'log_file': extra_obj.get('log_file'),
                    'pid': extra_obj.get('pid') or mem.get('pid'),
                })
            return jsonify(out), 200

        @api.post('/rclone/mounts')
        def api_rclone_mounts_create():
            if not _rclone_exe():
                return jsonify({'error': 'rclone not installed'}), 400
            try:
                body = request.get_json(silent=True) or {}
                remote = str(body.get('remote') or '').strip()
                subpath = str(body.get('subpath') or '').strip().lstrip('/')
                name = _sanitize_name(str(body.get('name') or ''))
                read_only = bool(body.get('read_only', True))
                if not remote:
                    return jsonify({'error': 'remote required'}), 400
                if not remote.endswith(':'):
                    remote = remote + ':'
                if not name:
                    return jsonify({'error': 'name required'}), 400
                # Safety: restrict mountpoint and validate remote exists
                remotes = _listremotes()
                if remote not in remotes:
                    return jsonify({'error': f'remote not configured: {remote}'}), 400
                mdir = _mounts_dir()
                mpath = mdir / name
                try:
                    mpath.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    return jsonify({'error': f'failed to create mount dir: {e}'}), 500
                target = remote + (subpath if subpath else '')
                log_file = mdir / f"{name}.log"
                # Build rclone command
                args = [
                    _rclone_exe(), 'mount', target, str(mpath),
                    '--dir-cache-time', '60m',
                    '--poll-interval', '30s',
                    '--vfs-cache-mode', 'minimal',
                    '--log-format', 'DATE,TIME,LEVEL',
                    '--log-level', 'INFO',
                    '--log-file', str(log_file),
                ]
                if read_only:
                    args.append('--read-only')
                # Launch subprocess detached from terminal; logs go to file
                try:
                    fnull = open(os.devnull, 'wb')
                    proc = subprocess.Popen(args, stdout=fnull, stderr=fnull)
                except Exception as e:
                    return jsonify({'error': f'failed to start rclone: {e}'}), 500
                rec = {
                    'id': name,
                    'name': name,
                    'remote': remote,
                    'subpath': subpath,
                    'path': str(mpath),
                    'read_only': bool(read_only),
                    'started_at': time.time(),
                    'process': proc,
                    'pid': proc.pid if proc else None,
                    'log_file': str(log_file),
                }
                mounts = app.extensions['scidk'].setdefault('rclone_mounts', {})
                mounts[name] = rec
                # Persist mount definition to SQLite (best-effort)
                try:
                    from .core import path_index_sqlite as pix
                    from .core import migrations as _migs
                    import json as _json
                    conn = pix.connect()
                    try:
                        _migs.migrate(conn)
                        cur = conn.cursor()
                        extra = {
                            'subpath': subpath,
                            'path': str(mpath),
                            'read_only': bool(read_only),
                            'log_file': str(log_file),
                            'pid': rec.get('pid'),
                        }
                        cur.execute(
                            "INSERT OR REPLACE INTO provider_mounts(id, provider, root, created, status, extra_json) VALUES(?,?,?,?,?,?)",
                            (name, 'rclone', remote, float(rec.get('started_at') or time.time()), 'running', _json.dumps(extra))
                        )
                        conn.commit()
                    finally:
                        try:
                            conn.close()
                        except Exception:
                            pass
                except Exception:
                    pass
                return jsonify({'id': name, 'path': str(mpath)}), 201
            except Exception as e:
                return jsonify({'error': str(e)}), 500

        @api.delete('/rclone/mounts/<mid>')
        def api_rclone_mounts_delete(mid):
            mounts = app.extensions['scidk'].setdefault('rclone_mounts', {})
            m = mounts.get(mid)
            if not m:
                return jsonify({'error': 'not found'}), 404
            proc = m.get('process')
            mpath = m.get('path')
            try:
                if proc and (proc.poll() is None):
                    proc.terminate()
                    try:
                        proc.wait(timeout=5)
                    except Exception:
                        proc.kill()
                # Best-effort unmount
                try:
                    subprocess.run(['fusermount', '-u', mpath], check=False)
                except Exception:
                    pass
                try:
                    subprocess.run(['umount', mpath], check=False)
                except Exception:
                    pass
            except Exception:
                pass
            mounts.pop(mid, None)
            # Remove persisted mount definition (best-effort)
            try:
                from .core import path_index_sqlite as pix
                from .core import migrations as _migs
                conn = pix.connect()
                try:
                    _migs.migrate(conn)
                    cur = conn.cursor()
                    cur.execute("DELETE FROM provider_mounts WHERE id = ?", (mid,))
                    conn.commit()
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            except Exception:
                pass
            return jsonify({'ok': True}), 200

        @api.get('/rclone/mounts/<mid>/logs')
        def api_rclone_mounts_logs(mid):
            mounts = app.extensions['scidk'].setdefault('rclone_mounts', {})
            m = mounts.get(mid)
            if not m:
                return jsonify({'error': 'not found'}), 404
            tail_n = int(request.args.get('tail') or 200)
            path = m.get('log_file')
            lines = []
            try:
                with open(path, 'r', encoding='utf-8', errors='ignore') as fh:
                    lines = fh.readlines()
            except Exception:
                lines = []
            if tail_n > 0 and len(lines) > tail_n:
                lines = lines[-tail_n:]
            return jsonify({'lines': [ln.rstrip('\n') for ln in lines]}), 200

        @api.get('/rclone/mounts/<mid>/health')
        def api_rclone_mounts_health(mid):
            mounts = app.extensions['scidk'].setdefault('rclone_mounts', {})
            m = mounts.get(mid)
            if not m:
                return jsonify({'ok': False, 'error': 'not found'}), 404
            proc = m.get('process')
            alive = (proc is not None) and (proc.poll() is None)
            path = m.get('path')
            listable = False
            try:
                p = Path(path)
                listable = p.exists() and p.is_dir() and (len(list(p.iterdir())) >= 0)
            except Exception:
                listable = False
            return jsonify({'ok': bool(alive and listable), 'alive': bool(alive), 'listable': bool(listable)}), 200

    @api.get('/fs/list')
    def api_fs_list():
        """List immediate children within a scanned base directory.
        Query params:
          - base (required): must equal a previously scanned directory path
          - path (optional): if provided, must resolve under base; otherwise list base
        Returns JSON with breadcrumb and items. Prevents path traversal outside base.
        """
        base = (request.args.get('base') or '').strip()
        rel_path = (request.args.get('path') or '').strip()
        if not base:
            return jsonify({"error": "missing base"}), 400
        dirs = app.extensions['scidk'].get('directories', {})
        if base not in dirs:
            return jsonify({"error": "unknown base (run a scan first)"}), 400
        try:
            base_p = Path(base).resolve()
            cur_p = Path(rel_path).resolve() if rel_path else base_p
            # Ensure cur_p is under base
            try:
                cur_p.relative_to(base_p)
            except Exception:
                cur_p = base_p
            if not cur_p.exists() or not cur_p.is_dir():
                return jsonify({"error": "path not a directory"}), 400
            # Build breadcrumb from base to cur
            breadcrumb = []
            # iterate ancestors from base to cur
            parts = []
            tmp = cur_p
            while True:
                parts.append(tmp)
                if tmp == base_p:
                    break
                tmp = tmp.parent
                if tmp == tmp.parent:  # reached filesystem root
                    break
            parts.reverse()
            for p in parts:
                try:
                    breadcrumb.append({"name": p.name or str(p), "path": str(p)})
                except Exception:
                    breadcrumb.append({"name": str(p), "path": str(p)})
            # Precompute scanned dataset paths
            scanned_paths = {}
            for d in app.extensions['scidk']['graph'].list_datasets():
                scanned_paths[d.get('path')] = d.get('id')
            # List items
            items = []
            for child in cur_p.iterdir():
                try:
                    st = child.stat()
                    is_dir = child.is_dir()
                    item = {
                        'name': child.name,
                        'path': str(child.resolve()),
                        'is_dir': bool(is_dir),
                        'size_bytes': 0 if is_dir else int(st.st_size),
                        'modified': float(st.st_mtime),
                        'ext': '' if is_dir else child.suffix.lower(),
                        'scanned': False,
                        'dataset_id': None,
                    }
                    if not is_dir:
                        dsid = scanned_paths.get(str(child.resolve()))
                        if dsid:
                            item['scanned'] = True
                            item['dataset_id'] = dsid
                    items.append(item)
                except Exception:
                    continue
            # Sort: directories first, then files by name
            items.sort(key=lambda x: (0 if x['is_dir'] else 1, x['name'].lower()))
            return jsonify({
                'base': str(base_p),
                'path': str(cur_p),
                'breadcrumb': breadcrumb,
                'items': items,
            }), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @api.route('/scans', methods=['GET', 'POST'])
    def api_scans():
        # POST creates a new scan (alias of legacy /api/scan)
        if request.method == 'POST':
            return api_scan()
        # GET: Prefer SQLite-backed history with in-memory fallback
        summaries = []
        try:
            from .core import path_index_sqlite as pix
            import json as _json
            conn = pix.connect()
            try:
                from .core import migrations as _migs
                _migs.migrate(conn)
                cur = conn.cursor()
                cur.execute("SELECT id, root, started, completed, status, extra_json FROM scans ORDER BY coalesce(completed, started) DESC LIMIT 500")
                rows = cur.fetchall()
                for (sid, root, started, completed, status, extra) in rows:
                    extra_obj = {}
                    try:
                        if extra:
                            extra_obj = _json.loads(extra)
                    except Exception:
                        extra_obj = {}
                    summaries.append({
                        'id': sid,
                        'path': root,
                        'recursive': bool((extra_obj or {}).get('recursive')),
                        'started': started,
                        'ended': completed,
                        'duration_sec': (extra_obj or {}).get('duration_sec'),
                        'file_count': (extra_obj or {}).get('file_count'),
                        'by_ext': (extra_obj or {}).get('by_ext') or {},
                        'source': (extra_obj or {}).get('source'),
                        'checksum_count': len((extra_obj or {}).get('checksums') or []),
                        'committed': bool((extra_obj or {}).get('committed', False)),
                        'committed_at': (extra_obj or {}).get('committed_at'),
                        'status': status,
                    })
                # Merge in-memory committed flags to reflect immediate commits
                try:
                    inmem = {s.get('id'): s for s in app.extensions['scidk'].get('scans', {}).values()}
                    for i in range(len(summaries)):
                        sid = summaries[i].get('id')
                        if sid in inmem:
                            if bool(inmem[sid].get('committed')):
                                summaries[i]['committed'] = True
                                summaries[i]['committed_at'] = inmem[sid].get('committed_at')
                except Exception:
                    pass
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            summaries = []
        if not summaries:
            scans = list(app.extensions['scidk'].get('scans', {}).values())
            scans.sort(key=lambda s: s.get('ended') or s.get('started') or 0, reverse=True)
            summaries = [
                {
                    'id': s.get('id'),
                    'path': s.get('path'),
                    'recursive': s.get('recursive'),
                    'started': s.get('started'),
                    'ended': s.get('ended'),
                    'duration_sec': s.get('duration_sec'),
                    'file_count': s.get('file_count'),
                    'by_ext': s.get('by_ext', {}),
                    'source': s.get('source'),
                    'checksum_count': len(s.get('checksums') or []),
                    'committed': bool(s.get('committed', False)),
                    'committed_at': s.get('committed_at'),
                }
                for s in scans
            ]
        return jsonify(summaries), 200

    @api.get('/scans/<scan_id>')
    def api_scan_detail(scan_id):
        s = app.extensions['scidk'].get('scans', {}).get(scan_id)
        if not s:
            # Try to reconstruct minimal scan dict from SQLite persistence
            try:
                from .core import path_index_sqlite as pix
                from .core import migrations as _migs
                import json as _json
                conn = pix.connect()
                try:
                    _migs.migrate(conn)
                    cur = conn.cursor()
                    cur.execute("SELECT id, root, started, completed, status, extra_json FROM scans WHERE id = ?", (scan_id,))
                    row = cur.fetchone()
                    if row:
                        sid, root, started, completed, status, extra = row
                        extra_obj = {}
                        try:
                            if extra:
                                extra_obj = _json.loads(extra)
                        except Exception:
                            extra_obj = {}
                        s = {
                            'id': sid,
                            'path': root,
                            'recursive': bool((extra_obj or {}).get('recursive')),
                            'started': started,
                            'ended': completed,
                            'duration_sec': (extra_obj or {}).get('duration_sec'),
                            'file_count': (extra_obj or {}).get('file_count'),
                            'by_ext': (extra_obj or {}).get('by_ext') or {},
                            'source': (extra_obj or {}).get('source'),
                            'checksums': (extra_obj or {}).get('checksums') or [],
                            'committed': bool((extra_obj or {}).get('committed', False)),
                            'committed_at': (extra_obj or {}).get('committed_at'),
                            'provider_id': (extra_obj or {}).get('provider_id'),
                            'host_type': (extra_obj or {}).get('host_type'),
                            'host_id': (extra_obj or {}).get('host_id'),
                            'root_id': (extra_obj or {}).get('root_id'),
                            'root_label': (extra_obj or {}).get('root_label'),
                        }
                        # Cache minimal record in-memory to help downstream endpoints
                        app.extensions['scidk'].setdefault('scans', {})[scan_id] = s
                    else:
                        return jsonify({"error": "not found"}), 404
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            except Exception:
                return jsonify({"error": "not found"}), 404
        return jsonify(s), 200

    @api.post('/interpret/scan/<scan_id>')
    def api_interpret_scan(scan_id):
        """
        Interpret files for a completed scan, streaming remote content via rclone when needed (no mounts required).

        JSON body (optional):
          - include: ["*.py","*.ipynb","*.txt"]
          - exclude: ["*.csv"]
          - max_files: 200
          - max_size_bytes: 1048576
          - interpreters: ["python","ipynb","txt"]
          - overwrite: false
          - timeout_sec: 60
        """
        import fnmatch, json as _json
        body = request.get_json(force=True, silent=True) or {}
        include = body.get('include') or []
        exclude = body.get('exclude') or []
        # Client-requested batch size (subject to server cap)
        try:
            req_max = int(body.get('max_files')) if body.get('max_files') not in (None, '') else None
        except Exception:
            req_max = None
        cap = int(app.config.get('rclone.interpret.max_files_per_batch', 1000))
        cap = min(max(100, cap), 2000)
        max_files = int(req_max) if (req_max is not None) else cap
        max_files = min(max(1, max_files), cap)
        max_size_bytes = body.get('max_size_bytes')
        max_size_bytes = int(max_size_bytes) if max_size_bytes not in (None, '') else None
        only_interps = set((body.get('interpreters') or []))
        overwrite = bool(body.get('overwrite', False))
        timeout_sec = float(body.get('timeout_sec') or 60.0)

        # Ensure we have a scan record (use existing detail reconstruction if missing)
        scan = app.extensions['scidk'].get('scans', {}).get(scan_id)
        if not scan:
            resp = api_scan_detail(scan_id)
            # api_scan_detail returns (json, code) in some fallbacks; support both
            try:
                code = resp[1]
            except Exception:
                code = getattr(resp, 'status_code', 200)
            if code != 200:
                return resp
            try:
                scan = resp.get_json()  # Response
            except Exception:
                try:
                    scan = resp[0].json  # tuple from our internal call pattern
                except Exception:
                    scan = None
        if not scan:
            return jsonify({'status': 'error', 'error': 'scan not found'}), 404

        provider_id = (scan or {}).get('provider_id') or 'local_fs'
        provs = app.extensions['scidk'].get('providers') or {}
        rclone = provs.get('rclone') if provider_id == 'rclone' else None
        if provider_id == 'rclone' and not rclone:
            return jsonify({'status': 'error', 'error': 'rclone provider not available'}), 400

        # Load candidate files from SQLite for this scan_id
        try:
            from .core import path_index_sqlite as pix
            from .core import migrations as _migs
            conn = pix.connect(); _migs.migrate(conn)
            cur = conn.cursor()
            # Build base query with optional overwrite and cursor, then apply LIMIT
            if overwrite:
                if (body.get('after_rowid') not in (None, '')):
                    try:
                        after_rowid = int(body.get('after_rowid'))
                    except Exception:
                        after_rowid = None
                else:
                    after_rowid = None
                if after_rowid is not None:
                    cur.execute(
                        "SELECT rowid, path, size, file_extension FROM files WHERE scan_id=? AND type='file' AND rowid > ? ORDER BY rowid ASC LIMIT ?",
                        (scan_id, after_rowid, max_files)
                    )
                else:
                    cur.execute(
                        "SELECT rowid, path, size, file_extension FROM files WHERE scan_id=? AND type='file' ORDER BY rowid ASC LIMIT ?",
                        (scan_id, max_files)
                    )
            else:
                if (body.get('after_rowid') not in (None, '')):
                    try:
                        after_rowid = int(body.get('after_rowid'))
                    except Exception:
                        after_rowid = None
                else:
                    after_rowid = None
                if after_rowid is not None:
                    cur.execute(
                        "SELECT rowid, path, size, file_extension FROM files WHERE scan_id=? AND type='file' AND (interpreted_as IS NULL OR interpreted_as='') AND rowid > ? ORDER BY rowid ASC LIMIT ?",
                        (scan_id, after_rowid, max_files)
                    )
                else:
                    cur.execute(
                        "SELECT rowid, path, size, file_extension FROM files WHERE scan_id=? AND type='file' AND (interpreted_as IS NULL OR interpreted_as='') ORDER BY rowid ASC LIMIT ?",
                        (scan_id, max_files)
                    )
            rows = cur.fetchall() or []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # Interpreter selection based on registry rules and extension map
        registry: InterpreterRegistry = app.extensions['scidk']['registry']
        # Build extension -> interpreters map using existing registry data
        ext_to_interpreters = {}
        try:
            # registry.by_extension is available; also respect only_interps filter by id
            for ext_key, interps in getattr(registry, 'by_extension', {}).items():
                cand = []
                for interp in interps:
                    iid = getattr(interp, 'id', '')
                    if only_interps and iid not in only_interps:
                        continue
                    # Use registry enabled state if any explicitly enabled
                    if getattr(registry, 'enabled_interpreters', None):
                        if iid not in registry.enabled_interpreters:
                            continue
                    cand.append(interp)
                if cand:
                    ext_to_interpreters[ext_key.lower()] = cand
        except Exception:
            ext_to_interpreters = {}

        def _wanted(path: str) -> bool:
            if include and not any(fnmatch.fnmatch(path, pat) for pat in include):
                return False
            if exclude and any(fnmatch.fnmatch(path, pat) for pat in exclude):
                return False
            return True

        # Diagnostics counters
        files_seen = 0
        filtered_by_size = 0
        filtered_by_include = 0
        filtered_no_interpreter = 0

        candidates = []
        last_rowid = 0
        for (rowid, path, size, ext) in rows:
            try:
                last_rowid = int(rowid)
            except Exception:
                pass
            files_seen += 1
            try:
                size_i = int(size or 0)
            except Exception:
                size_i = 0
            if (max_size_bytes is not None) and size_i > max_size_bytes:
                filtered_by_size += 1
                continue
            if not _wanted(path):
                filtered_by_include += 1
                continue
            interps = (ext_to_interpreters.get((ext or '').lower()) or [])
            if not interps:
                # Allow any interpreter with empty ext rule if present in map
                interps = ext_to_interpreters.get('', []) or []
            if not interps:
                filtered_no_interpreter += 1
                continue
            candidates.append((path, size_i, ext, interps))
            if len(candidates) >= max_files:
                break

        processed, errors = [], []

        def _run_interp(interp, target_path: str, content_bytes: bytes):
            # Prefer interpret_bytes/text when available, fallback to temp-file + interpret(path)
            try:
                if hasattr(interp, 'interpret_bytes'):
                    return interp.interpret_bytes(content_bytes, path_hint=target_path)
            except Exception:
                pass
            try:
                if hasattr(interp, 'interpret_text'):
                    return interp.interpret_text(content_bytes.decode('utf-8', errors='replace'), path_hint=target_path)
            except Exception:
                pass
            import tempfile
            from pathlib import Path as _P
            with tempfile.NamedTemporaryFile(delete=True) as tf:
                try:
                    tf.write(content_bytes)
                    tf.flush()
                except Exception:
                    pass
                return interp.interpret(_P(tf.name))

        for (fpath, fsize, ext, interps) in candidates:
            last_err = None
            try:
                if provider_id == 'rclone':
                    content = rclone.cat(fpath, max_bytes=max_size_bytes, timeout_sec=timeout_sec)  # type: ignore[attr-defined]
                else:
                    with open(fpath, 'rb') as f:
                        content = f.read(max_size_bytes or (16 * 1024 * 1024))
                success = False
                for interp in interps:
                    try:
                        result = _run_interp(interp, fpath, content) or {}
                        payload = {
                            'status': result.get('status', 'success'),
                            'data': result.get('data', result),
                            'interpreter_version': getattr(interp, 'version', '0.0.1'),
                        }
                        try:
                            from .core import path_index_sqlite as pix
                            conn_i = pix.connect(); pix.init_db(conn_i)
                            cur_i = conn_i.cursor()
                            cur_i.execute(
                                "UPDATE files SET interpreted_as=?, interpretation_json=? WHERE path=? AND type='file' AND scan_id=?",
                                (getattr(interp, 'id', None), _json.dumps(payload.get('data')), fpath, scan_id)
                            )
                            conn_i.commit(); conn_i.close()
                        except Exception:
                            pass
                        processed.append({'path': fpath, 'size': fsize, 'interpreter': getattr(interp, 'id', None), 'status': 'ok'})
                        success = True
                        break
                    except Exception as e:
                        last_err = str(e)
                        continue
                if not success:
                    errors.append({'path': fpath, 'error': last_err or 'no interpreter succeeded'})
            except Exception as e:
                errors.append({'path': fpath, 'error': str(e)})

        return jsonify({
            'status': 'ok',
            'scan_id': scan_id,
            'provider_id': provider_id,
            'processed_count': len(processed),
            'error_count': len(errors),
            'files_seen': files_seen,
            'filtered_by_size': filtered_by_size,
            'filtered_by_include': filtered_by_include,
            'filtered_no_interpreter': filtered_no_interpreter,
            'processed': processed[:100],
            'errors': errors[:50],
        }), 200

    @api.post('/scans/<scan_id>/reinterpret')
    def api_scan_reinterpret(scan_id):
        """
        Re-run interpreters for files in an existing scan and persist results into SQLite.
        - Operates only on local files (absolute paths). Remote canonical paths like "remote:sub/path" are skipped
          unless they are mounted locally and resolvable; this MVP does not stream remote bytes.
        - Honors current effective interpreter enablement.
        - Matching is by filename against each interpreter's globs (fnmatch), with normalization:
          patterns like ".csv" are treated as "*.csv". We also fall back to extension-based matching.
        Returns a summary with counts, including files_seen and files_matched.
        """
        try:
            from .core import path_index_sqlite as pix
            from .core import migrations as _migs
            import json as _json
            import fnmatch
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        # Resolve effective enabled interpreters
        reg = app.extensions['scidk']['registry']
        istate = app.extensions.get('scidk', {}).get('interpreters', {})
        eff_set = set(istate.get('effective_enabled') or [])
        if not eff_set:
            # Fallback to current registry defaults if snapshot missing
            eff_set = set([iid for iid in reg.by_id.keys() if getattr(reg.by_id[iid], 'default_enabled', True)])
        enabled_interpreters = [reg.by_id[i] for i in reg.by_id.keys() if i in eff_set]
        # Prepare minimal debug snapshot of interpreter patterns
        _dbg_interps = []
        try:
            for it in enabled_interpreters:
                try:
                    _dbg_interps.append({
                        'id': getattr(it, 'id', None),
                        'globs': list((getattr(it, 'globs', []) or [])),
                        'extensions': list((getattr(it, 'extensions', []) or [])),
                    })
                except Exception:
                    pass
        except Exception:
            _dbg_interps = []
        # Open DB
        conn = pix.connect()
        _migs.migrate(conn)
        updated = 0
        skipped_remote = 0
        not_found = 0
        errors = 0
        files_seen = 0
        files_matched = 0
        try:
            cur = conn.cursor()
            cur.execute("SELECT path, name, file_extension FROM files WHERE scan_id=? AND type='file'", (scan_id,))
            rows = cur.fetchall()
            for (path_val, name_val, ext_val) in rows:
                files_seen += 1
                try:
                    # Skip remote canonical paths (e.g., "remote:...")
                    if isinstance(path_val, str) and ':' in path_val and not path_val.startswith('/'):
                        skipped_remote += 1
                        continue
                    fpath = Path(path_val)
                    if not fpath.exists():
                        not_found += 1
                        continue
                    # Choose applicable interpreters using registry rule engine with extension fallback
                    ds = {
                        'path': path_val,
                        'extension': (ext_val or '').strip().lower(),
                        'name': name_val,
                    }
                    matched = reg.select_for_dataset(ds) or []
                    if not matched:
                        continue
                    files_matched += 1
                    # Run first matching interpreter (MVP). If multiple, prefer first.
                    interp = matched[0]
                    try:
                        result = interp.interpret(fpath)
                        payload = {
                            'status': result.get('status', 'success'),
                            'data': result.get('data', result),
                            'interpreter_version': getattr(interp, 'version', '0.0.1'),
                        }
                    except Exception as ie:
                        payload = {
                            'status': 'error',
                            'data': {'error': str(ie)},
                            'interpreter_version': getattr(interp, 'version', '0.0.1'),
                        }
                    # Persist into SQLite
                    cur.execute(
                        "UPDATE files SET interpreted_as = ?, interpretation_json = ? WHERE path = ? AND type = 'file' AND scan_id = ?",
                        (interp.id, _json.dumps(payload.get('data')), str(fpath), scan_id)
                    )
                    updated += (1 if cur.rowcount else 0)
                except Exception:
                    errors += 1
                    continue
            conn.commit()
        finally:
            try:
                conn.close()
            except Exception:
                pass
        # Minimal server-side log for diagnostics
        try:
            print(f"[reinterpret] scan={scan_id} updated={updated} skipped_remote={skipped_remote} not_found={not_found} errors={errors} files_seen={files_seen} files_matched={files_matched}")
        except Exception:
            pass
        # Minimal server-side log for diagnostics
        try:
            print(f"[reinterpret] scan={scan_id} updated={updated} skipped_remote={skipped_remote} not_found={not_found} errors={errors} files_seen={files_seen} files_matched={files_matched}")
            if files_matched == 0 and files_seen > 0:
                # Log first few file names and extensions plus interpreter patterns
                try:
                    print(f"[reinterpret] enabled_interpreters={_dbg_interps}")
                except Exception:
                    pass
        except Exception:
            pass
        return jsonify({'status': 'ok', 'updated': int(updated), 'skipped_remote': int(skipped_remote), 'not_found': int(not_found), 'errors': int(errors), 'files_seen': int(files_seen), 'files_matched': int(files_matched)}), 200

    @api.get('/index/search')
    def api_index_search():
        # Feature-flagged minimal search over the SQLite files table
        if not _ff_index:
            return jsonify({"error": "file index disabled"}), 404
        from .core import path_index_sqlite as pix
        q = (request.args.get('q') or '').strip()
        ext = (request.args.get('ext') or '').strip().lower() or None
        prefix = (request.args.get('prefix') or '').strip() or None
        scan_id = (request.args.get('scan_id') or '').strip() or None
        try:
            conn = pix.connect()
            pix.init_db(conn)
            cur = conn.cursor()
            clauses = []
            params = []
            if q:
                clauses.append('(name LIKE ? OR path LIKE ?)')
                like = f"%{q}%"
                params.extend([like, like])
            if ext:
                clauses.append('file_extension = ?')
                params.append(ext if ext.startswith('.') else ('.' + ext if ext else ext))
            if prefix:
                # Ensure trailing slash for folder-like prefixes unless remote root
                pfx = prefix if prefix.endswith(':') else (prefix.rstrip('/') + '/')
                clauses.append('path LIKE ?')
                params.append(f"{pfx}%")
            if scan_id:
                clauses.append('scan_id = ?')
                params.append(scan_id)
            where = (' WHERE ' + ' AND '.join(clauses)) if clauses else ''
            sql = f"SELECT path,name,size,file_extension,mime_type,scan_id FROM files{where} LIMIT 500"
            cur.execute(sql, params)
            rows = cur.fetchall()
            out = [
                {
                    'path': r[0],
                    'name': r[1],
                    'size': r[2],
                    'ext': r[3],
                    'mime': r[4],
                    'scan_id': r[5],
                } for r in rows
            ]
            conn.close()
            return jsonify({'results': out, 'count': len(out)}), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @api.get('/index/duplicates')
    def api_index_duplicates():
        if not _ff_index:
            return jsonify({"error": "file index disabled"}), 404
        from .core import path_index_sqlite as pix
        method = (request.args.get('method') or 'size_name').strip()
        scan_id = (request.args.get('scan_id') or '').strip() or None
        try:
            conn = pix.connect()
            pix.init_db(conn)
            cur = conn.cursor()
            where = ''
            params = []
            if scan_id:
                where = ' WHERE scan_id = ?'
                params = [scan_id]
            out = []
            if method == 'hash':
                cur.execute(f"""
                    SELECT hash, COUNT(*) as n, GROUP_CONCAT(path)
                    FROM files{where} 
                    WHERE hash IS NOT NULL AND hash <> ''
                    GROUP BY hash HAVING n > 1
                    ORDER BY n DESC LIMIT 200
                """, params)
                for h, n, paths in cur.fetchall():
                    out.append({'hash': h, 'count': n, 'paths': (paths or '').split(',')})
            else:
                cur.execute(f"""
                    SELECT name, size, COUNT(*) as n, GROUP_CONCAT(path)
                    FROM files{where}
                    WHERE size > 0 AND name <> ''
                    GROUP BY name, size HAVING n > 1
                    ORDER BY n DESC, size DESC LIMIT 200
                """, params)
                for name, size, n, paths in cur.fetchall():
                    out.append({'name': name, 'size': size, 'count': n, 'paths': (paths or '').split(',')})
            conn.close()
            return jsonify({'duplicates': out, 'count': len(out), 'method': method}), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # Lightweight status endpoint with progress counters
    @api.get('/scans/<scan_id>/status')
    def api_scan_status(scan_id):
        s = app.extensions['scidk'].get('scans', {}).get(scan_id)
        if not s:
            return jsonify({"error": "not found"}), 404
        # Derive simple status and counters
        started = s.get('started')
        ended = s.get('ended')
        status = 'complete' if ended else 'running'
        return jsonify({
            'id': s.get('id'),
            'status': status,
            'started': started,
            'ended': ended,
            'duration_sec': s.get('duration_sec'),
            'file_count': s.get('file_count'),
            'ingested_rows': s.get('ingested_rows', 0),
            'by_ext': s.get('by_ext', {}),
            'folder_count': s.get('folder_count'),
            'source': s.get('source'),
        }), 200

    @api.get('/scans/<scan_id>/fs')
    def api_scan_fs(scan_id):
        idx = _get_or_build_scan_index(scan_id)
        if not idx:
            return jsonify({'error': 'scan not found'}), 404
        from pathlib import Path as _P
        req_path = (request.args.get('path') or '').strip()
        folder_info = idx['folder_info']
        children_folders = idx['children_folders']
        children_files = idx['children_files']
        roots = idx['roots']
        # Virtual root listing when no path specified
        if not req_path:
            # Auto-enter the scan base folder for a stable, expected view
            s = app.extensions['scidk'].get('scans', {}).get(scan_id) or {}
            base_path = s.get('path') or ''
            if base_path:
                # Ensure base exists in folder_info for consistent naming
                try:
                    from .core.path_utils import parse_remote_path, parent_remote_path
                    binfo = parse_remote_path(base_path)
                    if base_path not in folder_info:
                        if binfo.get('is_remote'):
                            bname = (binfo.get('parts')[-1] if binfo.get('parts') else binfo.get('remote_name') or base_path)
                            bparent = parent_remote_path(base_path)
                        else:
                            _bp = _P(base_path)
                            bname = _bp.name or base_path
                            bparent = str(_bp.parent)
                        folder_info[base_path] = {'path': base_path, 'name': bname, 'parent': bparent}
                except Exception:
                    pass
                req_path = base_path
                # Build breadcrumb and children for the base path
                breadcrumb = [
                    {'name': '(scan base)', 'path': ''},
                    {'name': folder_info.get(req_path, {}).get('name', _P(req_path).name), 'path': req_path},
                ]
                sub_folders = [
                    {'name': folder_info.get(p, {}).get('name', _P(p).name), 'path': p, 'file_count': len(children_files.get(p, []))}
                    for p in children_folders.get(req_path, [])
                ]
                sub_folders.sort(key=lambda r: r['name'].lower())
                files = children_files.get(req_path, [])
                return jsonify({
                    'scan_id': scan_id,
                    'path': req_path,
                    'breadcrumb': breadcrumb,
                    'folders': sub_folders,
                    'files': files,
                    'roots': idx['roots'],
                    'folder_info': folder_info,
                    'children_folders': children_folders,
                    'children_files': children_files,
                }), 200
            # If no base_path, fall back to showing roots
            folders = [{'name': _P(p).name, 'path': p, 'file_count': len(children_files.get(p, []))} for p in roots]
            folders.sort(key=lambda r: r['name'].lower())
            breadcrumb = [{'name': '(scan roots)', 'path': ''}]
            return jsonify({'scan_id': scan_id, 'path': '', 'breadcrumb': breadcrumb, 'folders': folders, 'files': [], 'roots': roots, 'folder_info': folder_info, 'children_folders': children_folders, 'children_files': children_files}), 200
        # Validate path exists in snapshot
        if req_path not in folder_info:
            return jsonify({'error': 'folder not found in scan'}), 404
        # Breadcrumb from this scan’s perspective
        bc_chain = []
        cur = req_path
        while cur and cur in folder_info:
            bc_chain.append(cur)
            par = folder_info[cur].get('parent')
            if par == cur:
                break
            cur = par
        bc_chain.reverse()
        breadcrumb = [{'name': '(scan roots)', 'path': ''}] + [{'name': _P(p).name, 'path': p} for p in bc_chain]
        # Children
        sub_folders = [{'name': _P(p).name, 'path': p, 'file_count': len(children_files.get(p, []))} for p in children_folders.get(req_path, [])]
        sub_folders.sort(key=lambda r: r['name'].lower())
        files = children_files.get(req_path, [])
        return jsonify({'scan_id': scan_id, 'path': req_path, 'breadcrumb': breadcrumb, 'folders': sub_folders, 'files': files, 'roots': roots, 'folder_info': folder_info, 'children_folders': children_folders, 'children_files': children_files}), 200

    # New: SQLite-index-backed browse for a given scan
    @api.get('/scans/<scan_id>/browse')
    def api_scan_browse(scan_id):
        """Browse direct children from the SQLite index for a scan.
        Delegates to FSIndexService.browse_children.
        Query params:
          - path (optional): parent folder; defaults to scan base path
          - page_size (optional, default 100)
          - next_page_token (optional)
          - extension / ext (optional)
          - type (optional)
        """
        from .services.fs_index_service import FSIndexService
        svc = FSIndexService(app)
        req_path = (request.args.get('path') or '').strip()
        # page_size
        try:
            page_size = int(request.args.get('page_size') or 100)
        except Exception:
            page_size = 100
        token = (request.args.get('next_page_token') or '').strip()
        filters = {
            'extension': (request.args.get('extension') or request.args.get('ext') or '').strip().lower(),
            'type': (request.args.get('type') or '').strip().lower(),
        }
        return svc.browse_children(scan_id, req_path, page_size, token, filters)

    @api.post('/ro-crates/referenced')
    def api_ro_crates_referenced():
        """Create a referenced RO-Crate from dataset_ids and/or explicit files.
        Environment flags:
          - SCIDK_ENABLE_ROCRATE_REFERENCED: if not truthy, returns 404.
          - SCIDK_ROCRATE_DIR: base directory to store crates (default: ~/.scidk/crates).
        Payload (JSON): { dataset_ids?: [str], files?: [obj], title?: str }
        Returns: { status: 'ok', crate_id: str, path: str }
        """
        # Feature gate
        flag = str(os.environ.get('SCIDK_ENABLE_ROCRATE_REFERENCED', '')).strip().lower()
        if flag not in ('1', 'true', 'yes', 'on', 'enabled'):  # disabled by default
            return jsonify({'error': 'not found'}), 404

        data = request.get_json(force=True, silent=True) or {}
        dataset_ids = data.get('dataset_ids') or []
        files = data.get('files') or []
        title = (data.get('title') or 'Referenced RO-Crate').strip() or 'Referenced RO-Crate'

        import time as _t, hashlib as _h, json as _json
        now = _t.time()
        crate_id = _h.sha1(f"{title}|{now}".encode()).hexdigest()[:12]
        base_dir = os.environ.get('SCIDK_ROCRATE_DIR') or os.path.expanduser('~/.scidk/crates')
        out_dir = os.path.join(base_dir, crate_id)
        try:
            os.makedirs(out_dir, exist_ok=True)
        except Exception as e:
            return jsonify({"status": "error", "error": f"could not create crate dir: {e}"}), 500

        # Gather items from dataset_ids and/or files
        items = []
        try:
            g = app.extensions['scidk']['graph']
        except Exception:
            g = None
        if dataset_ids and g is not None:
            ds_map = getattr(g, 'datasets', {})
            for did in dataset_ids:
                d = ds_map.get(did)
                if not d:
                    continue
                items.append({
                    'path': d.get('path'),
                    'name': d.get('filename') or Path(d.get('path') or '').name,
                    'size': int(d.get('size_bytes') or 0),
                    'mime_type': d.get('mime_type'),
                    'modified_time': float(d.get('modified') or 0.0),
                    'checksum': d.get('checksum'),
                })
        for f in files:
            items.append({
                'path': f.get('path') or f.get('url') or f.get('contentUrl'),
                'name': f.get('name'),
                'size': f.get('size') or f.get('size_bytes') or 0,
                'mime_type': f.get('mime') or f.get('mime_type'),
                'modified_time': f.get('modified') or f.get('modified_time') or 0.0,
                'checksum': f.get('checksum'),
            })

        def to_rclone_url(p: Optional[str]) -> Optional[str]:
            if not p or not isinstance(p, str):
                return None
            if '://' in p:
                return p
            if ':' in p:
                remote, rest = p.split(':', 1)
                rest = (rest or '').lstrip('/')
                return f"rclone://{remote}/{rest}" if rest else f"rclone://{remote}/"
            try:
                return f"file://{str(Path(p).resolve())}"
            except Exception:
                return f"file://{p}"

        graph = []
        graph.append({
            "@id": "ro-crate-metadata.json",
            "@type": "CreativeWork",
            "about": {"@id": "./"}
        })
        has_parts = []
        file_nodes = []
        import datetime as _dt
        for it in items:
            url = to_rclone_url(it.get('path'))
            if not url:
                continue
            has_parts.append({"@id": url})
            node = {"@id": url, "@type": "File", "contentUrl": url}
            if it.get('name'):
                node['name'] = it.get('name')
            try:
                node['contentSize'] = int(it.get('size') or 0)
            except Exception:
                pass
            if it.get('mime_type'):
                node['encodingFormat'] = it.get('mime_type')
            try:
                mt = float(it.get('modified_time') or 0.0)
                if mt:
                    node['dateModified'] = _dt.datetime.utcfromtimestamp(mt).isoformat() + 'Z'
            except Exception:
                pass
            if it.get('checksum'):
                node['checksum'] = it.get('checksum')
            file_nodes.append(node)
        root = {"@id": "./", "@type": "Dataset", "name": title, "hasPart": has_parts}
        graph.append(root)
        graph.extend(file_nodes)
        ro = {"@context": "https://w3id.org/ro/crate/1.1/context", "@graph": graph}
        try:
            with open(os.path.join(out_dir, 'ro-crate-metadata.json'), 'w', encoding='utf-8') as fh:
                _json.dump(ro, fh, indent=2)
        except Exception as e:
            return jsonify({"status": "error", "error": f"could not write ro-crate: {e}"}), 500
        return jsonify({"status": "ok", "crate_id": crate_id, "path": out_dir}), 200

        
    @api.post('/ro-crates/<crate_id>/export')
    def api_ro_crates_export(crate_id):
        """Export a referenced RO-Crate directory as a ZIP (metadata-only).
        Query param: target=zip (required)
        Errors:
          - 400 for missing/invalid target or inaccessible path
          - 404 when crateId directory does not exist
        """
        target = (request.args.get('target') or '').strip().lower()
        if target not in ('zip', 'application/zip', 'zipfile'):
            return jsonify({'error': 'invalid or missing target; expected target=zip'}), 400
        base_dir = os.environ.get('SCIDK_ROCRATE_DIR') or os.path.expanduser('~/.scidk/crates')
        crate_dir = os.path.join(base_dir, crate_id)
        try:
            from pathlib import Path as _P
            p = _P(crate_dir)
            if not p.exists():
                return jsonify({'error': 'crate not found'}), 404
            if not p.is_dir():
                return jsonify({'error': 'crate path is not a directory'}), 400
            # Build a ZIP of the crate directory (metadata files only live here in referenced mode)
            import io, zipfile
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
                # include all files under the crate directory (non-recursive safe walk)
                for root, dirs, files in os.walk(str(p)):
                    rel_root = os.path.relpath(root, str(p))
                    for fname in files:
                        fpath = os.path.join(root, fname)
                        arcname = fname if rel_root == '.' else os.path.join(rel_root, fname)
                        try:
                            zf.write(fpath, arcname)
                        except Exception:
                            # skip unreadable files but continue building the archive
                            continue
                    # Only shallow by default; but if metadata structure has subdirs, we include them
                    # so we do not break out of walk intentionally.
            buf.seek(0)
            from flask import send_file as _send_file
            dl_name = f"{crate_id}.zip"
            return _send_file(buf, mimetype='application/zip', as_attachment=True, download_name=dl_name)
        except PermissionError:
            return jsonify({'error': 'inaccessible crate path'}), 400
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @api.get('/scans/<scan_id>/commit_preview')
    def api_scan_commit_preview(scan_id):
        """Dev-only: preview rows/folders the app would commit for this scan (index mode builder)."""
        scans = app.extensions['scidk'].setdefault('scans', {})
        s = scans.get(scan_id)
        if not s:
            # Try to load minimal scan from SQLite to allow preview
            try:
                from .core import path_index_sqlite as pix
                from .core import migrations as _migs
                import json as _json
                conn = pix.connect()
                try:
                    _migs.migrate(conn)
                    cur = conn.cursor()
                    cur.execute("SELECT id, root, started, completed, status, extra_json FROM scans WHERE id = ?", (scan_id,))
                    row = cur.fetchone()
                    if row:
                        sid, root, started, completed, status, extra = row
                        ex = {}
                        try:
                            if extra:
                                ex = _json.loads(extra)
                        except Exception:
                            ex = {}
                        s = {'id': sid, 'path': root, 'started': started, 'ended': completed}
                        scans[scan_id] = s
                    else:
                        return jsonify({"error": "scan not found"}), 404
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            except Exception:
                return jsonify({"error": "scan not found"}), 404
        try:
            from .core.commit_rows_from_index import build_rows_for_scan_from_index
            rows, folder_rows = build_rows_for_scan_from_index(scan_id, s, include_hierarchy=True)
            return jsonify({
                'scan_id': scan_id,
                'base': s.get('path'),
                'counts': {'files': len(rows), 'folders': len(folder_rows)},
                'sample_files': [{"path": r.get("path"), "folder": r.get("folder")} for r in (rows[:20] if rows else [])],
                'sample_folders': [{"path": r.get("path"), "parent": r.get("parent")} for r in (folder_rows[:20] if folder_rows else [])]
            }), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @api.get('/scans/<scan_id>/hierarchy')
    def api_scan_hierarchy(scan_id):
        scans = app.extensions['scidk'].setdefault('scans', {})
        s = scans.get(scan_id)
        if not s:
            # Attempt to reconstruct minimal scan from SQLite so the hierarchy can be built from index
            try:
                from .core import path_index_sqlite as pix
                from .core import migrations as _migs
                import json as _json
                conn = pix.connect()
                try:
                    _migs.migrate(conn)
                    cur = conn.cursor()
                    cur.execute("SELECT id, root, started, completed, status, extra_json FROM scans WHERE id = ?", (scan_id,))
                    row = cur.fetchone()
                    if row:
                        sid, root, started, completed, status, extra = row
                        extra_obj = {}
                        try:
                            if extra:
                                extra_obj = _json.loads(extra)
                        except Exception:
                            extra_obj = {}
                        s = {
                            'id': sid,
                            'path': root,
                            'started': started,
                            'ended': completed,
                            'provider_id': (extra_obj or {}).get('provider_id') or 'local_fs',
                            'host_type': (extra_obj or {}).get('host_type'),
                            'host_id': (extra_obj or {}).get('host_id'),
                            'root_id': (extra_obj or {}).get('root_id') or '/',
                            'root_label': (extra_obj or {}).get('root_label'),
                        }
                        scans[scan_id] = s
                    else:
                        return jsonify({"error": "scan not found"}), 404
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            except Exception:
                return jsonify({"error": "scan not found"}), 404
        try:
            # Prefer index path when feature enabled
            use_index = (os.environ.get('SCIDK_COMMIT_FROM_INDEX') or '').strip().lower() in ('1','true','yes','y','on')
            rows = []
            folder_rows = []
            if use_index:
                from .core import path_index_sqlite as pix
                conn = pix.connect()
                try:
                    pix.init_db(conn)
                    cur = conn.cursor()
                    cur.execute(
                        "SELECT path, parent_path, name, depth, type, size, modified_time, file_extension, mime_type FROM files WHERE scan_id = ?",
                        (scan_id,)
                    )
                    items = cur.fetchall()
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
                def _parent(path: str) -> str:
                    try:
                        from .core.path_utils import parse_remote_path, parent_remote_path
                        info = parse_remote_path(path)
                        if info.get('is_remote'):
                            return parent_remote_path(path)
                    except Exception:
                        pass
                    try:
                        from pathlib import Path as _P
                        return str(_P(path).parent)
                    except Exception:
                        return ''
                def _name(path: str) -> str:
                    try:
                        from .core.path_utils import parse_remote_path
                        info = parse_remote_path(path)
                        if info.get('is_remote'):
                            parts = info.get('parts') or []
                            return (info.get('remote_name') or '') if not parts else parts[-1]
                    except Exception:
                        pass
                    try:
                        from pathlib import Path as _P
                        return _P(path).name
                    except Exception:
                        return path
                folders_seen = set()
                for (p, parent, name, depth, typ, size, mtime, ext, mime) in items:
                    if typ == 'folder':
                        if p in folders_seen:
                            continue
                        folders_seen.add(p)
                        par = (parent or '').strip() or _parent(p)
                        folder_rows.append({
                            'path': p,
                            'name': name or _name(p),
                            'parent': par,
                            'parent_name': _name(par),
                        })
                    else:
                        par = (parent or '').strip() or _parent(p)
                        rows.append({
                            'path': p,
                            'filename': name or _name(p),
                            'extension': ext or '',
                            'size_bytes': int(size or 0),
                            'created': 0.0,
                            'modified': float(mtime or 0.0),
                            'mime_type': mime,
                            'folder': par,
                        })
            else:
                g = app.extensions['scidk']['graph']
                ds_map = getattr(g, 'datasets', {})
                rows, folder_rows = build_commit_rows(s, ds_map)
            # Debug: first 10 examples from hierarchy view
            try:
                app.logger.debug({
                    "event": "hierarchy_preview_debug",
                    "sample_files": [{"path": r.get("path"), "folder": r.get("folder")} for r in (rows[:10] if rows else [])],
                    "sample_folders": [{"path": r.get("path"), "parent": r.get("parent")} for r in (folder_rows[:10] if folder_rows else [])]
                })
            except Exception:
                pass
            # Ensure complete hierarchy
            try:
                from .core.folder_hierarchy import build_complete_folder_hierarchy
                folder_rows = build_complete_folder_hierarchy(rows, folder_rows, s)
            except Exception:
                pass
            # Build relationships list
            relationships = []
            seen = set()
            for fr in folder_rows:
                p = fr.get('parent') or ''
                c = fr.get('path') or ''
                if p and c and (p, c) not in seen and p != c:
                    relationships.append([p, c])
                    seen.add((p, c))
            return jsonify({
                'folders': folder_rows,
                'relationships': relationships,
                'count_folders': len({fr.get('path') for fr in folder_rows}),
                'count_relationships': len(relationships)
            }), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @api.post('/scans/<scan_id>/commit')
    def api_scan_commit(scan_id):
        scans = app.extensions['scidk'].setdefault('scans', {})
        s = scans.get(scan_id)
        if not s:
            return jsonify({"status": "error", "error": "scan not found"}), 404
        try:
            use_index = (os.environ.get('SCIDK_COMMIT_FROM_INDEX') or '').strip().lower() in ('1','true','yes','y','on')
            g = app.extensions['scidk']['graph']
            if use_index:
                # Build rows directly from SQLite index for this scan using shared builder
                from .core.commit_rows_from_index import build_rows_for_scan_from_index
                rows, folder_rows = build_rows_for_scan_from_index(scan_id, s, include_hierarchy=True)
                # Debug: first 10 mappings from index
                try:
                    app.logger.debug({
                        "event": "index_mapping_debug",
                        "sample_files": [{"path": r.get("path"), "folder": r.get("folder")} for r in (rows[:10] if rows else [])],
                        "sample_folders": [{"path": r.get("path"), "parent": r.get("parent")} for r in (folder_rows[:10] if folder_rows else [])]
                    })
                except Exception:
                    pass
                # For compatibility with schema endpoint, still add a Scan node to in-memory graph
                try:
                    g.commit_scan(s)
                except Exception:
                    pass
                # Totals for reporting
                total = len([r for r in rows if r.get('path')])
                present = total  # index-driven commit considers all indexed files as present
                missing = 0
                s['committed'] = True
                import time as _t
                s['committed_at'] = _t.time()
            else:
                # Legacy in-memory path
                checksums = s.get('checksums') or []
                total = len(checksums)
                # How many of these files are present in the current graph
                present = sum(1 for ch in checksums if ch in getattr(g, 'datasets', {}))
                missing = total - present
                # Commit into graph (idempotent add of Scan + edges)
                g.commit_scan(s)
                ds_map = getattr(g, 'datasets', {})
                rows, folder_rows = build_commit_rows(s, ds_map)
            import time as _t
            s['committed'] = True
            s['committed_at'] = _t.time()

            # Attempt Neo4j write if configuration is present (do not rely solely on connected flag)
            neo_state = app.extensions['scidk'].get('neo4j_state', {})
            neo_attempted = False
            neo_written = 0
            neo_error = None
            db_verified = None
            db_files = 0
            db_folders = 0
            uri, user, pwd, database, auth_mode = _get_neo4j_params()
            if uri and ((auth_mode == 'none') or (user and pwd)):
                neo_attempted = True
                try:
                    # If committing from index, use the rows built above from SQLite; otherwise build from legacy datasets
                    if use_index:
                        def _prog(ev, payload):
                            try:
                                app.logger.info(f"neo4j {ev}: {payload}")
                            except Exception:
                                pass
                        if app.config.get('TESTING'):
                            result = commit_to_neo4j(rows, folder_rows, s, (uri, user, pwd, database, auth_mode))
                        else:
                            result = commit_to_neo4j_batched(
                                rows=rows,
                                folder_rows=folder_rows,
                                scan=s,
                                neo4j_params=(uri, user, pwd, database, auth_mode),
                                file_batch_size=int(os.environ.get('SCIDK_NEO4J_FILE_BATCH') or 5000),
                                folder_batch_size=int(os.environ.get('SCIDK_NEO4J_FOLDER_BATCH') or 5000),
                                max_retries=2,
                                on_progress=_prog,
                            )
                    else:
                        ds_map = getattr(g, 'datasets', {})
                        rows, folder_rows = build_commit_rows(s, ds_map)
                        result = commit_to_neo4j_batched(
                            rows=rows,
                            folder_rows=folder_rows,
                            scan=s,
                            neo4j_params=(uri, user, pwd, database, auth_mode),
                            file_batch_size=int(os.environ.get('SCIDK_NEO4J_FILE_BATCH') or 5000),
                            folder_batch_size=int(os.environ.get('SCIDK_NEO4J_FOLDER_BATCH') or 5000),
                            max_retries=2,
                            on_progress=_neo4j_progress_default,
                        )
                    neo_written = int(result.get('written_files', 0)) + int(result.get('written_folders', 0))
                    # Capture DB verification if provided
                    if 'db_verified' in result:
                        db_verified = bool(result.get('db_verified'))
                        db_files = int(result.get('db_files') or 0)
                        db_folders = int(result.get('db_folders') or 0)
                    if result.get('error'):
                        raise Exception(result['error'])
                    # Update state on success
                    neo_state['connected'] = True
                    neo_state['last_error'] = None
                except Exception as ne:
                    neo_error = str(ne)
                    neo_state['connected'] = False
                    neo_state['last_error'] = neo_error
                # Note: even on Neo4j error we still return 200 for in-memory commit but include error details

            # In our in-memory model, linked_edges_added ~= present (File→Scan per matched dataset)
            payload = {
                "status": "ok",
                "committed": True,
                "scan_id": scan_id,
                "files_in_scan": total,
                "matched_in_graph": present,
                "missing_from_graph": max(0, missing),
                "linked_edges_added": present,
                "neo4j_attempted": neo_attempted,
                "neo4j_written_files": neo_written,
                "commit_mode": ("index" if use_index else "legacy"),
                # DB verification results (if commit attempted)
                "neo4j_db_verified": db_verified,
                "neo4j_db_files": db_files,
                "neo4j_db_folders": db_folders,
            }
            # Diagnostics about commit path and row counts
            if use_index:
                try:
                    payload["neo4j_rows_files"] = len(rows)
                    payload["neo4j_rows_folders"] = len(folder_rows)
                except Exception:
                    pass
            if neo_error:
                payload["neo4j_error"] = neo_error
            # Add user-facing warnings
            if total == 0:
                payload["warning"] = "This scan has 0 files; nothing was linked."
            elif present == 0:
                payload["warning"] = "None of the scanned files are currently present in the graph; verify you scanned in this session or refresh."
            # Neo4j-specific warning when verification fails but no explicit error was raised
            if neo_attempted and (db_verified is not None) and (not db_verified) and (not neo_error):
                payload["neo4j_warning"] = (
                    "Neo4j post-commit verification found 0 SCANNED_IN edges for this scan. "
                    "Verify: URI, credentials or set NEO4J_AUTH=none for no-auth, and database name. "
                    "Also ensure the scan has files present in this session's graph."
                )
            try:
                from .services.metrics import inc_counter
                # Consider files written as "rows" proxy for MVP
                inc_counter(app, 'rows_ingested_total', int(payload.get('neo4j_written_files') or 0))
            except Exception:
                pass
            return jsonify(payload), 200
        except Exception as e:
            return jsonify({"status": "error", "error": "commit failed", "error_detail": str(e)}), 500

    @api.delete('/scans/<scan_id>')
    def api_scan_delete(scan_id):
        scans = app.extensions['scidk'].setdefault('scans', {})
        existed = scan_id in scans
        # Remove from graph first
        app.extensions['scidk']['graph'].delete_scan(scan_id)
        if existed:
            del scans[scan_id]
        return jsonify({"status": "ok", "deleted": True, "scan_id": scan_id, "existed": existed}), 200

    @api.get('/graph/schema')
    def api_graph_schema():
        try:
            limit = int(request.args.get('limit') or 500)
        except Exception:
            limit = 500
        data = app.extensions['scidk']['graph'].schema_triples(limit=limit)
        return jsonify(data), 200

    @api.get('/graph/schema.csv')
    def api_graph_schema_csv():
        # Build a simple CSV with two sections: NodeLabels and RelationshipTypes
        g = app.extensions['scidk']['graph']
        triples = g.schema_triples(limit=int(request.args.get('limit') or 0 or 0))
        # If limit is 0 treat as no limit
        if (request.args.get('limit') or '').strip() == '0':
            triples = g.schema_triples(limit=0)
        nodes = triples.get('nodes', [])
        edges = triples.get('edges', [])
        lines = []
        lines.append('NodeLabels')
        lines.append('label,count')
        for n in nodes:
            lines.append(f"{n.get('label','')},{n.get('count',0)}")
        lines.append('')
        lines.append('RelationshipTypes')
        lines.append('start_label,rel_type,end_label,count')
        for e in edges:
            lines.append(f"{e.get('start_label','')},{e.get('rel_type','')},{e.get('end_label','')},{e.get('count',0)}")
        csv_text = "\n".join(lines) + "\n"
        from flask import Response
        return Response(csv_text, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename="schema.csv"'})

    @api.get('/graph/subschema')
    def api_graph_subschema():
        # Filters: name (optional), labels (csv), rel_types (csv), limit (int)
        params = {k: (request.args.get(k) or '').strip() for k in ['name','labels','rel_types','limit']}
        # Named queries
        if params['name']:
            if params['name'].lower() == 'interpreted_as':
                params['rel_types'] = 'INTERPRETED_AS' if not params['rel_types'] else params['rel_types']
            # Future named queries can be added here
        # Parse filters
        labels = set([s for s in (params['labels'].split(',')) if s]) if params['labels'] else set()
        rel_types = set([s for s in (params['rel_types'].split(',')) if s]) if params['rel_types'] else set()
        try:
            limit = int(params['limit']) if params['limit'] else 500
        except Exception:
            limit = 500
        g = app.extensions['scidk']['graph']
        base = g.schema_triples(limit=0 if limit == 0 else 1000000)  # fetch all, filter then trim
        edges = base.get('edges', [])
        # Apply filters
        def edge_ok(e):
            if rel_types and e.get('rel_type') not in rel_types:
                return False
            if labels and (e.get('start_label') not in labels and e.get('end_label') not in labels):
                return False
            return True
        filtered_edges = [e for e in edges if edge_ok(e)]
        # Truncate if needed
        truncated = False
        if limit and limit > 0 and len(filtered_edges) > limit:
            filtered_edges = filtered_edges[:limit]
            truncated = True
        # Build nodes: start with base nodes filtered by labels (if any)
        base_nodes = {n['label']: n.get('count', 0) for n in base.get('nodes', [])}
        node_map = {}
        # Include from filtered edges endpoints
        for e in filtered_edges:
            for lab in [e.get('start_label'), e.get('end_label')]:
                if lab not in node_map:
                    node_map[lab] = base_nodes.get(lab, 1 if lab else 1)
        # If labels filter provided, ensure inclusion even if no edges
        for lab in labels:
            if lab not in node_map:
                node_map[lab] = base_nodes.get(lab, 0)
        out = {
            'nodes': [{'label': k, 'count': v} for k, v in node_map.items()],
            'edges': filtered_edges,
            'truncated': truncated,
        }
        return jsonify(out), 200

    @api.get('/graph/instances')
    def api_graph_instances():
        label = (request.args.get('label') or '').strip()
        if not label:
            return jsonify({"error": "missing label"}), 400
        rows = app.extensions['scidk']['graph'].list_instances(label)
        # preview only (cap to 100 rows)
        limit = 100
        if request.args.get('limit'):
            try:
                limit = int(request.args.get('limit'))
            except Exception:
                pass
        return jsonify({
            'label': label,
            'count': len(rows),
            'rows': rows[:max(0, limit)]
        }), 200

    @api.get('/graph/instances.csv')
    def api_graph_instances_csv():
        label = (request.args.get('label') or '').strip()
        if not label:
            return jsonify({"error": "missing label"}), 400
        rows = app.extensions['scidk']['graph'].list_instances(label)
        # Build CSV
        if not rows:
            headers = ['id']
        else:
            # union columns
            cols = set()
            for r in rows:
                cols.update(r.keys())
            headers = sorted(list(cols))
        import io, csv
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, '') for k in headers})
        from flask import Response
        return Response(buf.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename="instances_{label}.csv"'})

    @api.get('/graph/instances.xlsx')
    def api_graph_instances_xlsx():
        try:
            import openpyxl  # type: ignore
        except Exception:
            return jsonify({"error": "xlsx export requires openpyxl"}), 501
        label = (request.args.get('label') or '').strip()
        if not label:
            return jsonify({"error": "missing label"}), 400
        rows = app.extensions['scidk']['graph'].list_instances(label)
        # Determine headers
        if not rows:
            headers = ['id']
        else:
            cols = set()
            for r in rows:
                cols.update(r.keys())
            headers = sorted(list(cols))
        from openpyxl import Workbook  # type: ignore
        wb = Workbook()
        ws = wb.active
        ws.title = label or 'Sheet1'
        ws.append(headers)
        for r in rows:
            ws.append([r.get(k, '') for k in headers])
        import io
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        from flask import Response
        return Response(bio.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="instances_{label}.xlsx"'})

    @api.get('/graph/instances.pkl')
    def api_graph_instances_pickle():
        import pickle
        label = (request.args.get('label') or '').strip()
        if not label:
            return jsonify({"error": "missing label"}), 400
        rows = app.extensions['scidk']['graph'].list_instances(label)
        payload = pickle.dumps(rows, protocol=pickle.HIGHEST_PROTOCOL)
        from flask import Response
        return Response(payload, mimetype='application/octet-stream', headers={'Content-Disposition': f'attachment; filename="instances_{label}.pkl"'})

    @api.get('/graph/instances.arrow')
    def api_graph_instances_arrow():
        try:
            import pyarrow as pa  # type: ignore
            import pyarrow.ipc as pa_ipc  # type: ignore
        except Exception:
            return jsonify({"error": "arrow export requires pyarrow"}), 501
        label = (request.args.get('label') or '').strip()
        if not label:
            return jsonify({"error": "missing label"}), 400
        rows = app.extensions['scidk']['graph'].list_instances(label)
        # Normalize rows to a table (handle missing keys by union of columns)
        cols = set()
        for r in rows:
            cols.update(r.keys())
        cols = sorted(list(cols)) if rows else ['id']
        arrays = {c: [] for c in cols}
        for r in rows:
            for c in cols:
                arrays[c].append(r.get(c))
        table = pa.table({c: pa.array(arrays[c]) for c in cols})
        sink = pa.BufferOutputStream()
        with pa_ipc.new_stream(sink, table.schema) as writer:
            writer.write_table(table)
        buf = sink.getvalue()
        from flask import Response
        return Response(buf.to_pybytes(), mimetype='application/vnd.apache.arrow.stream', headers={'Content-Disposition': f'attachment; filename="instances_{label}.arrow"'})

    @api.get('/graph/schema.neo4j')
    def api_graph_schema_neo4j():
        # Cypher-only triple derivation from a Neo4j instance if configured
        uri, user, pwd, database, auth_mode = _get_neo4j_params()
        if not uri:
            return jsonify({"error": "neo4j not configured (set in Settings or env: NEO4J_URI, and NEO4J_USER/NEO4J_PASSWORD or NEO4J_AUTH=none)"}), 501
        try:
            from neo4j import GraphDatabase  # type: ignore
        except Exception:
            return jsonify({"error": "neo4j driver not installed"}), 501
        try:
            driver = None
            try:
                driver = GraphDatabase.driver(uri, auth=None if auth_mode == 'none' else (user, pwd))
                with driver.session(database=database) as sess:
                    # Node label counts
                    q_nodes = "MATCH (n) WITH head(labels(n)) AS l, count(*) AS c RETURN l AS label, c ORDER BY c DESC"
                    nodes = [dict(record) for record in sess.run(q_nodes)]
                    # Unique triples counts
                    q_edges = (
                        "MATCH (s)-[r]->(t) "
                        "WITH head(labels(s)) AS sl, type(r) AS rt, head(labels(t)) AS tl, count(*) AS c "
                        "RETURN sl AS start_label, rt AS rel_type, tl AS end_label, c ORDER BY c DESC"
                    )
                    edges = [dict(record) for record in sess.run(q_edges)]
            finally:
                try:
                    if driver is not None:
                        driver.close()
                except Exception:
                    pass
            return jsonify({"nodes": nodes, "edges": edges, "truncated": False}), 200
        except Exception as e:
            return jsonify({"error": f"neo4j query failed: {str(e)}"}), 502

    @api.get('/graph/schema.apoc')
    def api_graph_schema_apoc():
        # APOC-based schema where available; fall back is not done here; use /graph/schema or /graph/schema.neo4j otherwise
        uri, user, pwd, database, auth_mode = _get_neo4j_params()
        if not uri:
            return jsonify({"error": "neo4j not configured (set in Settings or env: NEO4J_URI, and NEO4J_USER/NEO4J_PASSWORD or NEO4J_AUTH=none)"}), 501
        try:
            from neo4j import GraphDatabase  # type: ignore
        except Exception:
            return jsonify({"error": "neo4j driver not installed"}), 501
        try:
            driver = None
            try:
                driver = GraphDatabase.driver(uri, auth=None if auth_mode == 'none' else (user, pwd))
                with driver.session(database=database) as sess:
                    # Use apoc.meta.data() to derive nodes and edges
                    # Relationship triples aggregation
                    q_apoc = (
                        "CALL apoc.meta.data() YIELD label, other, elementType, type, count "
                        "WITH label, other, elementType, type, count "
                        "WHERE elementType = 'relationship' "
                        "RETURN label AS start_label, type AS rel_type, other AS end_label, count ORDER BY count DESC"
                    )
                    edges = [dict(record) for record in sess.run(q_apoc)]
                    # Node label counts via apoc (fallback to Cypher if needed)
                    q_nodes = "CALL apoc.meta.stats() YIELD labels RETURN [k IN keys(labels) | {label:k, count: labels[k]}] AS pairs"
                    rec = sess.run(q_nodes).single()
                    nodes = []
                    if rec and 'pairs' in rec:
                        for p in rec['pairs']:
                            nodes.append({'label': p['label'], 'count': p['count']})
            finally:
                try:
                    if driver is not None:
                        driver.close()
                except Exception:
                    pass
            return jsonify({"nodes": nodes, "edges": edges, "truncated": False}), 200
        except Exception as e:
            # If APOC procedures are missing or fail, inform the client
            return jsonify({"error": f"apoc schema failed: {str(e)}"}), 502

    @api.get('/health/graph')
    def api_health_graph():
        """Basic health for graph backend. In-memory is always OK; if Neo4j settings/env are provided, try a connection."""
        backend = os.environ.get('SCIDK_GRAPH_BACKEND', 'in_memory').lower() or 'in_memory'
        info = {
            'backend': backend,
            'in_memory_ok': True,
            'neo4j': {
                'configured': False,
                'connectable': False,
                'error': None,
            }
        }
        uri, user, pwd, database, auth_mode = _get_neo4j_params()
        if uri:
            info['neo4j']['configured'] = True
            try:
                from neo4j import GraphDatabase  # type: ignore
            except Exception as e:
                info['neo4j']['error'] = f"neo4j driver not installed: {e}"
                return jsonify(info), 200
            # Respect auth-failure backoff
            st = app.extensions['scidk'].setdefault('neo4j_state', {})
            import time as _t
            now = _t.time()
            next_after = float(st.get('next_connect_after') or 0)
            if next_after and now < next_after:
                info['neo4j']['error'] = f"backoff active; retry after {int(next_after-now)}s"
                return jsonify(info), 200
            try:
                driver = None
                try:
                    driver = GraphDatabase.driver(uri, auth=None if auth_mode == 'none' else (user, pwd))
                    with driver.session(database=database) as sess:
                        rec = sess.run("RETURN 1 AS ok").single()
                        if rec and rec.get('ok') == 1:
                            info['neo4j']['connectable'] = True
                finally:
                    try:
                        if driver is not None:
                            driver.close()
                    except Exception:
                        pass
            except Exception as e:
                info['neo4j']['error'] = str(e)
        return jsonify(info), 200

    @api.get('/health')
    def api_health():
        """Overall health focusing on SQLite availability and WAL mode."""
        from .core import path_index_sqlite as pix
        info = {
            'sqlite': {
                'path': None,
                'exists': False,
                'journal_mode': None,
                'select1': False,
                'error': None,
            }
        }
        try:
            dbp = pix._db_path()
            info['sqlite']['path'] = str(dbp)
            conn = pix.connect()
            try:
                mode = (conn.execute('PRAGMA journal_mode;').fetchone() or [''])[0]
                if isinstance(mode, str):
                    info['sqlite']['journal_mode'] = mode.lower()
                row = conn.execute('SELECT 1').fetchone()
                info['sqlite']['select1'] = bool(row and row[0] == 1)
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
            # After connecting, the DB file should exist on disk
            try:
                from pathlib import Path as _P
                info['sqlite']['exists'] = _P(info['sqlite']['path']).exists()
            except Exception:
                pass
        except Exception as e:
            info['sqlite']['error'] = str(e)
        # Always return 200 so UIs can render details; clients can decide on status
        return jsonify(info), 200

    @api.get('/metrics')
    def api_metrics():
        try:
            from .services.metrics import collect_metrics
            m = collect_metrics(app)
            return jsonify(m), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # Rclone interpretation settings (GET/POST)
    @api.get('/settings/rclone-interpret')
    def api_settings_rclone_interpret_get():
        return jsonify({
            'suggest_mount_threshold': int(app.config.get('rclone.interpret.suggest_mount_threshold', 400)),
            'max_files_per_batch': int(app.config.get('rclone.interpret.max_files_per_batch', 1000)),
        }), 200

    @api.post('/settings/rclone-interpret')
    def api_settings_rclone_interpret_set():
        data = request.get_json(force=True, silent=True) or {}
        try:
            suggest = int(data.get('suggest_mount_threshold')) if data.get('suggest_mount_threshold') not in (None, '') else None
        except Exception:
            suggest = None
        try:
            max_batch = int(data.get('max_files_per_batch')) if data.get('max_files_per_batch') not in (None, '') else None
        except Exception:
            max_batch = None
        # Validate and clamp
        if suggest is not None:
            suggest = max(0, int(suggest))
        if max_batch is not None:
            max_batch = min(max(100, int(max_batch)), 2000)
        # Persist best-effort
        try:
            from .core import path_index_sqlite as pix
            from .core import migrations as _migs
            conn = pix.connect()
            try:
                _migs.migrate(conn)
                cur = conn.cursor()
                if suggest is not None:
                    cur.execute("INSERT OR REPLACE INTO settings(key, value) VALUES(?, ?)", ('rclone.interpret.suggest_mount_threshold', str(suggest)))
                if max_batch is not None:
                    cur.execute("INSERT OR REPLACE INTO settings(key, value) VALUES(?, ?)", ('rclone.interpret.max_files_per_batch', str(max_batch)))
                conn.commit()
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass
        # Update in-memory config
        if suggest is not None:
            app.config['rclone.interpret.suggest_mount_threshold'] = int(suggest)
        if max_batch is not None:
            app.config['rclone.interpret.max_files_per_batch'] = int(max_batch)
        return jsonify({'ok': True, 'suggest_mount_threshold': int(app.config.get('rclone.interpret.suggest_mount_threshold', 400)), 'max_files_per_batch': int(app.config.get('rclone.interpret.max_files_per_batch', 1000))}), 200

    # Settings APIs for Neo4j configuration
    @api.get('/settings/neo4j')
    def api_settings_neo4j_get():
        cfg = app.extensions['scidk'].get('neo4j_config', {})
        state = app.extensions['scidk'].get('neo4j_state', {})
        # Do not return password
        out = {
            'uri': cfg.get('uri') or '',
            'user': cfg.get('user') or '',
            'database': cfg.get('database') or '',
            'connected': bool(state.get('connected')),
            'last_error': state.get('last_error'),
        }
        return jsonify(out), 200

    @api.post('/settings/neo4j')
    def api_settings_neo4j_set():
        data = request.get_json(force=True, silent=True) or {}
        cfg = app.extensions['scidk'].setdefault('neo4j_config', {})
        # Accept free text fields for uri, user, database
        for k in ['uri','user','database']:
            v = data.get(k)
            if v is not None:
                v = v.strip()
                cfg[k] = v if v else None
        # Password handling: only update if non-empty provided, unless clear_password=true
        if data.get('clear_password') is True:
            cfg['password'] = None
        else:
            if 'password' in data:
                v = data.get('password')
                if isinstance(v, str) and v.strip():
                    cfg['password'] = v.strip()
                # else: ignore empty password to avoid wiping stored secret
        # Reset state error on change
        st = app.extensions['scidk'].setdefault('neo4j_state', {})
        st['last_error'] = None
        return jsonify({'status':'ok'}), 200

    @api.post('/settings/neo4j/connect')
    def api_settings_neo4j_connect():
        uri, user, pwd, database, auth_mode = _get_neo4j_params()
        st = app.extensions['scidk'].setdefault('neo4j_state', {})
        st['connected'] = False
        st['last_error'] = None
        if not uri:
            st['last_error'] = 'Missing uri'
            return jsonify({'connected': False, 'error': st['last_error']}), 400
        try:
            from neo4j import GraphDatabase  # type: ignore
        except Exception as e:
            st['last_error'] = f'neo4j driver not installed: {e}'
            return jsonify({'connected': False, 'error': st['last_error']}), 501
        # Respect auth-failure backoff
        import time as _t
        now = _t.time()
        next_after = float(st.get('next_connect_after') or 0)
        if next_after and now < next_after:
            st['last_error'] = f"backoff active; retry after {int(next_after-now)}s"
            return jsonify({'connected': False, 'error': st['last_error']}), 429
        try:
            driver = None
            try:
                driver = GraphDatabase.driver(uri, auth=None if auth_mode == 'none' else (user, pwd))
                with driver.session(database=database) as sess:
                    rec = sess.run('RETURN 1 AS ok').single()
                    ok = bool(rec and rec.get('ok') == 1)
            finally:
                try:
                    if driver is not None:
                        driver.close()
                except Exception:
                    pass
            st['connected'] = ok
            # On success, clear backoff
            if ok:
                st['next_connect_after'] = 0
                st['last_error'] = None
            return jsonify({'connected': ok}), 200 if ok else 502
        except Exception as e:
            msg = str(e)
            st['last_error'] = msg
            st['connected'] = False
            # Apply backoff on auth errors
            try:
                emsg = msg.lower()
                if ('unauthorized' in emsg) or ('authentication' in emsg):
                    prev = float(st.get('next_connect_after') or 0)
                    base = 20.0
                    delay = base
                    if prev and now < prev:
                        rem = prev - now
                        delay = min(max(base*2, rem*2), 120.0)
                    st['next_connect_after'] = now + delay
            except Exception:
                pass
            return jsonify({'connected': False, 'error': st['last_error']}), 502

    @api.post('/settings/neo4j/disconnect')
    def api_settings_neo4j_disconnect():
        st = app.extensions['scidk'].setdefault('neo4j_state', {})
        st['connected'] = False
        return jsonify({'connected': False}), 200

    @api.get('/rocrate')
    def api_rocrate():
        """Return a minimal RO-Crate JSON-LD for a given directory (depth=1).
        Query: provider_id (default local_fs), root_id ('/'), path (directory path)
        Caps: at most 1000 immediate children; include meta.truncated when applied.
        """
        prov_id = (request.args.get('provider_id') or 'local_fs').strip() or 'local_fs'
        root_id = (request.args.get('root_id') or '/').strip() or '/'
        sel_path = (request.args.get('path') or '').strip() or root_id
        try:
            provs = app.extensions['scidk']['providers']
            prov = provs.get(prov_id)
            if not prov:
                return jsonify({'error': 'provider not available'}), 400
            from pathlib import Path as _P
            base = _P(root_id).resolve()
            target = _P(sel_path).resolve()
            # Ensure target resides under base (best-effort for local/mounted providers)
            try:
                target.relative_to(base)
            except Exception:
                # If not under base, fall back to base
                target = base
            if not target.exists() or not target.is_dir():
                return jsonify({'error': 'path not a directory'}), 400
            # Prepare root entity
            from datetime import datetime as _DT
            def iso(ts):
                try:
                    return _DT.fromtimestamp(float(ts)).isoformat()
                except Exception:
                    return None
            # Enumerate immediate children
            children = []
            total = 0
            LIMIT = 1000
            import mimetypes as _mt
            for child in target.iterdir():
                total += 1
                if len(children) >= LIMIT:
                    continue
                try:
                    st = child.stat()
                    is_dir = child.is_dir()
                    mime = None if is_dir else (_mt.guess_type(child.name)[0] or 'application/octet-stream')
                    children.append({
                        '@id': child.name + ('/' if is_dir else ''),
                        '@type': 'Dataset' if is_dir else 'File',
                        'name': child.name or str(child),
                        'contentSize': 0 if is_dir else int(st.st_size),
                        'dateModified': iso(st.st_mtime),
                        'encodingFormat': None if is_dir else mime,
                        'url': None if is_dir else (f"/api/files?provider_id={prov_id}&root_id={root_id}&path=" + str(child.resolve())),
                    })
                except Exception:
                    continue
            graph = [{
                '@id': './',
                '@type': 'Dataset',
                'name': target.name or str(target),
                'hasPart': [{'@id': c['@id']} for c in children],
            }] + children
            out = {
                '@context': 'https://w3id.org/ro/crate/1.1/context',
                '@graph': graph,
                'meta': {
                    'truncated': bool(total > len(children)),
                    'total_children': total,
                    'shown': len(children),
                }
            }
            return jsonify(out), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @api.get('/files')
    def api_files():
        """Stream a file's bytes with basic security and size limits.
        Query: provider_id, root_id, path
        Limits: default max 32MB unless SCIDK_FILE_MAX_BYTES is set.
        """
        prov_id = (request.args.get('provider_id') or 'local_fs').strip() or 'local_fs'
        root_id = (request.args.get('root_id') or '/').strip() or '/'
        file_path = (request.args.get('path') or '').strip()
        if not file_path:
            return jsonify({'error': 'missing path'}), 400
        try:
            from pathlib import Path as _P
            base = _P(root_id).resolve()
            target = _P(file_path).resolve()
            # Enforce that target is within base
            try:
                target.relative_to(base)
            except Exception:
                return jsonify({'error': 'path outside root'}), 400
            if not target.exists() or not target.is_file():
                return jsonify({'error': 'not a file'}), 400
            st = target.stat()
            max_bytes = int(os.environ.get('SCIDK_FILE_MAX_BYTES', '33554432'))  # 32MB
            if st.st_size > max_bytes:
                return jsonify({'error': 'file too large', 'limit': max_bytes, 'size': int(st.st_size)}), 413
            import mimetypes as _mt
            mime = _mt.guess_type(target.name)[0] or 'application/octet-stream'
            from flask import send_file as _send_file
            return _send_file(str(target), mimetype=mime, as_attachment=False, download_name=target.name)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @api.post('/research_objects')
    def api_research_objects_create():
        """Create or update a ResearchObject node for a directory path.
        Body JSON: { provider_id, root_id, path, name?, metadata? }
        Links to known files (datasets) under the directory and includes derived folder paths.
        """
        data = request.get_json(silent=True) or {}
        prov_id = (data.get('provider_id') or 'local_fs').strip() or 'local_fs'
        root_id = (data.get('root_id') or '/').strip() or '/'
        sel_path = (data.get('path') or '').strip() or root_id
        name = (data.get('name') or '').strip() or None
        extra_meta = data.get('metadata') or {}
        try:
            from pathlib import Path as _P
            base = _P(root_id).resolve()
            target = _P(sel_path).resolve()
            try:
                target.relative_to(base)
            except Exception:
                return jsonify({'error': 'path outside root'}), 400
            if not target.exists() or not target.is_dir():
                return jsonify({'error': 'path not a directory'}), 400
            # Build file checksum list by matching datasets whose path is under target
            g = app.extensions['scidk']['graph']
            file_checksums = []
            folder_paths = set()
            for ds in g.list_datasets():
                try:
                    p = _P(ds.get('path') or '').resolve()
                except Exception:
                    continue
                try:
                    p.relative_to(target)
                except Exception:
                    continue
                # Under target => include
                file_checksums.append(ds.get('checksum'))
                try:
                    folder_paths.add(str(p.parent))
                except Exception:
                    pass
            meta = {
                'name': name or (target.name or str(target)),
                'path': str(target),
                'provider_id': prov_id,
                'root_id': root_id,
            }
            # merge extra metadata
            try:
                if isinstance(extra_meta, dict):
                    meta.update({k: v for k, v in extra_meta.items() if k not in meta})
            except Exception:
                pass
            ro = g.upsert_research_object(meta, file_checksums, list(folder_paths))
            return jsonify({'status': 'ok', 'research_object': ro, 'file_links': len(file_checksums), 'folder_links': len(folder_paths)}), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @api.get('/research_objects')
    def api_research_objects_list():
        g = app.extensions['scidk']['graph']
        return jsonify(g.list_research_objects()), 200

    @api.get('/research_objects/<ro_id>')
    def api_research_objects_get(ro_id):
        g = app.extensions['scidk']['graph']
        ro = g.get_research_object(ro_id)
        if not ro:
            return jsonify({'error': 'not found'}), 404
        # expand lightweight links counts
        files = len(g.ro_files.get(ro_id, set()))
        folders = len(g.ro_folders.get(ro_id, set()))
        out = ro.copy()
        out.update({'file_count': files, 'folder_count': folders})
        return jsonify(out), 200

    # Selections & Annotations API (SQLite-backed)
    @api.post('/selections')
    def api_create_selection():
        payload = request.get_json(silent=True) or {}
        sel_id = (payload.get('id') or '').strip()
        name = (payload.get('name') or '').strip() or None
        import time as _t
        ts = _t.time()
        # If id not provided, derive short id from time
        if not sel_id:
            sel_id = hex(int(ts * 1000000))[2:]
        item = ann_db.create_selection(sel_id, name, ts)
        return jsonify(item), 201

    @api.post('/selections/<sel_id>/items')
    def api_add_selection_items(sel_id):
        payload = request.get_json(silent=True) or {}
        file_ids = payload.get('file_ids') or payload.get('files') or []
        if not isinstance(file_ids, list):
            return jsonify({'error': 'file_ids must be a list'}), 400
        import time as _t
        ts = _t.time()
        count = ann_db.add_selection_items(sel_id, [str(fid) for fid in file_ids], ts)
        return jsonify({'selection_id': sel_id, 'added': int(count)}), 200

    @api.post('/annotations')
    def api_create_annotation():
        payload = request.get_json(silent=True) or {}
        file_id = (payload.get('file_id') or '').strip()
        if not file_id:
            return jsonify({'error': 'file_id is required'}), 400
        kind = (payload.get('kind') or '').strip() or None
        label = (payload.get('label') or '').strip() or None
        note = payload.get('note')
        if isinstance(note, str):
            note = note
        elif note is None:
            note = None
        else:
            try:
                note = json.dumps(note)
            except Exception:
                note = str(note)
        data_json = payload.get('data_json')
        if not isinstance(data_json, (str, type(None))):
            try:
                data_json = json.dumps(data_json)
            except Exception:
                data_json = None
        import time as _t
        ts = _t.time()
        ann = ann_db.create_annotation(file_id, kind, label, note, data_json, ts)
        return jsonify(ann), 201

    @api.get('/annotations')
    def api_get_annotations():
        file_id = (request.args.get('file_id') or '').strip()
        if not file_id:
            return jsonify({'error': 'file_id query parameter is required'}), 400
        items = ann_db.list_annotations_by_file(file_id)
        return jsonify({'items': items, 'count': len(items)}), 200

    app.register_blueprint(api)

    # UI routes
    ui = Blueprint('ui', __name__)

    @ui.get('/')
    def index():
        datasets = app.extensions['scidk']['graph'].list_datasets()
        # Build lightweight summaries for the landing page
        by_ext = {}
        interp_types = set()
        for d in datasets:
            by_ext[d.get('extension') or ''] = by_ext.get(d.get('extension') or '', 0) + 1
            for k in (d.get('interpretations') or {}).keys():
                interp_types.add(k)
        schema_summary = app.extensions['scidk']['graph'].schema_summary()
        telemetry = app.extensions['scidk'].get('telemetry', {})
        directories = list(app.extensions['scidk'].get('directories', {}).values())
        directories.sort(key=lambda d: d.get('last_scanned') or 0, reverse=True)
        scans = list(app.extensions['scidk'].get('scans', {}).values())
        scans.sort(key=lambda s: s.get('ended') or s.get('started') or 0, reverse=True)
        # Add SQLite-backed scan_count for landing summary
        scan_count = None
        try:
            from .core import path_index_sqlite as pix
            conn = pix.connect()
            try:
                cur = conn.cursor()
                cur.execute("SELECT COUNT(1) FROM scans")
                row = cur.fetchone()
                if row:
                    scan_count = int(row[0])
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            scan_count = None
        return render_template('index.html', datasets=datasets, by_ext=by_ext, schema_summary=schema_summary, telemetry=telemetry, directories=directories, scans=scans, scan_count=scan_count)

    @ui.get('/chat')
    def chat():
        return render_template('chat.html')

    @ui.get('/map')
    def map_page():
        schema_summary = app.extensions['scidk']['graph'].schema_summary()
        return render_template('map.html', schema_summary=schema_summary)

    @ui.get('/datasets')
    def datasets():
        all_datasets = app.extensions['scidk']['graph'].list_datasets()
        scan_id = (request.args.get('scan_id') or '').strip()
        selected_scan = None
        if scan_id:
            selected_scan = app.extensions['scidk'].get('scans', {}).get(scan_id)
        if selected_scan:
            checks = set(selected_scan.get('checksums') or [])
            datasets = [d for d in all_datasets if d.get('checksum') in checks]
        else:
            datasets = all_datasets
        directories = list(app.extensions['scidk'].get('directories', {}).values())
        directories.sort(key=lambda d: d.get('last_scanned') or 0, reverse=True)
        recent_scans = list(app.extensions['scidk'].get('scans', {}).values())
        recent_scans.sort(key=lambda s: s.get('ended') or s.get('started') or 0, reverse=True)
        # Show only the most recent N scans for dropdown
        N = 20
        recent_scans = recent_scans[:N]
        # files viewer mode: allow query param override, else env, else classic
        files_viewer = (request.args.get('files_viewer') or os.environ.get('SCIDK_FILES_VIEWER') or 'classic').strip()
        return render_template('datasets.html', datasets=datasets, directories=directories, recent_scans=recent_scans, selected_scan=selected_scan, files_viewer=files_viewer)

    @ui.get('/datasets/<dataset_id>')
    def dataset_detail(dataset_id):
        item = app.extensions['scidk']['graph'].get_dataset(dataset_id)
        if not item:
            return render_template('dataset_detail.html', dataset=None), 404
        return render_template('dataset_detail.html', dataset=item)

    @ui.get('/workbook/<dataset_id>')
    def workbook_view(dataset_id):
        # Simple XLSX viewer: list sheets and preview first rows
        item = app.extensions['scidk']['graph'].get_dataset(dataset_id)
        if not item:
            return render_template('workbook.html', dataset=None, error="Dataset not found"), 404
        from pathlib import Path as _P
        file_path = _P(item['path'])
        if (item.get('extension') or '').lower() not in ['.xlsx', '.xlsm']:
            return render_template('workbook.html', dataset=item, error="Not an Excel workbook (.xlsx/.xlsm)"), 400
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
            sheetnames = wb.sheetnames
            previews = []
            max_rows = 20
            max_cols = 20
            for name in sheetnames:
                ws = wb[name]
                rows = []
                for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True), start=1):
                    rows.append(list(row))
                previews.append({'name': name, 'rows': rows})
            wb.close()
            return render_template('workbook.html', dataset=item, sheetnames=sheetnames, previews=previews, error=None)
        except Exception as e:
            return render_template('workbook.html', dataset=item, sheetnames=[], previews=[], error=str(e)), 500

    @ui.get('/plugins')
    def plugins():
        # Placeholder: no dynamic plugins yet. Show counts from registry for context.
        reg = app.extensions['scidk']['registry']
        ext_count = len(reg.by_extension)
        interp_count = len(reg.by_id)
        return render_template('plugins.html', ext_count=ext_count, interp_count=interp_count)

    @ui.get('/interpreters')
    def interpreters():
        # List registry mappings and selection rules
        reg = app.extensions['scidk']['registry']
        mappings = {ext: [getattr(i, 'id', 'unknown') for i in interps] for ext, interps in reg.by_extension.items()}
        rules = list(reg.rules.rules)
        return render_template('extensions.html', mappings=mappings, rules=rules)

    # Backward-compatible route
    @ui.get('/extensions')
    def extensions_legacy():
        return redirect(url_for('ui.interpreters'))

    @ui.get('/rocrate_view')
    def rocrate_view():
        # Lightweight wrapper page to preview RO-Crate JSON-LD and prepare for embedding Crate-O
        prov_id = (request.args.get('provider_id') or 'local_fs').strip() or 'local_fs'
        root_id = (request.args.get('root_id') or '/').strip() or '/'
        sel_path = (request.args.get('path') or root_id).strip()
        try:
            from urllib.parse import urlencode
            qs = urlencode({'provider_id': prov_id, 'root_id': root_id, 'path': sel_path})
        except Exception:
            qs = f"provider_id={prov_id}&root_id={root_id}&path={sel_path}"
        metadata_url = '/api/rocrate?' + qs
        embed_mode = (os.environ.get('SCIDK_ROCRATE_EMBED') or 'json').strip()
        return render_template('rocrate_view.html', metadata_url=metadata_url, embed_mode=embed_mode, prov_id=prov_id, root_id=root_id, path=sel_path)

    @ui.get('/settings')
    def settings():
        # Basic settings from environment and current in-memory sizes
        datasets = app.extensions['scidk']['graph'].list_datasets()
        reg = app.extensions['scidk']['registry']
        info = {
            'host': os.environ.get('SCIDK_HOST', '127.0.0.1'),
            'port': os.environ.get('SCIDK_PORT', '5000'),
            'debug': os.environ.get('SCIDK_DEBUG', '1'),
                        'feature_file_index': os.environ.get('SCIDK_FEATURE_FILE_INDEX', ''),
                        'hash_policy': os.environ.get('SCIDK_HASH_POLICY', 'auto'),
            'dataset_count': len(datasets),
            'interpreter_count': len(reg.by_id),
                        'channel': os.environ.get('SCIDK_CHANNEL', 'stable'),
                        'files_viewer': os.environ.get('SCIDK_FILES_VIEWER', ''),
                        'rclone_mounts': (os.environ.get('SCIDK_RCLONE_MOUNTS') or os.environ.get('SCIDK_FEATURE_RCLONE_MOUNTS') or ''),
                        'providers': os.environ.get('SCIDK_PROVIDERS', 'local_fs,mounted_fs'),
        }
        # Provide interpreter mappings and rules, and plugin summary counts for the Set page sections
        mappings = {ext: [getattr(i, 'id', 'unknown') for i in interps] for ext, interps in reg.by_extension.items()}
        rules = list(reg.rules.rules)
        ext_count = len(reg.by_extension)
        interp_count = len(reg.by_id)
        # Feature flag for rclone mounts UI
        rclone_mounts_feature = _feature_rclone_mounts()
        return render_template('settings.html', info=info, mappings=mappings, rules=rules, ext_count=ext_count, interp_count=interp_count, rclone_mounts_feature=rclone_mounts_feature)

    @ui.post('/scan')
    def ui_scan():
        path = request.form.get('path') or os.getcwd()
        recursive = request.form.get('recursive') == 'on'
        import time, hashlib
        # Pre-scan snapshot
        before = set(ds.get('checksum') for ds in app.extensions['scidk']['graph'].list_datasets())
        started = time.time()
        count = fs.scan_directory(Path(path), recursive=recursive)
        ended = time.time()
        duration = ended - started
        after = set(ds.get('checksum') for ds in app.extensions['scidk']['graph'].list_datasets())
        new_checksums = sorted(list(after - before))
        by_ext = {}
        ext_map = {ds.get('checksum'): ds.get('extension') or '' for ds in app.extensions['scidk']['graph'].list_datasets()}
        for ch in new_checksums:
            ext = ext_map.get(ch, '')
            by_ext[ext] = by_ext.get(ext, 0) + 1
        sid_src = f"{path}|{started}"
        scan_id = hashlib.sha1(sid_src.encode()).hexdigest()[:12]
        scan = {
            'id': scan_id,
            'path': str(path),
            'recursive': bool(recursive),
            'started': started,
            'ended': ended,
            'duration_sec': duration,
            'file_count': int(count),
            'checksums': new_checksums,
            'by_ext': by_ext,
            'source': getattr(fs, 'last_scan_source', 'python'),
            'errors': [],
        }
        scans = app.extensions['scidk'].setdefault('scans', {})
        scans[scan_id] = scan
        # Persist scan summary to SQLite (best-effort)
        try:
            from .core import path_index_sqlite as pix
            from .core import migrations as _migs
            import json as _json
            conn = pix.connect()
            try:
                _migs.migrate(conn)
                cur = conn.cursor()
                cur.execute(
                    "INSERT OR REPLACE INTO scans(id, root, started, completed, status, extra_json) VALUES(?,?,?,?,?,?)",
                    (
                        scan_id,
                        str(path),
                        float(started or 0.0),
                        float(ended or 0.0),
                        'completed',
                        _json.dumps({
                            'recursive': bool(recursive),
                            'duration_sec': duration,
                            'file_count': int(count),
                            'by_ext': by_ext,
                            'source': getattr(fs, 'last_scan_source', 'python'),
                            'checksums': new_checksums,
                            'committed': bool(scan.get('committed', False)),
                            'committed_at': scan.get('committed_at'),
                        })
                    )
                )
                conn.commit()
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass
        telem = app.extensions['scidk'].setdefault('telemetry', {})
        telem['last_scan'] = {
            'path': str(path),
            'recursive': bool(recursive),
            'scanned': int(count),
            'started': started,
            'ended': ended,
            'duration_sec': duration,
        }
        # Track scanned directories here as well
        dirs = app.extensions['scidk'].setdefault('directories', {})
        drec = dirs.setdefault(str(path), {'path': str(path), 'recursive': bool(recursive), 'scanned': 0, 'last_scanned': 0, 'scan_ids': []})
        drec.update({'recursive': bool(recursive), 'scanned': int(count), 'last_scanned': ended})
        drec.setdefault('scan_ids', []).append(scan_id)
        return redirect(url_for('ui.datasets', scan_id=scan_id))

    app.register_blueprint(ui)

    return app


def main():
    app = create_app()
    # Read host/port from env for convenience
    host = os.environ.get('SCIDK_HOST', '127.0.0.1')
    port = int(os.environ.get('SCIDK_PORT', '5000'))
    debug = os.environ.get('SCIDK_DEBUG', '1') == '1'
    app.run(host=host, port=port, debug=debug)


if __name__ == "__main__":
    main()
